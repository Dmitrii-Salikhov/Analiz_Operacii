# tests/test_session_unclassified.py
"""Тесты назначения существующей категории для «Не классифицировано»."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import pandas as pd
import pytest
import yaml

APP = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(APP))

from analyzers.category_registry import CategorySpec
from ui_flet.session import AppSession


def _minimal_app_dir(td: Path) -> Path:
    cfg = {
        "departments": {
            "main": "Оториноларингологическое отделение",
            "list": ["Оториноларингологическое отделение"],
        },
        "thresholds": {"pension_age": 60},
        "summary": {
            "category_rows": {"Аденотомия": 4, "Биопсия гортани ": 37},
            "totals_rows": {"children": 43, "patients": 45},
            "plan_categories": ["Аденотомия"],
            "emergency_categories": [],
            "sheet_names": {1: "Январь"},
            "year": 2026,
            "form_4001": {"enabled": True},
            "backup_keep": 5,
        },
        "surgery_categories": [
            {
                "category": "Аденотомия",
                "codes": ["A"],
                "line": "5.2",
                "histology": True,
                "name_keywords": ["аденотомия"],
                "group": "аденоиды",
            },
        ],
        "department_profiles": {
            "Оториноларингологическое отделение": {"summary_key": "lor"},
        },
    }
    (td / "config.yaml").write_text(yaml.safe_dump(cfg, allow_unicode=True), encoding="utf-8")
    (td / "VERSION").write_text("1.1.7", encoding="utf-8")
    return td


def test_assign_unclassified_category_manual():
    with tempfile.TemporaryDirectory() as tmp:
        app_dir = _minimal_app_dir(Path(tmp))
        session = AppSession(app_dir=app_dir)
        ops = pd.DataFrame(
            [
                {
                    "Дата": pd.Timestamp("2026-01-15"),
                    "КВС": 101,
                    "Код": "X",
                    "Услуга": "Тестовая операция",
                    "Категория": "Не классифицировано",
                    "Возраст": 30,
                }
            ]
        )
        session.store.ops = ops
        store_index = ops.index[0]

        session.assign_unclassified_category(store_index, "Аденотомия")

        row = session.store.ops.loc[store_index]
        assert row["Категория"] == "Аденотомия"
        assert bool(row["Ручная_категория"]) is True
        assert bool(row.get("Спор_ключей", False)) is False
        assert row.get("Строка_4001") == "5.2"
        assert bool(row.get("Гистология", False)) is True


def test_assign_unclassified_skips_missing_index():
    with tempfile.TemporaryDirectory() as tmp:
        app_dir = _minimal_app_dir(Path(tmp))
        session = AppSession(app_dir=app_dir)
        session.store.ops = pd.DataFrame(
            [{"Дата": pd.Timestamp("2026-01-01"), "Категория": "X", "КВС": 1}]
        )
        session.assign_unclassified_category(999, "Аденотомия")
        assert session.store.ops.iloc[0]["Категория"] == "X"


def _make_summary_xlsx(path: Path, *, one_blank_before_totals: bool = False) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Январь"
    ws["B37"] = "Биопсия гортани "
    ws["H37"] = "=SUM(C37:G37)"
    if one_blank_before_totals:
        ws["B39"] = "Всего операций"
        ws["B42"] = "Дети всего"
        ws["B43"] = "Человек"
    else:
        ws["B40"] = "Всего операций"
        ws["B43"] = "Дети всего"
        ws["B44"] = None
        ws["B45"] = "Человек"
    wb.create_sheet("Общая")
    wb.save(path)


def test_add_category_and_reclassify_inserts_excel_row():
    with tempfile.TemporaryDirectory() as tmp:
        app_dir = _minimal_app_dir(Path(tmp))
        summary = app_dir / "Операции сводная 2026.xlsx"
        _make_summary_xlsx(summary, one_blank_before_totals=True)

        session = AppSession(app_dir=app_dir)
        session.summary_path = str(summary)
        session.store.ops = pd.DataFrame(
            [
                {
                    "Дата": pd.Timestamp("2026-01-15"),
                    "КВС": 101,
                    "Код": "M01",
                    "Услуга": "мирингопластика",
                    "Категория": "Не классифицировано",
                    "Возраст": 30,
                }
            ]
        )

        spec = CategorySpec(
            name="Мирингопластика",
            codes=["M01"],
            name_keywords=["мирингопластика"],
            kind="plan",
            form_line="6",
            anchor_category="Биопсия гортани ",
        )
        res = session.add_category_and_reclassify(spec)
        assert res["added"] == "Мирингопластика"
        assert res["excel_row"] == 38
        assert res["excel"] is not None
        assert "Мирингопластика" in session.store.ops["Категория"].values


def test_add_category_and_reclassify_rolls_back_on_excel_error(monkeypatch):
    with tempfile.TemporaryDirectory() as tmp:
        app_dir = _minimal_app_dir(Path(tmp))
        summary = app_dir / "Операции сводная 2026.xlsx"
        _make_summary_xlsx(summary)

        session = AppSession(app_dir=app_dir)
        session.summary_path = str(summary)

        def _boom(*_a, **_k):
            raise RuntimeError("excel insert failed")

        monkeypatch.setattr(
            "ui_flet.session.add_category_row_to_summary",
            _boom,
        )

        spec = CategorySpec(
            name="Мирингопластика",
            codes=["M01"],
            name_keywords=["мирингопластика"],
            kind="plan",
            form_line="6",
            anchor_category="Биопсия гортани ",
        )
        with pytest.raises(RuntimeError):
            session.add_category_and_reclassify(spec)

        cats = session.summary_cfg.get("category_rows") or {}
        assert "Мирингопластика" not in cats


def test_add_category_and_reclassify_rolls_back_without_summary():
    with tempfile.TemporaryDirectory() as tmp:
        app_dir = _minimal_app_dir(Path(tmp))
        session = AppSession(app_dir=app_dir)
        session.summary_path = str(app_dir / "missing.xlsx")

        spec = CategorySpec(
            name="Мирингопластика",
            codes=["M01"],
            name_keywords=["мирингопластика"],
            kind="plan",
            form_line="6",
            anchor_category="Биопсия гортани ",
        )

        with pytest.raises(FileNotFoundError):
            session.add_category_and_reclassify(spec)

        cats = session.summary_cfg.get("category_rows") or {}
        assert "Мирингопластика" not in cats
