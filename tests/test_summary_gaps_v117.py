# tests/test_summary_gaps_v117.py
"""Тесты разделителей Excel (v1.1.7): 2 пустые строки перед «Всего операций»."""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
import yaml

APP = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(APP))

from analyzers.category_registry import CategorySpec, register_category
from analyzers.summary_layout import (
    add_category_row_to_summary,
    ensure_one_blank_before_totals,
    ensure_one_blank_between_labels,
)
from analyzers.summary_writer import SummaryWriter


def _minimal_summary_cfg():
    return {
        "year": 2026,
        "backup_keep": 5,
        "category_rows": {"Биопсия гортани ": 37, "Аденотомия": 4},
        "totals_rows": {"children": 43, "patients": 45},
        "plan_categories": ["Аденотомия", "Биопсия гортани "],
        "emergency_categories": [],
        "sheet_names": {1: "Январь"},
        "form_4001": {
            "enabled": True,
            "line_rows": {"6": 14},
            "cols": {"total": 14, "histology": 18},
        },
    }


def _make_reference_sheet(ws, *, totals_row: int = 40):
    """Структура как в эталонном бэкапе: bio 37, 2 пустые, totals 40."""
    ws.title = "Январь"
    ws["B37"] = "Биопсия гортани "
    ws["H37"] = "=SUM(C37:G37)"
    ws[f"B{totals_row}"] = "Всего операций"
    ws[f"C{totals_row}"] = "=SUM(C4:C37)"
    ws[f"B{totals_row + 1}"] = "Экстренно операций"
    ws[f"B{totals_row + 2}"] = "План операций"
    ws[f"B{totals_row + 3}"] = "Дети всего"
    ws[f"B{totals_row + 4}"] = None  # разделитель
    ws[f"B{totals_row + 5}"] = "Человек"


def test_two_blanks_before_totals_from_one_gap():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B37"] = "Биопсия гортани "
    ws["B39"] = "Всего операций"
    gap = ensure_one_blank_before_totals(ws, protected_last_category_row=37)
    assert gap["inserted"] == 1
    assert gap["delta"] == 1
    assert ws["B38"].value in (None, "")
    assert ws["B39"].value in (None, "")
    assert ws["B40"].value == "Всего операций"


def test_two_blanks_already_ok_no_change():
    wb = openpyxl.Workbook()
    ws = wb.active
    _make_reference_sheet(ws, totals_row=40)
    gap = ensure_one_blank_before_totals(ws, protected_last_category_row=37)
    assert gap["inserted"] == 0
    assert gap["deleted"] == 0
    assert gap["delta"] == 0
    assert ws["B38"].value in (None, "")
    assert ws["B39"].value in (None, "")
    assert ws["B40"].value == "Всего операций"


def test_trim_extra_blanks_before_totals():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B37"] = "Биопсия"
    ws["B41"] = "Всего операций"
    gap = ensure_one_blank_before_totals(ws)
    assert gap["deleted"] >= 1
    assert gap["delta"] < 0
    totals = None
    for r in range(1, 50):
        if ws.cell(r, 2).value == "Всего операций":
            totals = r
            break
    assert totals is not None
    assert ws.cell(totals - 1, 2).value in (None, "")
    assert ws.cell(totals - 2, 2).value in (None, "")


def test_ensure_one_blank_between_labels_trims_extras():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B43"] = "Дети всего"
    ws["B46"] = "Человек"
    gap = ensure_one_blank_between_labels(ws, top_label="Дети всего", bottom_label="Человек")
    assert gap["deleted"] >= 1
    assert ws["B44"].value in (None, "")
    assert ws["B45"].value == "Человек"


def test_ensure_one_blank_between_labels_already_ok():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B43"] = "Дети всего"
    ws["B44"] = None
    ws["B45"] = "Человек"
    gap = ensure_one_blank_between_labels(ws, top_label="Дети всего", bottom_label="Человек")
    assert gap["inserted"] == 0
    assert gap["deleted"] == 0
    assert ws["B44"].value in (None, "")


def test_insert_after_biopsy_keeps_two_blanks(tmp_path: Path):
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    _make_reference_sheet(ws, totals_row=40)
    wb.save(path)

    cfg_path = tmp_path / "config.yaml"
    cfg = {"summary": _minimal_summary_cfg(), "surgery_categories": []}
    with cfg_path.open("w", encoding="utf-8") as f:
        yaml.safe_dump(cfg, f, allow_unicode=True)

    _cfg, result = register_category(
        cfg_path,
        CategorySpec(
            name="Новая операция",
            codes=["Z01"],
            name_keywords=["новая"],
            kind="plan",
            form_line="6",
            anchor_category="Биопсия гортани ",
        ),
        config=cfg,
    )
    rep = add_category_row_to_summary(
        path,
        category_name=result.name,
        excel_row=result.excel_row,
        form_line="6",
        sheet_names={1: "Январь"},
        form_cfg=_cfg["summary"].get("form_4001") or {},
        kind="plan",
        backup=False,
    )
    assert rep["excel_row"] == 38

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Январь"]
    assert ws2["B37"].value == "Биопсия гортани "
    assert ws2["B38"].value == "Новая операция"
    assert ws2["B39"].value in (None, "")
    assert ws2["B40"].value in (None, "")
    assert ws2["B41"].value == "Всего операций"


def test_ensure_one_blank_between_labels_no_labels():
    wb = openpyxl.Workbook()
    ws = wb.active
    gap = ensure_one_blank_between_labels(ws, top_label="Дети всего", bottom_label="Человек")
    assert gap["inserted"] == 0
    assert gap["deleted"] == 0


def test_ensure_one_blank_before_totals_no_totals_label():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B37"] = "Биопсия"
    gap = ensure_one_blank_before_totals(ws)
    assert gap["inserted"] == 0
    assert gap["delta"] == 0


def test_summary_writer_normalizes_gaps(tmp_path: Path):
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Январь"
    ws["B37"] = "Биопсия гортани "
    ws["B39"] = "Всего операций"
    ws["B42"] = "Дети всего"
    ws["B43"] = "Человек"
    ws["C2"] = (pd.Timestamp("2026-01-01") - pd.Timestamp("1899-12-30")).days
    wb.save(path)

    cfg = _minimal_summary_cfg()
    cfg["category_rows"]["Биопсия гортани "] = 37
    writer = SummaryWriter(path, cfg, department="ЛОР", categories=[])
    report = writer.write(
        pd.DataFrame(),
        output_path=str(path),
        backup=False,
        write_weeks=False,
        write_form=False,
    )
    assert report["blank_delta"] >= 1

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Январь"]
    totals_row = None
    for r in range(1, 50):
        if ws2.cell(r, 2).value == "Всего операций":
            totals_row = r
            break
    assert totals_row is not None
    assert ws2.cell(totals_row - 1, 2).value in (None, "")
    assert ws2.cell(totals_row - 2, 2).value in (None, "")


def test_ensure_one_blank_invalid_protected_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B37"] = "Биопсия"
    ws["B39"] = "Всего операций"
    gap = ensure_one_blank_before_totals(ws, protected_last_category_row="bad")
    assert gap["inserted"] >= 1


def test_ensure_one_blank_no_category_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B10"] = "Всего операций"
    gap = ensure_one_blank_before_totals(ws)
    assert gap["delta"] == 0


def test_ensure_between_labels_stops_on_nonempty():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B43"] = "Дети всего"
    ws["B44"] = "заполнено"
    ws["B45"] = "ещё"
    ws["B46"] = "Человек"
    gap = ensure_one_blank_between_labels(ws, top_label="Дети всего", bottom_label="Человек")
    assert gap["deleted"] == 0


def test_ensure_between_labels_inverted_rows():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B45"] = "Дети всего"
    ws["B43"] = "Человек"
    gap = ensure_one_blank_between_labels(ws, top_label="Дети всего", bottom_label="Человек")
    assert gap["inserted"] == 0


def test_add_category_row_missing_file(tmp_path: Path):
    with pytest.raises(FileNotFoundError):
        add_category_row_to_summary(
            tmp_path / "missing.xlsx",
            category_name="X",
            excel_row=5,
            form_line="6",
            sheet_names={1: "Январь"},
        )


def test_add_category_row_with_overview_sheet(tmp_path: Path):
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Январь"
    ws["B37"] = "Биопсия гортани "
    ws["B40"] = "Всего операций"
    ws["B43"] = "Дети всего"
    ws["B44"] = None
    ws["B45"] = "Человек"
    ov = wb.create_sheet("Общая")
    ov["B40"] = "Всего операций"
    wb.save(path)

    rep = add_category_row_to_summary(
        path,
        category_name="Новая",
        excel_row=38,
        form_line="6",
        sheet_names={1: "Январь"},
        form_cfg={"enabled": True, "line_rows": {"6": 14}, "cols": {"total": 14}},
        kind="plan",
        backup=False,
    )
    assert "Общая" in rep.get("overview", [])
