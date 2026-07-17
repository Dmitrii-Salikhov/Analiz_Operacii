"""Дымовой тест: апрельский журнал → недели + форма 4001."""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest
import yaml

APP = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(APP))

from analyzers.form_4001 import compute_form_4001
from analyzers.io_utils import read_table
from analyzers.surgery import SurgeryAnalyzer, build_summary_tables
from analyzers.summary_writer import compute_month_weeks, read_sheet_weeks

JOURNAL = APP / "Отчет по выполненным операциям и операционным столам (17).xlsx"
SUMMARY = APP / "Операции сводная 2026.xlsx"


def _require_data():
    if not JOURNAL.exists() or not SUMMARY.exists():
        pytest.skip("локальные Excel-файлы журнала/сводной отсутствуют (CI без данных)")


def _config():
    with open(APP / "config.yaml", encoding="utf-8") as f:
        return yaml.safe_load(f)


def test_april_journal_extracts_ops():
    _require_data()
    cfg = _config()
    df = read_table(str(JOURNAL))
    analyzer = SurgeryAnalyzer(
        df,
        cfg["departments"]["main"],
        cfg["surgery_categories"],
        emk_df=None,
    )
    ops = analyzer.extract_operations()
    assert len(ops) > 0


def test_weeks_and_form_4001():
    _require_data()
    cfg = _config()
    df = read_table(str(JOURNAL))
    analyzer = SurgeryAnalyzer(
        df, cfg["departments"]["main"], cfg["surgery_categories"], emk_df=None
    )
    ops = analyzer.extract_operations()
    summary_cfg = cfg.get("summary", {})
    cat_table, totals_df, weeks = build_summary_tables(
        ops, summary_cfg, cfg["surgery_categories"]
    )
    assert cat_table is not None and not cat_table.empty
    assert totals_df is not None and not totals_df.empty
    assert len(weeks) >= 1

    form = compute_form_4001(
        ops,
        categories=cfg["surgery_categories"],
        pension_age=int(cfg.get("thresholds", {}).get("pension_age", 60)),
    )
    assert isinstance(form, dict)
    lines = form.get("lines") or {}
    assert "5.1" in lines and "5.2" in lines
    assert form.get("total", {}).get("total", 0) > 0


def test_summary_sheet_weeks_readable():
    _require_data()
    wb = openpyxl.load_workbook(SUMMARY, data_only=False)
    sheet = wb["Апрель"] if "Апрель" in wb.sheetnames else wb[wb.sheetnames[0]]
    weeks = read_sheet_weeks(sheet)
    wb.close()
    assert weeks, "не прочитаны недели из шаблона (даты как Excel serial?)"
    assert compute_month_weeks(2026, 4)
