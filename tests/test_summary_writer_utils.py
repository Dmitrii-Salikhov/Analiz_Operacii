# tests/test_summary_writer_utils.py
"""Юнит-тесты SummaryWriter (даты, недели, запись)."""
from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

APP = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(APP))

from analyzers.summary_writer import (
    SummaryWriter,
    _as_date,
    compute_month_weeks,
    date_to_week_col,
)


def test_as_date_serial_and_datetime():
    assert _as_date(46023) == date(2026, 1, 1)
    assert _as_date(datetime(2026, 1, 15, 12, 0)) == date(2026, 1, 15)
    assert _as_date("15.01.2026") == date(2026, 1, 15)
    assert _as_date(None) is None
    assert _as_date("=SUM(A1)") is None


def test_date_to_week_col():
    weeks = compute_month_weeks(2026, 1)
    assert date_to_week_col(date(2026, 1, 5), weeks) == 4
    assert date_to_week_col(date(2025, 12, 31), weeks) is None


def test_summary_writer_writes_ops(tmp_path: Path):
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Январь"
    ws["B4"] = "Аденотомия"
    ws["B40"] = "Всего операций"
    ws["B43"] = "Дети всего"
    ws["B44"] = None
    ws["B45"] = "Человек"
    ws["C2"] = (pd.Timestamp("2026-01-01") - pd.Timestamp("1899-12-30")).days
    wb.save(path)

    cfg = {
        "year": 2026,
        "category_rows": {"Аденотомия": 4},
        "totals_rows": {"children": 43, "patients": 45},
        "sheet_names": {1: "Январь"},
        "form_4001": {"enabled": False},
        "backup_keep": 3,
    }
    ops = pd.DataFrame(
        [
            {
                "Дата": pd.Timestamp("2026-01-08"),
                "КВС": 1,
                "Категория": "Аденотомия",
                "Возраст": 10,
            }
        ]
    )
    writer = SummaryWriter(path, cfg, department="ЛОР", categories=[])
    report = writer.write(ops, output_path=str(path), backup=False, write_form=False)
    assert report["cells_written"] >= 1
    assert report["months"]["Январь"]["ops"] == 1

    wb2 = openpyxl.load_workbook(path)
    assert wb2["Январь"]["D4"].value == 1


def test_summary_writer_protected_row_fallback(tmp_path: Path, monkeypatch):
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Январь"
    ws["B40"] = "Всего операций"
    ws["B43"] = "Дети всего"
    ws["B44"] = None
    ws["B45"] = "Человек"
    wb.create_sheet("Archive")
    wb["Archive"]["B40"] = "Всего операций"
    wb.save(path)

    cfg = {
        "year": 2026,
        "category_rows": {"A": 4},
        "totals_rows": {"children": 43, "patients": 45},
        "sheet_names": {1: "Январь"},
        "form_4001": {"enabled": False},
    }
    monkeypatch.setattr(
        "analyzers.summary_layout.apply_category_kind_fills",
        lambda *a, **k: None,
    )
    writer = SummaryWriter(path, cfg)
    writer.category_rows = {"A": "not-int"}
    report = writer.write(pd.DataFrame(), output_path=str(path), backup=False, write_weeks=False)
    assert "blank_delta" in report


def test_summary_writer_skips_extra_sheet_without_totals(tmp_path: Path):
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Январь"
    ws["B40"] = "Всего операций"
    wb.create_sheet("Notes")
    wb.save(path)

    cfg = {
        "year": 2026,
        "category_rows": {},
        "sheet_names": {1: "Январь"},
        "form_4001": {"enabled": False},
    }
    writer = SummaryWriter(path, cfg)
    writer.write(pd.DataFrame(), output_path=str(path), backup=False, write_weeks=False)


def test_summary_writer_missing_template(tmp_path: Path):
    cfg = {"year": 2026, "category_rows": {}, "sheet_names": {1: "Январь"}}
    writer = SummaryWriter(tmp_path / "nope.xlsx", cfg)
    with pytest.raises(FileNotFoundError):
        writer.write(pd.DataFrame())
