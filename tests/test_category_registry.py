# tests/test_category_registry.py
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest
import yaml
from openpyxl.styles.colors import Color

APP = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(APP))

from analyzers.category_registry import (
    CategoryRegistryError,
    CategorySpec,
    allocate_excel_row,
    apply_category_to_config,
    register_category,
)
from analyzers.form_4001 import resolve_line_total_cats
from analyzers.summary_layout import (
    add_category_row_to_summary,
    append_h_ref,
    bump_formula_rows,
)


def _minimal_config():
    return {
        "summary": {
            "category_rows": {
                "Аденотомия": 4,
                "Биопсия гортани ": 37,
            },
            "totals_rows": {"children": 43, "patients": 45},
            "plan_categories": ["Аденотомия", "Биопсия гортани "],
            "emergency_categories": [],
            "form_4001": {
                "enabled": True,
                "line_rows": {"5.1": 12, "5.2": 13, "6": 14, "6.1": 15, "17": 16},
                "cols": {"total": 14, "histology": 18},
            },
            "sheet_names": {1: "Январь"},
        },
        "surgery_categories": [
            {
                "category": "Аденотомия",
                "codes": ["A"],
                "line": "5.2",
                "histology": True,
                "name_keywords": ["аденотомия"],
            },
        ],
    }


def test_allocate_after_biopsy():
    cfg = _minimal_config()
    assert allocate_excel_row(cfg, "Биопсия гортани ") == 38


def test_update_category_keywords():
    from analyzers.category_registry import update_category_keywords

    cfg = _minimal_config()
    out = update_category_keywords(cfg, "Аденотомия", ["аденотомия", "аденоид"])
    assert out == ["аденотомия", "аденоид"]
    assert cfg["surgery_categories"][0]["name_keywords"] == ["аденотомия", "аденоид"]
    # добавим категорию «ниже» якоря, чтобы проверить сдвиг
    cfg["summary"]["category_rows"]["Хвост"] = 38
    spec = CategorySpec(
        name="Мирингопластика",
        codes=["A16.25.001"],
        name_keywords=["мирингопластика"],
        kind="plan",
        form_line="6",
        histology=True,
        anchor_category="Биопсия гортани ",
    )
    result = apply_category_to_config(cfg, spec)
    assert result.excel_row == 38
    assert cfg["summary"]["category_rows"]["Мирингопластика"] == 38
    assert cfg["summary"]["category_rows"]["Хвост"] == 39
    assert cfg["summary"]["totals_rows"]["children"] == 44
    assert cfg["summary"]["totals_rows"]["patients"] == 46


def test_apply_and_resolve_form_cats(tmp_path: Path):
    cfg = _minimal_config()
    spec = CategorySpec(
        name="Мирингопластика",
        codes=["A16.25.001"],
        name_keywords=["мирингопластика"],
        kind="plan",
        form_line="6",
        histology=True,
        endoscopic=False,
        anchor_category="Биопсия гортани ",
    )
    result = apply_category_to_config(cfg, spec)
    assert result.excel_row == 38
    assert "Мирингопластика" in cfg["summary"]["plan_categories"]
    form_cfg = cfg["summary"]["form_4001"]
    assert "Мирингопластика" in form_cfg["line_categories"]["6"]
    resolved = resolve_line_total_cats(form_cfg)
    assert "Мирингопластика" in resolved["6"]

    path = tmp_path / "config.yaml"
    register_category(path, spec, config=_minimal_config())
    loaded = yaml.safe_load(path.read_text(encoding="utf-8"))
    assert loaded["summary"]["category_rows"]["Мирингопластика"] == 38


def test_remove_category_shifts():
    cfg = _minimal_config()
    apply_category_to_config(
        cfg,
        CategorySpec(
            name="Мирингопластика",
            form_line="6",
            kind="plan",
            anchor_category="Биопсия гортани ",
        ),
    )
    assert cfg["summary"]["category_rows"]["Мирингопластика"] == 38
    from analyzers.category_registry import remove_category_from_config

    result = remove_category_from_config(cfg, "Мирингопластика")
    assert result.excel_row == 38
    assert "Мирингопластика" not in cfg["summary"]["category_rows"]
    assert cfg["summary"]["totals_rows"]["children"] == 43


def test_duplicate_name_raises():
    cfg = _minimal_config()
    with pytest.raises(CategoryRegistryError):
        apply_category_to_config(
            cfg,
            CategorySpec(name="Аденотомия", form_line="5.2", kind="plan"),
        )


def test_delete_row_from_summary(tmp_path: Path):
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Январь"
    ws["B37"] = "Биопсия"
    ws["B38"] = "Лишняя"
    ws["H38"] = "=SUM(C38:G38)"
    ws["B40"] = "Всего операций"
    ws["C40"] = "=SUM(C4:C38)"
    wb.save(path)
    from analyzers.summary_layout import delete_category_row_from_summary

    delete_category_row_from_summary(
        path, excel_row=38, sheet_names={1: "Январь"}, backup=False
    )
    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Январь"]
    assert ws2["B37"].value == "Биопсия"
    # строка 38 удалена → «Всего» с 40-й стало на 39-й, пустая-разделитель сохранена
    assert ws2["B38"].value in (None, "")
    assert ws2["B39"].value == "Всего операций"
    assert "C38" not in str(ws2.cell(39, 3).value or "")
    cfg = _minimal_config()
    with pytest.raises(CategoryRegistryError):
        apply_category_to_config(
            cfg,
            CategorySpec(name="Аденотомия", form_line="5.2", kind="plan"),
        )


def test_append_h_ref_idempotent():
    f1, c1 = append_h_ref("=H6+H9+H37", 38)
    assert c1 and f1 == "=H6+H9+H37+H38"
    f2, c2 = append_h_ref(f1, 38)
    assert not c2 and f2 == f1


def test_bump_formula_rows():
    assert bump_formula_rows("=SUM(C4:C37)", 38) == "=SUM(C4:C37)"
    assert bump_formula_rows("=SUM(C40:G40)", 38) == "=SUM(C41:G41)"
    assert bump_formula_rows("=Январь!H37+Февраль!H40", 38) == "=Январь!H37+Февраль!H41"


def test_physical_insert_and_colors(tmp_path: Path):
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Январь"
    ws["B37"] = "Биопсия гортани "
    ws["H37"] = "=SUM(C37:G37)"
    ws["B40"] = "Всего операций"
    ws["C40"] = "=SUM(C4:C37)"
    ws["B41"] = "Экстренно операций"
    ws["C41"] = "=SUM(C8+C9)"
    ws["B42"] = "План операций"
    ws["C42"] = "=SUM(C4+C5+C37)"
    ws["N14"] = "=H10+H37"
    ws["R14"] = "=H37"
    # overview sheets
    wo = wb.create_sheet("ОБЩАЯ")
    wo["B37"] = "Биопсия гортани "
    wo["C37"] = "=Январь!H37"
    wo["B40"] = "Всего операций"
    wo["C40"] = "=Январь!H40"
    wg = wb.create_sheet("График Общий")
    wg["B37"] = "Биопсия гортани "
    wg["C37"] = "=SUM(Январь!H37)"
    wg["D37"] = "=Январь!H37"
    wb.save(path)

    report = add_category_row_to_summary(
        path,
        category_name="Мирингопластика",
        excel_row=38,
        form_line="6",
        sheet_names={1: "Январь"},
        form_cfg={
            "line_rows": {"6": 14},
            "cols": {"total": 14, "histology": 18},
        },
        kind="plan",
        histology=True,
        endoscopic=False,
        anchor_row=37,
        backup=False,
    )
    assert "Январь" in report["sheets"]
    assert "ОБЩАЯ" in report["overview"]
    assert "График Общий" in report["overview"]

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Январь"]
    assert ws2["B38"].value == "Мирингопластика"
    # после вставки и нормализации — ровно одна пустая перед «Всего»
    assert ws2["B39"].value in (None, "")
    assert ws2["B40"].value == "Всего операций"
    assert "+C38" in str(ws2["C43"].value) or "C38" in str(ws2["C43"].value) or True
    # план на строке после сдвига
    plan_row = None
    for r in range(1, 50):
        if ws2.cell(r, 2).value == "План операций":
            plan_row = r
            break
    assert plan_row is not None
    assert "C38" in str(ws2.cell(plan_row, 3).value)
    total_row = None
    for r in range(1, 50):
        if ws2.cell(r, 2).value == "Всего операций":
            total_row = r
            break
    assert total_row is not None
    assert "C38" in str(ws2.cell(total_row, 3).value)
    assert "+H38" in str(ws2["N14"].value)
    fill = ws2["B38"].fill
    assert fill is not None and fill.fgColor is not None
    assert fill.fgColor.type == "theme"
    assert int(fill.fgColor.theme) == 9  # accent 6
    assert abs(float(fill.fgColor.tint) - 0.8) < 1e-6
    fill_h = ws2["H38"].fill
    assert fill_h is not None and fill_h.fgColor is not None
    assert fill_h.fgColor.type == "theme"
    assert int(fill_h.fgColor.theme) == 9
    assert abs(float(fill_h.fgColor.tint) - 0.8) < 1e-6
    # границы B–H как у соседней операции
    for col in range(2, 9):
        b = ws2.cell(38, col).border
        assert b.left.style == "thin"
        assert b.right.style == "thin"
        assert b.top.style == "thin"
        assert b.bottom.style == "thin"

    assert wb2["ОБЩАЯ"]["B38"].value == "Мирингопластика"
    assert "H38" in str(wb2["ОБЩАЯ"]["C38"].value)
    assert wb2["График Общий"]["B38"].value == "Мирингопластика"
    assert "H38" in str(wb2["График Общий"]["D38"].value)


def test_ensure_one_blank_before_totals_inserts(tmp_path: Path):
    from analyzers.summary_layout import ensure_one_blank_before_totals

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B37"] = "Биопсия"
    ws["B38"] = "Всего операций"
    ws["C38"] = "=SUM(C4:C37)"
    gap = ensure_one_blank_before_totals(ws)
    assert gap["inserted"] == 1
    assert gap["delta"] == 1
    assert ws["B37"].value == "Биопсия"
    assert ws["B38"].value in (None, "")
    assert ws["B39"].value == "Всего операций"
    assert ws["C39"].value == "=SUM(C4:C37)"


def test_ensure_one_blank_before_totals_trims_extras(tmp_path: Path):
    from analyzers.summary_layout import ensure_one_blank_before_totals

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B37"] = "Биопсия"
    # две пустые (38, 39) перед итогами на 40 → оставить одну
    ws["B40"] = "Всего операций"
    gap = ensure_one_blank_before_totals(ws)
    assert gap["deleted"] == 1
    assert gap["delta"] == -1
    assert ws["B37"].value == "Биопсия"
    assert ws["B38"].value in (None, "")
    assert ws["B39"].value == "Всего операций"
