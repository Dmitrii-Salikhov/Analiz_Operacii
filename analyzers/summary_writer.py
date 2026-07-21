# analyzers/summary_writer.py
"""Запись недельных счётчиков в шаблон «Операции сводная YYYY.xlsx»."""
from __future__ import annotations

import calendar
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

from analyzers.form_4001 import write_form_4001


MONTH_RU = {
    1: "январь",
    2: "февраль",
    3: "март",
    4: "апрель",
    5: "май",
    6: "июнь",
    7: "июль",
    8: "август",
    9: "сентябрь",
    10: "октябрь",
    11: "ноябрь",
    12: "декабрь",
}


def _as_date(val) -> Optional[date]:
    """Дата из ячейки Excel: datetime, строка или serial (дней с 1899-12-30)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        s = val.strip()
        if not s or s.startswith("="):
            return None
        try:
            return pd.to_datetime(s, dayfirst=True).date()
        except Exception:
            return None
    # Excel serial number (например 46054 = 01.02.2026)
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        n = float(val)
        # разумный диапазон дат Excel ~1900–2100
        if 1 <= n <= 80000:
            return date(1899, 12, 30) + timedelta(days=int(n))
        return None
    try:
        return pd.to_datetime(val, dayfirst=True).date()
    except Exception:
        return None


def excel_weekday2(d: date) -> int:
    """Excel WEEKDAY(d, 2): Mon=1 .. Sun=7."""
    return d.isoweekday()


def compute_month_weeks(year: int, month: int) -> List[Tuple[date, date]]:
    """Цепочка недель как в формулах шаблона (лист «Январь»)."""
    first = date(year, month, 1)
    last = date(year, month, calendar.monthrange(year, month)[1])
    weeks = []
    start = first
    end = min(start + timedelta(days=7 - excel_weekday2(start)), last)
    weeks.append((start, end))
    while end < last and len(weeks) < 5:
        start = end + timedelta(days=1)
        if start > last:
            break
        end = min(start + timedelta(days=6), last)
        weeks.append((start, end))
    return weeks


def read_sheet_weeks(ws) -> List[Tuple[date, date]]:
    """
    Читает границы недель из строк 2–3 (C–G).
    Если в шапке формулы или битые значения — считает от даты в C2 / месяца листа.
    """
    c2 = _as_date(ws.cell(2, 3).value)
    parsed: List[Optional[Tuple[date, date]]] = []
    for col in range(3, 8):
        start = _as_date(ws.cell(2, col).value)
        end = _as_date(ws.cell(3, col).value)
        if start and end and start.year >= 2000 and end.year >= 2000:
            parsed.append((start, end))
        elif ws.cell(2, col).value is None and ws.cell(3, col).value is None:
            break
        else:
            # формула / неполный заголовок — считаем по C2
            parsed = []
            break

    if parsed:
        return parsed  # type: ignore[return-value]

    if c2 is not None and c2.year >= 2000:
        return compute_month_weeks(c2.year, c2.month)
    return []


def date_to_week_col(op_date: date, weeks: List[Tuple[date, date]]) -> Optional[int]:
    """Возвращает номер колонки Excel (3=C …) для даты операции."""
    for i, (start, end) in enumerate(weeks):
        if start <= op_date <= end:
            return 3 + i
    return None


class SummaryWriter:
    def __init__(
        self,
        template_path: str,
        summary_cfg: dict,
        department: str = "",
        categories: Optional[List[dict]] = None,
        pension_age: int = 60,
    ):
        self.template_path = Path(template_path)
        self.cfg = summary_cfg
        self.department = department
        self.categories = categories or []
        self.pension_age = int(pension_age)
        self.category_rows: Dict[str, int] = dict(summary_cfg.get("category_rows", {}))
        self.sheet_names: Dict[int, str] = {
            int(k): v for k, v in summary_cfg.get("sheet_names", {}).items()
        }
        self.children_row = summary_cfg.get("totals_rows", {}).get("children", 43)
        self.patients_row = summary_cfg.get("totals_rows", {}).get("patients", 45)
        self.form_4001_cfg = summary_cfg.get("form_4001") or {}

    def write(
        self,
        ops_df: pd.DataFrame,
        output_path: Optional[str] = None,
        overwrite_from=None,
        overwrite_to=None,
        backup: bool = True,
        write_weeks: bool = True,
        write_form: bool = False,
    ) -> dict:
        """
        Пишет счётчики в шаблон.
        write_weeks — колонки недель C–G (категории / дети / человек).
        write_form — форма 4001 (P/Q/S и пенсионная строка; формулы N/R не затираются).
        По умолчанию пишутся только недели; форму — отдельной кнопкой.
        """
        if not self.template_path.exists():
            raise FileNotFoundError(f"Шаблон не найден: {self.template_path}")

        out = Path(output_path) if output_path else self.template_path
        report_bak = None
        if backup and out.exists():
            from analyzers.backup_utils import DEFAULT_BACKUP_KEEP, make_backup

            report_bak = str(
                make_backup(out, keep=int(self.cfg.get("backup_keep", DEFAULT_BACKUP_KEEP))) or ""
            ) or None

        wb = openpyxl.load_workbook(self.template_path)
        report = {
            "months": {},
            "unmapped_dates": 0,
            "cells_written": 0,
            "write_weeks": write_weeks,
            "write_form": write_form,
            "blank_delta": 0,
        }
        if report_bak:
            report["backup"] = report_bak

        # Ровно одна пустая строка между операциями и «Всего операций»
        from analyzers.summary_layout import (
            _classify_sheet,
            _find_label_row,
            ensure_one_blank_before_totals,
        )

        blank_delta_set = False
        month_name_set = set(self.sheet_names.values())
        for sn in list(wb.sheetnames):
            ws = wb[sn]
            if sn not in month_name_set and not _classify_sheet(sn):
                if not _find_label_row(ws, "Всего операций"):
                    continue
            gap = ensure_one_blank_before_totals(ws)
            if not blank_delta_set:
                report["blank_delta"] = int(gap.get("delta") or 0)
                blank_delta_set = True
        if report["blank_delta"]:
            self.children_row = int(self.children_row) + report["blank_delta"]
            self.patients_row = int(self.patients_row) + report["blank_delta"]

        # Заливка названия (B) и суммы (H) — акцент план/экстр
        from analyzers.summary_layout import apply_category_kind_fills

        emergency = self.cfg.get("emergency_categories") or []
        for sn in month_name_set:
            if sn in wb.sheetnames:
                apply_category_kind_fills(
                    wb[sn], self.category_rows, emergency_categories=emergency
                )

        ops = pd.DataFrame() if ops_df is None else ops_df.copy()
        if not ops.empty:
            ops["Дата"] = pd.to_datetime(ops["Дата"], errors="coerce")
            ops = ops.dropna(subset=["Дата"])
            ops["_month"] = ops["Дата"].dt.month
            ops["_year"] = ops["Дата"].dt.year

        year = self.cfg.get("year", 2026)

        # месяцы для обработки: из данных + из диапазона перезаписи
        months = set()
        if not ops.empty:
            months |= set(int(m) for m in ops["_month"].unique())
        if overwrite_from is not None and overwrite_to is not None:
            cur = pd.Timestamp(overwrite_from).to_period("M")
            end = pd.Timestamp(overwrite_to).to_period("M")
            while cur <= end:
                months.add(int(cur.month))
                cur += 1

        ow_from = pd.Timestamp(overwrite_from).normalize() if overwrite_from is not None else None
        ow_to = pd.Timestamp(overwrite_to).normalize() if overwrite_to is not None else None

        for month in sorted(months):
            sheet_name = self.sheet_names.get(int(month))
            if not sheet_name or sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            month_ops = ops[ops["_month"] == month] if not ops.empty else pd.DataFrame()

            weeks = read_sheet_weeks(ws)
            if not weeks:
                weeks = compute_month_weeks(year, int(month))
            elif not month_ops.empty:
                # если в шаблоне другой год — считаем недели по году данных
                data_year = int(pd.to_datetime(month_ops["Дата"]).dt.year.mode().iloc[0])
                if weeks[0][0].year != data_year:
                    weeks = compute_month_weeks(data_year, int(month))
                    # и синхронизируем C2, чтобы Excel пересчитал шапку (дата без времени)
                    c2 = ws.cell(2, 3)
                    c2.value = (
                        date(data_year, int(month), 1) - date(1899, 12, 30)
                    ).days
                    c2.number_format = "DD/MM/YYYY"

            cols_touched = set()
            if write_weeks:
                for i, (start, end) in enumerate(weeks):
                    col = 3 + i
                    has_ops = False
                    if not month_ops.empty:
                        has_ops = bool(month_ops["Дата"].dt.date.between(start, end).any())
                    in_overwrite = False
                    if ow_from is not None and ow_to is not None:
                        in_overwrite = end >= ow_from.date() and start <= ow_to.date()
                    if has_ops or in_overwrite:
                        cols_touched.add(col)

                for d in month_ops["Дата"] if not month_ops.empty else []:
                    if date_to_week_col(d.date(), weeks) is None:
                        report["unmapped_dates"] += 1

                for col in cols_touched:
                    self._clear_week_column(ws, col)
                    report["cells_written"] += self._fill_week_column(ws, col, month_ops, weeks)

                title = f"{self.department} за {MONTH_RU.get(int(month), '')} {year}".strip()
                if self.department:
                    form_on = bool((self.form_4001_cfg or {}).get("enabled", False))
                    if form_on:
                        # ЛОР: заголовок в A1
                        ws.cell(1, 1).value = title
                    else:
                        # сводные отделений: заголовок в B1, A узкий
                        ws.cell(1, 1).value = None
                        cell_title = ws.cell(1, 2)
                        cell_title.value = title
                        try:
                            from openpyxl.styles import Font

                            cell_title.font = Font(bold=True)
                        except Exception:
                            pass
                        ws.column_dimensions["A"].width = 13.0

            form_info = {}
            if write_form and self.form_4001_cfg.get("enabled", True):
                form_info = write_form_4001(
                    ws,
                    month_ops,
                    self.categories,
                    self.form_4001_cfg,
                    pension_age=self.pension_age,
                )
                report["cells_written"] += int(form_info.get("written", 0))

            sample = {}
            if write_weeks:
                for cat, row in list(self.category_rows.items())[:5]:
                    sample[cat] = [ws.cell(row, c).value for c in sorted(cols_touched)]

            report["months"][sheet_name] = {
                "weeks": [(str(s), str(e)) for s, e in weeks],
                "cols": sorted(cols_touched),
                "ops": len(month_ops),
                "sample": sample,
                "form_4001": form_info.get("stats"),
            }

        wb.save(out)
        # openpyxl может вернуть старую тему — для non-LOR снова подставляем тему ЛОР
        if not bool((self.form_4001_cfg or {}).get("enabled", False)):
            try:
                from analyzers.dept_template import apply_lor_workbook_theme

                apply_lor_workbook_theme(out)
            except Exception:
                pass
        report["output"] = str(out)
        return report

    def _clear_week_column(self, ws, col: int):
        for row in self.category_rows.values():
            ws.cell(row, col).value = None
        ws.cell(self.children_row, col).value = None
        ws.cell(self.patients_row, col).value = None

    def _fill_week_column(self, ws, col: int, month_ops: pd.DataFrame, weeks: List[Tuple[date, date]]) -> int:
        """Пишет счётчики по категориям; 0 не ставим — ячейка остаётся пустой."""
        idx = col - 3
        if idx < 0 or idx >= len(weeks):
            return 0
        start, end = weeks[idx]
        if month_ops is None or month_ops.empty:
            week_ops = pd.DataFrame(columns=["Категория", "Возраст", "КВС"])
        else:
            mask = month_ops["Дата"].dt.date.between(start, end)
            week_ops = month_ops.loc[mask]
        written = 0

        counts = week_ops["Категория"].value_counts() if not week_ops.empty else pd.Series(dtype=int)
        for cat, row in self.category_rows.items():
            val = int(counts.get(cat, 0))
            # нет операций этой категории — пусто, не ноль
            ws.cell(row, col).value = val if val else None
            written += 1

        if week_ops.empty:
            children = 0
            patients = 0
        else:
            children = int(week_ops.loc[week_ops["Возраст"].fillna(99) < 18, "КВС"].nunique())
            patients = int(week_ops["КВС"].nunique())
        # итоги «Дети» / «Человек»: 0 тоже не пишем
        ws.cell(self.children_row, col).value = children if children else None
        ws.cell(self.patients_row, col).value = patients if patients else None
        written += 2
        return written
