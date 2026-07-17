# analyzers/write_verify.py
"""Проверка, что после записи ячейки в Excel совпадают с ожидаемым."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

import openpyxl


def verify_write_report(summary_path: str | Path, report: Dict[str, Any]) -> Dict[str, Any]:
    """
    Сверяет sample из отчёта записи с фактическими ячейками файла.
    Возвращает {ok, checked, mismatches:[{sheet, cat, col, expected, actual}]}.
    """
    path = Path(summary_path)
    result: Dict[str, Any] = {"ok": True, "checked": 0, "mismatches": []}
    if not path.exists():
        result["ok"] = False
        result["error"] = f"файл не найден: {path}"
        return result

    months = report.get("months") or {}
    if not months:
        result["note"] = "нечего проверять (нет месяцев в отчёте)"
        return result

    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        for sheet_name, info in months.items():
            if sheet_name not in wb.sheetnames:
                result["ok"] = False
                result["mismatches"].append(
                    {
                        "sheet": sheet_name,
                        "cat": "—",
                        "col": "—",
                        "expected": "лист есть",
                        "actual": "листа нет",
                    }
                )
                continue
            ws = wb[sheet_name]
            cols: List[int] = list(info.get("cols") or [])
            sample: Dict[str, list] = dict(info.get("sample") or {})
            for cat, values in sample.items():
                for i, expected in enumerate(values):
                    if i >= len(cols):
                        break
                    col = cols[i]
                    # найти строку категории по подписи в колонке B
                    row = _find_category_row(ws, cat)
                    if row is None:
                        continue
                    actual = ws.cell(row, col).value
                    result["checked"] += 1
                    exp_n = _as_int(expected)
                    act_n = _as_int(actual)
                    if exp_n != act_n:
                        result["ok"] = False
                        result["mismatches"].append(
                            {
                                "sheet": sheet_name,
                                "cat": cat,
                                "col": col,
                                "expected": exp_n,
                                "actual": act_n,
                            }
                        )
    finally:
        wb.close()

    return result


def _as_int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        return -1


def _find_category_row(ws, cat: str) -> int | None:
    needle = str(cat).strip()
    for row in range(4, 40):
        val = ws.cell(row, 2).value
        if val is None:
            continue
        if str(val).strip() == needle:
            return row
    return None


def format_verify_message(result: Dict[str, Any]) -> str:
    if result.get("error"):
        return f"Проверка не удалась: {result['error']}"
    if result.get("ok"):
        n = result.get("checked", 0)
        return f"Проверка записи: OK ({n} ячеек совпали с ожидаемым)."
    lines = ["Проверка записи: есть расхождения:"]
    for m in (result.get("mismatches") or [])[:12]:
        lines.append(
            f"  [{m.get('sheet')}] {m.get('cat')} col={m.get('col')}: "
            f"ждали {m.get('expected')}, в файле {m.get('actual')}"
        )
    extra = len(result.get("mismatches") or []) - 12
    if extra > 0:
        lines.append(f"  … ещё {extra}")
    return "\n".join(lines)
