# analyzers/year_template.py
"""Создание сводной на новый год из шаблона предыдущего."""
from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path
from typing import Dict, Optional

import openpyxl

from analyzers.summary_writer import MONTH_RU

# Excel epoch (совместимо с openpyxl / Windows Excel)
_EXCEL_EPOCH = date(1899, 12, 30)


def _set_sheet_month_start(ws, year: int, month: int) -> None:
    """C2 = 1-е число месяца как дата Excel без времени (иначе #### в узком столбце)."""
    cell = ws.cell(2, 3)
    d = date(year, month, 1)
    # целое serial — без datetime/00:00:00; формат как у соседних дат шапки (слэш, не точка/запятая)
    cell.value = (d - _EXCEL_EPOCH).days
    cell.number_format = "DD/MM/YYYY"


def _clear_if_not_formula(cell) -> None:
    v = cell.value
    if isinstance(v, str) and v.startswith("="):
        return
    cell.value = None


def create_year_summary(
    template_path: str,
    new_year: int,
    output_path: Optional[str] = None,
    sheet_names: Optional[Dict[int, str]] = None,
    clear_values: bool = True,
    category_row_max: Optional[int] = None,
    totals_rows: Optional[Dict[str, int]] = None,
) -> Path:
    """
    Копирует шаблон сводной и готовит файл на new_year:
    - C2 каждого месячного листа = 1-е число месяца new_year (даты недель пересчитает Excel);
    - при clear_values очищает числовые ячейки категорий C–G и ручные ячейки формы 4001
      (в т.ч. колонку S «Всего (из гр.3 т.4000)»); формулы не трогает.
    """
    src = Path(template_path)
    if not src.exists():
        raise FileNotFoundError(src)
    out = Path(output_path) if output_path else src.with_name(f"Операции сводная {new_year}.xlsx")
    shutil.copy2(src, out)

    names = sheet_names or {
        1: "Январь",
        2: "Февраль",
        3: "Март",
        4: "Апрель",
        5: "Май",
        6: "Июнь",
        7: "Июль",
        8: "Август",
        9: "Сентябрь",
        10: "Октябрь ",
        11: "Ноябрь",
        12: "Декабрь",
    }

    cat_hi = int(category_row_max) if category_row_max else 37
    cat_hi = max(4, cat_hi)
    tot = totals_rows or {"children": 43, "patients": 45}
    tot_rows = sorted({int(v) for v in tot.values()})

    wb = openpyxl.load_workbook(out)
    for month, sheet in names.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        _set_sheet_month_start(ws, new_year, month)

        title = ws.cell(1, 1).value
        if isinstance(title, str):
            for y in range(new_year - 5, new_year + 2):
                title = title.replace(str(y), str(new_year))
            ws.cell(1, 1).value = title

        if clear_values:
            # категории — C–G до фактического max(category_rows)
            for row in range(4, cat_hi + 1):
                for col in range(3, 8):
                    _clear_if_not_formula(ws.cell(row, col))
            for row in tot_rows:
                for col in range(3, 8):
                    _clear_if_not_formula(ws.cell(row, col))
            # форма 4001: N–S (14–19), строки 11–17
            # в т.ч. S «Всего (из гр.3 т.4000)»; формулы итогов не трогаем
            for row in range(11, 18):
                for col in range(14, 20):
                    _clear_if_not_formula(ws.cell(row, col))

    wb.save(out)
    return out


def suggest_summary_path(app_dir: Path, year: int) -> Path:
    return Path(app_dir) / f"Операции сводная {year}.xlsx"
