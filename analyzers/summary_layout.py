# analyzers/summary_layout.py
"""Физическая вставка строки категории в сводную, формулы, ОБЩАЯ/графики, цвета."""
from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

from analyzers.backup_utils import DEFAULT_BACKUP_KEEP, make_backup
from analyzers.form_4001 import FORM_ROWS

# Заливка названия: тема Excel — план accent 6 / экстр. accent 2, светлый 80%
# theme: accent1=4 … accent6=9
THEME_ACCENT_PLAN = 9       # акцент 6 (салатовый)
THEME_ACCENT_EMERGENCY = 5  # акцент 2 (оранжевый)
THEME_TINT_LIGHT = 0.8      # светлый цвет 80%

_H_REF_RE = re.compile(r"\bH(\d+)\b", re.IGNORECASE)
# A1 / $A$1 / Sheet!A1 / 'Октябрь '!H4
_CELL_REF_RE = re.compile(
    r"(?:(?P<sheet>'(?:[^']|'')+'|[A-Za-zА-Яа-яЁё][A-Za-zА-Яа-яЁё0-9_ ]*)!)?"
    r"(?P<col>\$?[A-Z]{1,3})(?P<row>\$?\d+)\b",
    re.IGNORECASE,
)

OVERVIEW_ALIASES = {
    "общая": "overview",
    "график общий": "chart",
}


def append_h_ref(formula: str, row: int) -> Tuple[str, bool]:
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula, False
    target = f"H{int(row)}"
    existing = {f"H{m.group(1)}" for m in _H_REF_RE.finditer(formula)}
    if any(e.upper() == target.upper() for e in existing):
        return formula, False
    return f"{formula}+{target}", True


def remove_h_ref(formula: str, row: int) -> Tuple[str, bool]:
    """Убирает H{row} / +H{row} из формулы."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula, False
    new_v, n = re.subn(rf"\+?H{int(row)}\b", "", formula, flags=re.IGNORECASE)
    if n == 0:
        return formula, False
    new_v = re.sub(r"=\+", "=", new_v)
    new_v = re.sub(r"\(\+", "(", new_v)
    new_v = re.sub(r"\+\+", "+", new_v)
    new_v = re.sub(r"\+$", "", new_v)
    new_v = re.sub(r"\+$", "", new_v)
    if new_v.endswith("+)"):
        new_v = new_v[:-2] + ")"
    return new_v, True


def bump_formula_rows(formula: str, insert_at: int, amount: int = 1) -> str:
    """Увеличивает номера строк >= insert_at в формуле (в т.ч. Sheet!A1)."""
    if not isinstance(formula, str) or amount == 0:
        return formula

    def repl(m: re.Match) -> str:
        sheet = m.group("sheet") or ""
        col = m.group("col")
        row_raw = m.group("row")
        abs_row = row_raw.startswith("$")
        row = int(row_raw.replace("$", ""))
        if row >= insert_at:
            row += amount
        row_s = f"${row}" if abs_row else str(row)
        prefix = f"{sheet}!" if sheet else ""
        return f"{prefix}{col}{row_s}"

    return _CELL_REF_RE.sub(repl, formula)


def bump_sheet_formulas(ws, insert_at: int, amount: int = 1) -> int:
    """После insert_rows: поправить ссылки в ячейках (openpyxl сам формулы не двигает)."""
    changed = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                new_v = bump_formula_rows(v, insert_at, amount)
                if new_v != v:
                    cell.value = new_v
                    changed += 1
    return changed


def bump_chart_refs(ws, insert_at: int, amount: int = 1) -> int:
    changed = 0
    for chart in getattr(ws, "_charts", []) or []:
        for series in chart.series:
            for attr in ("val", "cat"):
                src = getattr(series, attr, None)
                if src is None:
                    continue
                num = getattr(src, "numRef", None)
                if num is not None and getattr(num, "f", None):
                    new_f = bump_formula_rows(num.f, insert_at, amount)
                    # abs refs $C$4
                    new_f2 = re.sub(
                        r"\$([A-Z]{1,3})\$(\d+)",
                        lambda m: f"${m.group(1)}${int(m.group(2)) + amount}"
                        if int(m.group(2)) >= insert_at
                        else m.group(0),
                        new_f,
                    )
                    if new_f2 != num.f:
                        num.f = new_f2
                        changed += 1
                stre = getattr(src, "strRef", None)
                if stre is not None and getattr(stre, "f", None):
                    new_f = bump_formula_rows(stre.f, insert_at, amount)
                    new_f2 = re.sub(
                        r"\$([A-Z]{1,3})\$(\d+)",
                        lambda m: f"${m.group(1)}${int(m.group(2)) + amount}"
                        if int(m.group(2)) >= insert_at
                        else m.group(0),
                        new_f,
                    )
                    if new_f2 != stre.f:
                        stre.f = new_f2
                        changed += 1
            title = getattr(series, "title", None)
            if title is not None:
                stre = getattr(title, "strRef", None)
                if stre is not None and getattr(stre, "f", None):
                    new_f = bump_formula_rows(stre.f, insert_at, amount)
                    new_f2 = re.sub(
                        r"\$([A-Z]{1,3})\$(\d+)",
                        lambda m: f"${m.group(1)}${int(m.group(2)) + amount}"
                        if int(m.group(2)) >= insert_at
                        else m.group(0),
                        new_f,
                    )
                    if new_f2 != stre.f:
                        stre.f = new_f2
                        changed += 1
    return changed


def _quote_sheet(name: str) -> str:
    if re.search(r"[\s!']", name):
        return "'" + name.replace("'", "''") + "'"
    return name


def _year_sum_h_formula(month_sheet_names: Sequence[str], row: int) -> str:
    parts = [f"{_quote_sheet(s)}!H{row}" for s in month_sheet_names]
    return "=SUM(" + "+".join(parts) + ")"


def _fill_for_kind(kind: str) -> PatternFill:
    theme = THEME_ACCENT_EMERGENCY if kind == "emergency" else THEME_ACCENT_PLAN
    return PatternFill(
        fill_type="solid",
        fgColor=Color(theme=theme, tint=THEME_TINT_LIGHT),
    )


def _thin_border() -> Border:
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _border_is_set(border) -> bool:
    if border is None:
        return False
    for attr in ("left", "right", "top", "bottom"):
        side = getattr(border, attr, None)
        if side is not None and side.style:
            return True
    return False


def _border_from_template(template_cell) -> Border:
    if template_cell is not None and _border_is_set(template_cell.border):
        return copy(template_cell.border)
    return _thin_border()


def _apply_name_style(cell, kind: str, template_cell) -> None:
    """Стиль названия/суммы: заливка по типу, границы как у соседней операции."""
    if template_cell is not None and template_cell.has_style:
        cell.font = copy(template_cell.font)
        cell.alignment = copy(template_cell.alignment)
        cell.number_format = template_cell.number_format
    cell.border = _border_from_template(template_cell)
    cell.fill = _fill_for_kind(kind)


def _apply_week_cell_style(cell, template_cell) -> None:
    """Стиль ячеек недель C–G: границы/шрифт как у шаблона, без заливки акцента."""
    if template_cell is not None and template_cell.has_style:
        cell.font = copy(template_cell.font)
        cell.alignment = copy(template_cell.alignment)
        cell.number_format = template_cell.number_format
        if template_cell.fill and template_cell.fill.fill_type:
            cell.fill = copy(template_cell.fill)
    cell.border = _border_from_template(template_cell)


def _series_formula_rows(chart) -> List[int]:
    rows: List[int] = []
    for series in getattr(chart, "series", []) or []:
        for attr in ("val", "cat"):
            src = getattr(series, attr, None)
            if src is None:
                continue
            for ref_attr in ("numRef", "strRef"):
                ref = getattr(src, ref_attr, None)
                f = getattr(ref, "f", None) if ref is not None else None
                if not f:
                    continue
                for m in re.finditer(r"\$[A-Z]{1,3}\$(\d+)", f, re.I):
                    rows.append(int(m.group(1)))
    return rows


def _is_category_operations_chart(ws, chart) -> bool:
    """
    График по отдельным операциям (строки выше «Всего операций»).
    Сводный график итогов (План/Экстренно/Дети/…) — не трогаем.
    """
    rows = _series_formula_rows(chart)
    if not rows:
        return False
    totals_row = _find_label_row(ws, "Всего операций")
    if totals_row is None:
        # без метки — считаем категорийным, только если все серии «высокие» строки данных
        return max(rows) < 40
    # все серии строго выше блока итогов
    return max(rows) < int(totals_row)


def _find_label_row(ws, label_substr: str) -> Optional[int]:
    needle = label_substr.strip().lower()
    for r in range(1, (ws.max_row or 1) + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and needle in v.strip().lower():
            return r
    return None


def _last_category_row_before(ws, totals_row: int) -> Optional[int]:
    for r in range(totals_row - 1, 3, -1):
        v = ws.cell(r, 2).value
        if v is not None and str(v).strip():
            return r
    return None


def ensure_one_blank_before_totals(ws) -> dict:
    """
    Между последней операцией и «Всего операций» — ровно одна пустая строка.
    Возвращает {inserted, deleted, blank_at, delta} (delta = inserted - deleted).
    """
    result = {"inserted": 0, "deleted": 0, "blank_at": None, "delta": 0}
    totals_row = _find_label_row(ws, "Всего операций")
    if not totals_row:
        return result
    last_cat = _last_category_row_before(ws, totals_row)
    if last_cat is None:
        return result
    gap = totals_row - last_cat - 1
    if gap == 1:
        result["blank_at"] = last_cat + 1
        return result
    if gap == 0:
        ws.insert_rows(totals_row, amount=1)
        bump_sheet_formulas(ws, totals_row, 1)
        bump_chart_refs(ws, totals_row, 1)
        result["inserted"] = 1
        result["delta"] = 1
        result["blank_at"] = totals_row
        return result
    # gap > 1 — удалять пустые сразу над итогами, пока не останется одна
    while True:
        totals_row = _find_label_row(ws, "Всего операций")
        if not totals_row:
            break
        last_cat = _last_category_row_before(ws, totals_row)
        if last_cat is None:
            break
        gap = totals_row - last_cat - 1
        if gap <= 1:
            break
        del_row = totals_row - 1
        v = ws.cell(del_row, 2).value
        if v is not None and str(v).strip():
            break
        ws.delete_rows(del_row, amount=1)
        bump_sheet_formulas(ws, del_row + 1, -1)
        bump_chart_refs(ws, del_row + 1, -1)
        result["deleted"] += 1
        result["delta"] -= 1
    totals_row = _find_label_row(ws, "Всего операций")
    last_cat = _last_category_row_before(ws, totals_row) if totals_row else None
    if last_cat is not None:
        result["blank_at"] = last_cat + 1
    return result


def _extend_or_append_col_ref(formula: str, col: str, row: int) -> Tuple[str, bool]:
    """Расширяет SUM(C4:C37) или дописывает +C38 в список SUM(C4+C5+…)."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula, False
    col_u = col.upper()
    ref = f"{col_u}{row}"
    if re.search(rf"\b{col_u}{row}\b", formula, re.IGNORECASE):
        return formula, False

    m = re.search(rf"SUM\(({col_u})(\d+):({col_u})(\d+)\)", formula, re.IGNORECASE)
    if m:
        start, end = int(m.group(2)), int(m.group(4))
        new_end = max(end, row)
        new_start = min(start, row)
        new_frag = f"SUM({col_u}{new_start}:{col_u}{new_end})"
        return formula[: m.start()] + new_frag + formula[m.end() :], True

    # SUM(C4+C5+C6) или C4+C5 без SUM
    if "+" in formula and re.search(rf"\b{col_u}\d+\b", formula, re.IGNORECASE):
        if formula.endswith(")"):
            return formula[:-1] + f"+{ref})", True
        return formula + f"+{ref}", True

    return formula, False


def patch_totals_and_kind_formulas(ws, *, excel_row: int, kind: str) -> Dict[str, int]:
    """Всего операций (диапазон) + Экстренно/План (список) для столбцов C–G."""
    stats = {"total": 0, "kind": 0}
    total_row = _find_label_row(ws, "Всего операций")
    if total_row:
        for col in range(3, 8):  # C–G
            cell = ws.cell(total_row, col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                letter = get_column_letter(col)
                new_v, changed = _extend_or_append_col_ref(cell.value, letter, excel_row)
                if changed:
                    cell.value = new_v
                    stats["total"] += 1

    label = "Экстренно" if kind == "emergency" else "План операций"
    kind_row = _find_label_row(ws, label)
    if kind == "plan" and kind_row is None:
        kind_row = _find_label_row(ws, "План")
    if kind_row:
        for col in range(3, 8):
            cell = ws.cell(kind_row, col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                letter = get_column_letter(col)
                new_v, changed = _extend_or_append_col_ref(cell.value, letter, excel_row)
                if changed:
                    cell.value = new_v
                    stats["kind"] += 1
    return stats


def append_col_ref(formula: str, col: str, row: int) -> Tuple[str, bool]:
    """Дописывает +ColRow к формуле, если ссылки ещё нет."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula, False
    col_u = col.upper()
    ref = f"{col_u}{int(row)}"
    if re.search(rf"\b{col_u}{int(row)}\b", formula, re.IGNORECASE):
        return formula, False
    return f"{formula}+{ref}", True


def patch_overview_form_c_refs(
    ws,
    *,
    excel_row: int,
    form_line: str,
    form_cfg: dict,
    histology: bool = False,
    endoscopic: bool = False,
) -> Dict[str, int]:
    """На ОБЩАЯ итоги формы считают SUM(C…) в колонке I (и M для морфологии)."""
    patched = {"n": 0, "r": 0, "endo": 0}
    leaf_row = _form_line_excel_row(form_cfg, form_line)
    if leaf_row is not None:
        cell_i = ws.cell(leaf_row, 9)  # I
        if isinstance(cell_i.value, str) and cell_i.value.startswith("="):
            new_v, changed = append_col_ref(cell_i.value, "C", excel_row)
            if changed:
                cell_i.value = new_v
                patched["n"] += 1
        if histology:
            cell_m = ws.cell(leaf_row, 13)  # M — морфология на ОБЩАЯ
            if isinstance(cell_m.value, str) and cell_m.value.startswith("="):
                new_v, changed = append_col_ref(cell_m.value, "C", excel_row)
                if changed:
                    cell_m.value = new_v
                    patched["r"] += 1
    if endoscopic:
        endo_row = _endo_excel_row(form_cfg)
        cell_e = ws.cell(endo_row, 9)
        if isinstance(cell_e.value, str) and cell_e.value.startswith("="):
            new_v, changed = append_col_ref(cell_e.value, "C", excel_row)
            if changed:
                cell_e.value = new_v
                patched["endo"] += 1
    return patched
    line_rows = (form_cfg or {}).get("line_rows") or {}
    if line in line_rows:
        return int(line_rows[line])
    for spec in FORM_ROWS:
        if spec.get("line") == line and spec.get("kind") == "leaf":
            return int(spec["excel_row"])
    return None


def _form_line_excel_row(form_cfg: dict, line: str) -> Optional[int]:
    line_rows = (form_cfg or {}).get("line_rows") or {}
    if line in line_rows:
        return int(line_rows[line])
    for spec in FORM_ROWS:
        if spec.get("line") == line and spec.get("kind") == "leaf":
            return int(spec["excel_row"])
    return None


def _endo_excel_row(form_cfg: dict) -> int:
    for spec in FORM_ROWS:
        if spec.get("kind") == "endo":
            return int(spec["excel_row"])
    return 20


def patch_form_formulas_on_sheet(
    ws,
    *,
    excel_row: int,
    form_line: str,
    form_cfg: dict,
    histology: bool = False,
    endoscopic: bool = False,
) -> Dict[str, int]:
    cols = (form_cfg or {}).get("cols") or {}
    col_total = int(cols.get("total", 14))
    col_hist = int(cols.get("histology", 18))
    patched = {"n": 0, "r": 0, "endo": 0, "skipped": 0}

    leaf_row = _form_line_excel_row(form_cfg, form_line)
    if leaf_row is None:
        patched["skipped"] += 1
    else:
        cell_n = ws.cell(leaf_row, col_total)
        if isinstance(cell_n.value, str) and cell_n.value.startswith("="):
            new_v, changed = append_h_ref(cell_n.value, excel_row)
            if changed:
                cell_n.value = new_v
                patched["n"] += 1
        else:
            patched["skipped"] += 1
        if histology:
            cell_r = ws.cell(leaf_row, col_hist)
            if isinstance(cell_r.value, str) and cell_r.value.startswith("="):
                new_v, changed = append_h_ref(cell_r.value, excel_row)
                if changed:
                    cell_r.value = new_v
                    patched["r"] += 1

    if endoscopic:
        endo_row = _endo_excel_row(form_cfg)
        cell_e = ws.cell(endo_row, col_total)
        if isinstance(cell_e.value, str) and cell_e.value.startswith("="):
            new_v, changed = append_h_ref(cell_e.value, excel_row)
            if changed:
                cell_e.value = new_v
                patched["endo"] += 1
    return patched


def _add_month_chart_series(ws, excel_row: int) -> int:
    """Добавляет серию C:G / подпись B только на график категорий (не на итоги)."""
    from copy import deepcopy

    added = 0
    for chart in getattr(ws, "_charts", []) or []:
        if not _is_category_operations_chart(ws, chart):
            continue
        if not chart.series:
            continue
        try:
            proto = chart.series[-1]
            new_s = deepcopy(proto)
            q = _quote_sheet(ws.title)
            if new_s.val and new_s.val.numRef:
                new_s.val.numRef.f = f"{q}!$C${excel_row}:$G${excel_row}"
            if new_s.title and getattr(new_s.title, "strRef", None):
                new_s.title.strRef.f = f"{q}!$B${excel_row}"
            chart.series.append(new_s)
            added += 1
        except Exception:
            pass
    return added


def _add_overview_chart_series(ws, excel_row: int) -> int:
    """График Общий: серия D:O только на график операций, не на сводный итогов."""
    from copy import deepcopy

    added = 0
    for chart in getattr(ws, "_charts", []) or []:
        if not _is_category_operations_chart(ws, chart):
            continue
        if not chart.series:
            continue
        # дополнительно: серии категорий обычно D:O по одной строке операции
        sample = ""
        num = getattr(getattr(chart.series[0], "val", None), "numRef", None)
        if num is not None:
            sample = getattr(num, "f", "") or ""
        if not re.search(r"\$D\$\d+:\$O\$\d+", sample, re.I):
            continue
        try:
            proto = chart.series[-1]
            new_s = deepcopy(proto)
            q = _quote_sheet(ws.title)
            if new_s.val and new_s.val.numRef:
                new_s.val.numRef.f = f"{q}!$D${excel_row}:$O${excel_row}"
            if new_s.title and getattr(new_s.title, "strRef", None):
                new_s.title.strRef.f = f"{q}!$B${excel_row}"
            elif getattr(new_s, "tx", None) and getattr(new_s.tx, "strRef", None):
                new_s.tx.strRef.f = f"{q}!$B${excel_row}"
            chart.series.append(new_s)
            added += 1
        except Exception:
            pass
    return added


def _fill_month_category_row(ws, excel_row: int, category_name: str, kind: str) -> None:
    template_row = excel_row - 1 if excel_row > 1 else None
    template_b = ws.cell(template_row, 2) if template_row else None
    cell_b = ws.cell(excel_row, 2)
    cell_b.value = category_name
    _apply_name_style(cell_b, kind, template_b)
    for col in range(3, 8):
        cell = ws.cell(excel_row, col)
        if not (isinstance(cell.value, str) and cell.value.startswith("=")):
            cell.value = None
        template_c = ws.cell(template_row, col) if template_row else None
        _apply_week_cell_style(cell, template_c)
    cell_h = ws.cell(excel_row, 8)
    cell_h.value = f"=SUM(C{excel_row}:G{excel_row})"
    template_h = ws.cell(template_row, 8) if template_row else None
    _apply_name_style(cell_h, kind, template_h)


def apply_category_kind_fills(
    ws,
    category_rows: Dict[str, int],
    *,
    emergency_categories: Optional[Iterable[str]] = None,
) -> int:
    """
    Заливка B и H + границы B–H по типу план/экстр.
    Границы берутся с соседней строки категории или тонкая чёрная.
    """
    emergency = {str(x).strip() for x in (emergency_categories or [])}
    rows_sorted = sorted(
        ((str(name).strip(), int(row)) for name, row in (category_rows or {}).items()),
        key=lambda kv: kv[1],
    )
    n = 0
    for i, (name, r) in enumerate(rows_sorted):
        kind = "emergency" if name in emergency else "plan"
        # шаблон — предыдущая категория, иначе следующая
        template_row = rows_sorted[i - 1][1] if i > 0 else (rows_sorted[i + 1][1] if i + 1 < len(rows_sorted) else None)
        for col in range(2, 9):
            cell = ws.cell(r, col)
            tmpl = ws.cell(template_row, col) if template_row else None
            if col in (2, 8):
                _apply_name_style(cell, kind, tmpl)
            else:
                _apply_week_cell_style(cell, tmpl)
        n += 1
    return n


def _fill_overview_row(
    ws,
    excel_row: int,
    category_name: str,
    kind: str,
    month_sheets: Sequence[str],
    *,
    with_months: bool,
) -> None:
    template_row = excel_row - 1 if excel_row > 1 else None
    template_b = ws.cell(template_row, 2) if template_row else None
    cell_b = ws.cell(excel_row, 2)
    cell_b.value = category_name
    _apply_name_style(cell_b, kind, template_b)
    cell_c = ws.cell(excel_row, 3)
    cell_c.value = _year_sum_h_formula(month_sheets, excel_row)
    template_c = ws.cell(template_row, 3) if template_row else None
    _apply_week_cell_style(cell_c, template_c)
    if with_months:
        # D… = ссылки на месяцы (как у соседней строки)
        # колонки 4..15 = 12 месяцев
        for i, sheet in enumerate(month_sheets):
            col = 4 + i
            if col > 15:
                break
            cell = ws.cell(excel_row, col)
            cell.value = f"={_quote_sheet(sheet)}!H{excel_row}"
            tmpl = ws.cell(template_row, col) if template_row else None
            _apply_week_cell_style(cell, tmpl)


def _classify_sheet(name: str) -> Optional[str]:
    key = str(name).strip().lower().replace("ё", "е")
    return OVERVIEW_ALIASES.get(key)


def add_category_row_to_summary(
    summary_path: Path | str,
    *,
    category_name: str,
    excel_row: int,
    form_line: str,
    sheet_names: Dict[int, str],
    form_cfg: Optional[dict] = None,
    kind: str = "plan",
    histology: bool = False,
    endoscopic: bool = False,
    anchor_row: Optional[int] = None,
    backup: bool = True,
    backup_keep: int = DEFAULT_BACKUP_KEEP,
) -> dict:
    """
    Физическая вставка строки excel_row на месячных листах, ОБЩАЯ и График Общий:
    сдвиг формул, цвета, итоги план/экстр, форма 4001, серии графиков.
    """
    path = Path(summary_path)
    if not path.exists():
        raise FileNotFoundError(path)

    insert_at = int(excel_row)
    report: dict = {
        "path": str(path),
        "excel_row": insert_at,
        "sheets": [],
        "overview": [],
        "backup": None,
        "formulas": {"n": 0, "r": 0, "endo": 0, "skipped": 0, "total": 0, "kind": 0, "bumped": 0},
        "charts": 0,
        "blank_inserted": 0,
        "blank_deleted": 0,
        "blank_delta": 0,
    }

    if backup:
        bak = make_backup(path, keep=backup_keep)
        report["backup"] = str(bak) if bak else None

    wb = openpyxl.load_workbook(path)
    form_cfg = form_cfg or {}
    month_sheets = [sheet_names[k] for k in sorted(sheet_names.keys(), key=int) if sheet_names[k] in wb.sheetnames]

    # --- месячные листы ---
    blank_delta_set = False
    for name in month_sheets:
        ws = wb[name]
        ws.insert_rows(insert_at, amount=1)
        report["formulas"]["bumped"] += bump_sheet_formulas(ws, insert_at, 1)
        report["formulas"]["bumped"] += bump_chart_refs(ws, insert_at, 1)
        _fill_month_category_row(ws, insert_at, category_name, kind)
        tot = patch_totals_and_kind_formulas(ws, excel_row=insert_at, kind=kind)
        report["formulas"]["total"] += tot["total"]
        report["formulas"]["kind"] += tot["kind"]
        patch = patch_form_formulas_on_sheet(
            ws,
            excel_row=insert_at,
            form_line=form_line,
            form_cfg=form_cfg,
            histology=histology,
            endoscopic=endoscopic,
        )
        for k in ("n", "r", "endo", "skipped"):
            report["formulas"][k] += patch.get(k, 0)
        report["charts"] += _add_month_chart_series(ws, insert_at)
        gap = ensure_one_blank_before_totals(ws)
        report["blank_inserted"] += gap["inserted"]
        report["blank_deleted"] += gap["deleted"]
        if not blank_delta_set:
            report["blank_delta"] = gap["delta"]
            blank_delta_set = True
        report["sheets"].append(name)

    # --- ОБЩАЯ / График Общий ---
    for sheet_name in list(wb.sheetnames):
        role = _classify_sheet(sheet_name)
        if not role:
            continue
        ws = wb[sheet_name]
        ws.insert_rows(insert_at, amount=1)
        report["formulas"]["bumped"] += bump_sheet_formulas(ws, insert_at, 1)
        report["formulas"]["bumped"] += bump_chart_refs(ws, insert_at, 1)
        _fill_overview_row(
            ws,
            insert_at,
            category_name,
            kind,
            month_sheets,
            with_months=(role == "chart"),
        )
        if role == "overview":
            op = patch_overview_form_c_refs(
                ws,
                excel_row=insert_at,
                form_line=form_line,
                form_cfg=form_cfg,
                histology=histology,
                endoscopic=endoscopic,
            )
            for k in ("n", "r", "endo"):
                report["formulas"][k] += op.get(k, 0)
        if role == "chart":
            report["charts"] += _add_overview_chart_series(ws, insert_at)
        gap = ensure_one_blank_before_totals(ws)
        report["blank_inserted"] += gap["inserted"]
        report["blank_deleted"] += gap["deleted"]
        if not blank_delta_set:
            report["blank_delta"] = gap["delta"]
            blank_delta_set = True
        report["overview"].append(sheet_name)

    wb.save(path)
    return report


def find_anchor_row(category_rows: dict, anchor_category: str) -> Optional[int]:
    if not anchor_category:
        return None
    if anchor_category in category_rows:
        return int(category_rows[anchor_category])
    needle = str(anchor_category).strip()
    for key, row in category_rows.items():
        if str(key).strip() == needle:
            return int(row)
    return None


def remove_col_ref_from_formula(formula: str, col: str, row: int) -> Tuple[str, bool]:
    """Убирает ColRow из SUM(C4:C38) или из списка C4+C5+C38."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula, False
    col_u = col.upper()
    row = int(row)
    orig = formula

    m = re.search(rf"SUM\(({col_u})(\d+):({col_u})(\d+)\)", formula, re.IGNORECASE)
    if m:
        start, end = int(m.group(2)), int(m.group(4))
        if row == end and end > start:
            formula = (
                formula[: m.start()]
                + f"SUM({col_u}{start}:{col_u}{end - 1})"
                + formula[m.end() :]
            )
        elif row == start and end > start:
            formula = (
                formula[: m.start()]
                + f"SUM({col_u}{start + 1}:{col_u}{end})"
                + formula[m.end() :]
            )

    formula = re.sub(rf"\+{col_u}{row}\b", "", formula, flags=re.IGNORECASE)
    formula = re.sub(rf"\b{col_u}{row}\+", "", formula, flags=re.IGNORECASE)
    formula = re.sub(r"=\+", "=", formula)
    formula = re.sub(r"\(\+", "(", formula)
    formula = re.sub(r"\+\+", "+", formula)
    formula = re.sub(r"\+\)", ")", formula)
    return formula, formula != orig


def _strip_row_refs_on_sheet(ws, excel_row: int) -> int:
    changed = 0
    max_col = min(ws.max_column or 30, 40)
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=max_col):
        for cell in row_cells:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            new_v = v
            for col_i in range(1, max_col + 1):
                new_v, _ch = remove_col_ref_from_formula(new_v, get_column_letter(col_i), excel_row)
            new_v, _ch = remove_h_ref(new_v, excel_row)
            new_v2, n = re.subn(
                rf"\+?((?:'[^']+'|[A-Za-zА-Яа-яЁё0-9_ ]+)!)[A-Z]{{1,3}}{excel_row}\b",
                "",
                new_v,
                flags=re.IGNORECASE,
            )
            if n:
                new_v = new_v2
            new_v = re.sub(r"=\+", "=", new_v)
            new_v = re.sub(r"\(\+", "(", new_v)
            new_v = re.sub(r"\+\+", "+", new_v)
            new_v = re.sub(r"\+\)", ")", new_v)
            if new_v != v:
                cell.value = new_v
                changed += 1
    return changed


def _remove_chart_series_for_row(ws, excel_row: int) -> int:
    removed = 0
    for chart in getattr(ws, "_charts", []) or []:
        keep = []
        for series in list(chart.series):
            rows: List[int] = []
            for attr in ("val", "cat"):
                src = getattr(series, attr, None)
                if src is None:
                    continue
                for ref_attr in ("numRef", "strRef"):
                    ref = getattr(src, ref_attr, None)
                    f = getattr(ref, "f", None) if ref is not None else None
                    if f:
                        rows.extend(
                            int(m.group(1)) for m in re.finditer(r"\$[A-Z]{1,3}\$(\d+)", f, re.I)
                        )
            title = getattr(series, "title", None)
            if title is not None and getattr(title, "strRef", None) and title.strRef.f:
                rows.extend(
                    int(m.group(1))
                    for m in re.finditer(r"\$[A-Z]{1,3}\$(\d+)", title.strRef.f, re.I)
                )
            if rows and all(r == excel_row for r in rows):
                removed += 1
                continue
            keep.append(series)
        if len(keep) != len(list(chart.series)) or removed:
            try:
                chart.series = keep
            except Exception:
                chart._series = keep  # type: ignore[attr-defined]
    return removed


def delete_category_row_from_summary(
    summary_path: Path | str,
    *,
    excel_row: int,
    sheet_names: Dict[int, str],
    backup: bool = True,
    backup_keep: int = DEFAULT_BACKUP_KEEP,
) -> dict:
    """Физическое удаление строки категории на месяцах / ОБЩАЯ / График Общий."""
    path = Path(summary_path)
    if not path.exists():
        raise FileNotFoundError(path)
    delete_at = int(excel_row)
    report: dict = {
        "path": str(path),
        "excel_row": delete_at,
        "sheets": [],
        "overview": [],
        "backup": None,
        "charts_removed": 0,
        "stripped": 0,
        "bumped": 0,
        "blank_inserted": 0,
        "blank_deleted": 0,
        "blank_delta": 0,
    }
    if backup:
        bak = make_backup(path, keep=backup_keep)
        report["backup"] = str(bak) if bak else None

    wb = openpyxl.load_workbook(path)
    month_sheets = [
        sheet_names[k]
        for k in sorted(sheet_names.keys(), key=int)
        if sheet_names[k] in wb.sheetnames
    ]
    targets = list(month_sheets)
    for sn in wb.sheetnames:
        if _classify_sheet(sn) and sn not in targets:
            targets.append(sn)

    blank_delta_set = False
    for name in targets:
        ws = wb[name]
        report["charts_removed"] += _remove_chart_series_for_row(ws, delete_at)
        report["stripped"] += _strip_row_refs_on_sheet(ws, delete_at)
        ws.delete_rows(delete_at, amount=1)
        report["bumped"] += bump_sheet_formulas(ws, delete_at + 1, -1)
        report["bumped"] += bump_chart_refs(ws, delete_at + 1, -1)
        gap = ensure_one_blank_before_totals(ws)
        report["blank_inserted"] += gap["inserted"]
        report["blank_deleted"] += gap["deleted"]
        if not blank_delta_set:
            report["blank_delta"] = gap["delta"]
            blank_delta_set = True
        if name in month_sheets:
            report["sheets"].append(name)
        else:
            report["overview"].append(name)

    wb.save(path)
    return report


def ensure_blank_separator_in_summary(
    summary_path: Path | str,
    *,
    sheet_names: Optional[Dict[int, str]] = None,
    backup: bool = True,
    backup_keep: int = DEFAULT_BACKUP_KEEP,
) -> dict:
    """
    На всех месячных / ОБЩАЯ / График Общий листах — ровно одна пустая строка
    между последней операцией и «Всего операций».
    """
    path = Path(summary_path)
    if not path.exists():
        raise FileNotFoundError(path)
    report: dict = {
        "path": str(path),
        "sheets": [],
        "overview": [],
        "backup": None,
        "blank_inserted": 0,
        "blank_deleted": 0,
        "blank_delta": 0,
    }
    if backup:
        bak = make_backup(path, keep=backup_keep)
        report["backup"] = str(bak) if bak else None

    wb = openpyxl.load_workbook(path)
    month_sheets: List[str] = []
    if sheet_names:
        month_sheets = [
            sheet_names[k]
            for k in sorted(sheet_names.keys(), key=int)
            if sheet_names[k] in wb.sheetnames
        ]
    targets = list(month_sheets)
    for sn in wb.sheetnames:
        if sn not in targets and (_classify_sheet(sn) or sn in month_sheets):
            if sn not in targets:
                targets.append(sn)
    # если sheet_names не передали — все листы, где есть «Всего операций»
    if not targets:
        for sn in wb.sheetnames:
            if _find_label_row(wb[sn], "Всего операций"):
                targets.append(sn)

    blank_delta_set = False
    for name in targets:
        ws = wb[name]
        gap = ensure_one_blank_before_totals(ws)
        if not gap["inserted"] and not gap["deleted"]:
            # всё равно считаем лист обработанным, если метка есть
            if not _find_label_row(ws, "Всего операций"):
                continue
        report["blank_inserted"] += gap["inserted"]
        report["blank_deleted"] += gap["deleted"]
        if not blank_delta_set:
            report["blank_delta"] = gap["delta"]
            blank_delta_set = True
        if name in month_sheets or (sheet_names and name in month_sheets):
            report["sheets"].append(name)
        elif _classify_sheet(name):
            report["overview"].append(name)
        else:
            report["sheets"].append(name)

    wb.save(path)
    return report
