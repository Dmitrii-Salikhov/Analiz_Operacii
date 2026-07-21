# analyzers/dept_template.py
"""Создание Excel-сводной для отделения (без блока формы 4001)."""
from __future__ import annotations

import io
import zipfile
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

from analyzers.dept_config import DEFAULT_SHEET_NAMES
from analyzers.summary_writer import MONTH_RU, compute_month_weeks

_EXCEL_EPOCH = date(1899, 12, 30)
THIN = Side(style="thin", color="FF000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Как в ЛОР (тема Office 2013+): accent6=салат 70AD47, accent2=оранж ED7D31
# theme index: accent1=4 … accent6=9
PLAN_FILL = PatternFill(patternType="solid", fgColor=Color(theme=9, tint=0.8))
EMERG_FILL = PatternFill(patternType="solid", fgColor=Color(theme=5, tint=0.8))
# Столбец «Итого» у блока итогов — accent5 (голубой), как H39+ в ЛОР
TOTALS_H_FILL = PatternFill(patternType="solid", fgColor=Color(theme=8, tint=0.8))

ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

COL_A_WIDTH = 13.0
COL_B_WIDTH = 48.0
COL_WEEK_WIDTH = 11.0
COL_H_WIDTH = 12.0

# Тема из ЛОР-сводной: иначе openpyxl ставит старую тему (accent6=оранж вместо салата)
_APP_DIR = Path(__file__).resolve().parent.parent
LOR_THEME_SOURCES = (
    _APP_DIR / "Операции сводная 2026.xlsx",
    _APP_DIR / "Операции сводная 2027.xlsx",
)


def apply_lor_workbook_theme(xlsx_path: str | Path, theme_source: Optional[Path] = None) -> bool:
    """
    Подменяет xl/theme/theme1.xml на тему из ЛОР-сводной,
    чтобы plan (accent6) был салатовым, emergency (accent2) — оранжевым.
    """
    dest = Path(xlsx_path)
    if not dest.exists():
        return False
    src = Path(theme_source) if theme_source else None
    if src is None or not src.exists():
        src = next((p for p in LOR_THEME_SOURCES if p.exists()), None)
    if src is None:
        return False
    try:
        with zipfile.ZipFile(src, "r") as zin:
            theme_xml = zin.read("xl/theme/theme1.xml")
    except KeyError:
        return False

    buf = io.BytesIO()
    with zipfile.ZipFile(dest, "r") as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/theme/theme1.xml":
                data = theme_xml
            zout.writestr(item, data)
    dest.write_bytes(buf.getvalue())
    return True


def _serial(d: date) -> int:
    return (d - _EXCEL_EPOCH).days


def _totals_layout(category_rows: Dict[str, int]) -> dict:
    if not category_rows:
        last_cat = 3
    else:
        last_cat = max(int(v) for v in category_rows.values())
    blank = last_cat + 1
    total = blank + 1
    return {
        "last_cat": last_cat,
        "blank": blank,
        "total": total,
        "emergency": total + 1,
        "plan": total + 2,
        "children": total + 4,
        "adults": total + 5,
        "patients": total + 6,
    }


def _sum_cells_expr(col_letter: str, rows: List[int]) -> Optional[str]:
    if not rows:
        return None
    if len(rows) == 1:
        return f"={col_letter}{rows[0]}"
    parts = "+".join(f"{col_letter}{r}" for r in rows)
    return f"={parts}"


def _apply_column_widths(ws) -> None:
    ws.column_dimensions["A"].width = COL_A_WIDTH
    ws.column_dimensions["B"].width = COL_B_WIDTH
    for col in range(3, 8):
        ws.column_dimensions[get_column_letter(col)].width = COL_WEEK_WIDTH
    ws.column_dimensions["H"].width = COL_H_WIDTH


def _write_week_header(ws, year: int, month: int) -> List[tuple]:
    weeks = compute_month_weeks(year, month)
    b2 = ws.cell(2, 2)
    b2.value = "Неделя с"
    b2.alignment = ALIGN_RIGHT
    b3 = ws.cell(3, 2)
    b3.value = "по"
    b3.alignment = ALIGN_RIGHT
    c2 = ws.cell(2, 3)
    c2.value = _serial(date(year, month, 1))
    c2.number_format = "DD/MM/YYYY"
    for i, (start, end) in enumerate(weeks):
        col = 3 + i
        ws.cell(2, col).value = _serial(start)
        ws.cell(2, col).number_format = "DD/MM/YYYY"
        ws.cell(3, col).value = _serial(end)
        ws.cell(3, col).number_format = "DD/MM/YYYY"
    ws.cell(2, 8).value = "Итого"
    ws.cell(2, 8).alignment = Alignment(horizontal="center", vertical="center")
    return weeks


def _write_category_block(
    ws,
    category_rows: Dict[str, int],
    plan_categories: List[str],
    emergency_categories: List[str],
    layout: dict,
) -> None:
    first = min(category_rows.values()) if category_rows else 4
    last = layout["last_cat"]
    emerg_set = set(emergency_categories or [])
    plan_set = set(plan_categories or [])

    for name, row in sorted(category_rows.items(), key=lambda x: x[1]):
        cell_b = ws.cell(row, 2)
        cell_b.value = name
        cell_b.alignment = ALIGN_LEFT_WRAP
        ws.cell(row, 8).value = f"=SUM(C{row}:G{row})"
        fill = EMERG_FILL if name in emerg_set else PLAN_FILL
        for col in range(2, 9):
            cell = ws.cell(row, col)
            cell.border = THIN_BORDER
            if col in (2, 8):
                cell.fill = fill

    ws.cell(layout["blank"], 2).value = None

    t = layout["total"]
    e_row = layout["emergency"]
    p_row = layout["plan"]
    ch = layout["children"]
    ad = layout["adults"]
    pa = layout["patients"]

    labels = [
        (t, "Всего операций", None),
        (e_row, "Экстренно операций", EMERG_FILL),
        (p_row, "План операций", PLAN_FILL),
        (ch, "Дети всего", None),
        (ad, "Взрослые", None),
        (pa, "Человек", None),
    ]
    for row, label, row_fill in labels:
        cell_b = ws.cell(row, 2)
        cell_b.value = label
        cell_b.alignment = ALIGN_LEFT
        for col in range(2, 9):
            cell = ws.cell(row, col)
            cell.border = THIN_BORDER
            if col == 2 and row_fill is not None:
                cell.fill = row_fill
            if col == 8:
                cell.fill = TOTALS_H_FILL

    for col in range(3, 8):
        letter = get_column_letter(col)
        ws.cell(t, col).value = f"=SUM({letter}{first}:{letter}{last})"
    ws.cell(t, 8).value = f"=SUM(C{t}:G{t})"

    emerg_rows = sorted({category_rows[c] for c in emerg_set if c in category_rows})
    plan_rows = sorted({category_rows[c] for c in plan_set if c in category_rows})
    if not plan_rows and not emerg_rows:
        plan_rows = list(range(first, last + 1))

    for col in range(3, 8):
        letter = get_column_letter(col)
        e_expr = _sum_cells_expr(letter, emerg_rows)
        p_expr = _sum_cells_expr(letter, plan_rows)
        ws.cell(e_row, col).value = e_expr if e_expr else 0
        ws.cell(p_row, col).value = p_expr if p_expr else 0

    ws.cell(e_row, 8).value = f"=SUM(C{e_row}:G{e_row})"
    ws.cell(p_row, 8).value = f"=SUM(C{p_row}:G{p_row})"

    for col in range(3, 8):
        letter = get_column_letter(col)
        ws.cell(ad, col).value = f"={letter}{pa}-{letter}{ch}"
    ws.cell(ch, 8).value = f"=SUM(C{ch}:G{ch})"
    ws.cell(ad, 8).value = f"=SUM(C{ad}:G{ad})"
    ws.cell(pa, 8).value = f"=SUM(C{pa}:G{pa})"


def _write_overview_sheet(
    wb,
    sheet_name: str,
    month_sheets: List[str],
    category_rows: dict,
    layout: dict,
    plan_categories: List[str],
    emergency_categories: List[str],
) -> None:
    ws = wb.create_sheet(sheet_name)
    emerg_set = set(emergency_categories or [])

    for name, row in sorted(category_rows.items(), key=lambda x: x[1]):
        cell_b = ws.cell(row, 2)
        cell_b.value = name
        cell_b.alignment = ALIGN_LEFT_WRAP
        parts = [f"{m}!H{row}" for m in month_sheets]
        ws.cell(row, 3).value = "=" + "+".join(parts) if parts else None
        ws.cell(row, 8).value = f"=C{row}"
        fill = EMERG_FILL if name in emerg_set else PLAN_FILL
        for col in (2, 3, 8):
            ws.cell(row, col).fill = fill
            ws.cell(row, col).border = THIN_BORDER

    ws.cell(layout["blank"], 2).value = None
    t = layout["total"]
    e_row = layout["emergency"]
    p_row = layout["plan"]
    ch = layout["children"]
    ad = layout["adults"]
    pa = layout["patients"]

    for row, label, row_fill in (
        (t, "Всего операций", None),
        (e_row, "Экстренно операций", EMERG_FILL),
        (p_row, "План операций", PLAN_FILL),
        (ch, "Дети всего", None),
        (ad, "Взрослые", None),
        (pa, "Человек", None),
    ):
        ws.cell(row, 2).value = label
        for col in range(2, 4):
            ws.cell(row, col).border = THIN_BORDER
        if row_fill is not None:
            ws.cell(row, 2).fill = row_fill
        ws.cell(row, 3).fill = TOTALS_H_FILL

    parts_total = [f"{m}!H{t}" for m in month_sheets]
    ws.cell(t, 3).value = "=" + "+".join(parts_total) if parts_total else None
    for label_row in (e_row, p_row, ch, pa):
        parts = [f"{m}!H{label_row}" for m in month_sheets]
        ws.cell(label_row, 3).value = "=" + "+".join(parts) if parts else None
    ws.cell(ad, 3).value = f"=C{pa}-C{ch}"
    _apply_column_widths(ws)


def create_department_summary(
    output_path: str | Path,
    *,
    department: str,
    year: int,
    category_rows: Dict[str, int],
    plan_categories: Optional[List[str]] = None,
    emergency_categories: Optional[List[str]] = None,
    sheet_names: Optional[Dict[int, str]] = None,
    totals_rows: Optional[dict] = None,
) -> Path:
    """
    Создаёт xlsx-сводную: 12 месяцев + ОБЩАЯ + График Общий.
    Блок формы 4001 не создаётся.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    names = sheet_names or DEFAULT_SHEET_NAMES
    plan = plan_categories if plan_categories is not None else list(category_rows.keys())
    emerg = emergency_categories or []
    layout = _totals_layout(category_rows)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    month_names: List[str] = []

    for month in range(1, 13):
        sname = names.get(month) or MONTH_RU.get(month, str(month)).capitalize()
        month_names.append(sname)
        ws = wb.create_sheet(sname)
        title = f"{department} за {MONTH_RU.get(month, '')} {year}".strip()
        # Заголовок в B1 (A остаётся узким, как в ЛОР)
        ws.cell(1, 1).value = None
        b1 = ws.cell(1, 2)
        b1.value = title
        b1.font = Font(bold=True)
        b1.alignment = ALIGN_LEFT
        _write_week_header(ws, year, month)
        _write_category_block(ws, category_rows, plan, emerg, layout)
        _apply_column_widths(ws)

    _write_overview_sheet(
        wb, "ОБЩАЯ", month_names, category_rows, layout, plan, emerg
    )
    chart = wb.create_sheet("График Общий")
    for name, row in sorted(category_rows.items(), key=lambda x: x[1]):
        chart.cell(row, 2).value = name
        chart.cell(row, 2).alignment = ALIGN_LEFT_WRAP
        parts = [f"{m}!H{row}" for m in month_names]
        chart.cell(row, 3).value = "=" + "+".join(parts) if parts else None
    chart.cell(layout["blank"], 2).value = None
    chart.cell(layout["total"], 2).value = "Всего операций"
    _apply_column_widths(chart)

    wb.save(out)
    # Тема ЛОР: иначе plan/emergency выглядят не салат/оранж
    apply_lor_workbook_theme(out)
    return out


def create_from_summary_cfg(
    output_path: str | Path,
    summary_cfg: dict,
    department: str,
) -> Path:
    category_rows = dict(summary_cfg.get("category_rows") or {})
    return create_department_summary(
        output_path,
        department=department,
        year=int(summary_cfg.get("year") or 2026),
        category_rows=category_rows,
        plan_categories=list(summary_cfg.get("plan_categories") or []),
        emergency_categories=list(summary_cfg.get("emergency_categories") or []),
        sheet_names={int(k): v for k, v in (summary_cfg.get("sheet_names") or {}).items()},
        totals_rows=summary_cfg.get("totals_rows"),
    )
