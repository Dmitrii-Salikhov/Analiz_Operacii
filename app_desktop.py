"""
Анализ операционной деятельности → «Операции сводная».
Простой UI на классическом tk (надёжно на macOS Tk 8.5).
"""
from __future__ import annotations

import logging
import os
import subprocess
import sys
import traceback
from datetime import datetime, timedelta
from pathlib import Path

os.environ.setdefault("TK_SILENCE_DEPRECATION", "1")

import openpyxl
import pandas as pd
import tkinter as tk
import yaml
from openpyxl.styles import Alignment, Font
from tkinter import BooleanVar, Menu, StringVar, filedialog, messagebox, ttk

from analyzers.app_log import AppLog
from analyzers.emk_compare import compare_plan_emergency, format_mismatch_report
from analyzers.emk_kind_classify import (
    apply_kind_to_summary_cfg,
    classify_categories_by_emk,
    disputed_to_dataframe,
    format_kind_report,
)
from analyzers.emk_loader import emk_department_stats, read_emk_stationary_report
from analyzers.export_report import export_month_like_summary
from analyzers.file_lock import excel_file_locked
from analyzers.form_4001 import compute_form_4001, form_4001_preview_rows
from analyzers.io_utils import OperationsStore, read_table
from analyzers.category_registry import (
    CategoryRegistryError,
    CategorySpec,
    default_anchor_category,
    register_category,
    save_config,
    shift_totals_rows_by_delta,
    suggest_keywords_from_name,
    unregister_category,
    update_category_keywords_file,
)
from analyzers.dept_config import (
    DEPT_REPORT_SOURCES,
    default_summary_filename,
    dept_summary_key,
    ensure_multi_dept_config,
    form_4001_enabled,
    get_summary_cfg,
    get_surgery_categories,
    is_lor_department,
    set_summary_cfg,
)
from analyzers.dept_inventory import (
    build_inventory_table,
    export_inventory_excel,
    inventory_from_source,
)
from analyzers.dept_template import create_from_summary_cfg
from analyzers.summary_layout import (
    add_category_row_to_summary,
    delete_category_row_from_summary,
    find_anchor_row,
)
from analyzers.ksg_catalog import get_catalog
from analyzers.problem_codes import build_problem_codes_table, format_config_draft
from analyzers.backup_utils import list_backups, restore_backup
from analyzers.surgery import (
    SurgeryAnalyzer,
    build_summary_tables,
    lookup_category_meta,
    reclassify_ops_by_keywords,
)
from analyzers.summary_writer import (
    MONTH_RU,
    SummaryWriter,
    _as_date,
    compute_month_weeks,
    read_sheet_weeks,
)
from analyzers.ui_settings import load_settings, save_settings
from analyzers.updater import (
    apply_update_from_zip,
    check_for_update,
    format_update_notes,
    read_local_version,
    resolve_token,
)
from analyzers.release_notes import format_whats_new
from analyzers.write_verify import format_verify_message, verify_write_report
from analyzers.year_template import create_year_summary, suggest_summary_path

try:
    from tkcalendar import Calendar
except ImportError:  # pragma: no cover
    Calendar = None  # type: ignore

APP_DIR = (
    Path(sys.executable).resolve().parent
    if getattr(sys, "frozen", False)
    else Path(__file__).resolve().parent
)
os.chdir(APP_DIR)
APP_LOG = AppLog(APP_DIR / "analysis.log", max_lines=500)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# Палитра светлой / тёмной темы и акцентов кнопок шапки
UI_THEMES = {
    "light": {
        "bg": "#F4F4F4",
        "fg": "#1A1A1A",
        "muted": "#555555",
        "panel": "#FFFFFF",
        "toolbar": "#EAEAEA",
        "entry_bg": "#FFFFFF",
        "entry_fg": "#1A1A1A",
        "btn_bg": "#E0E0E0",
        "btn_fg": "#1A1A1A",
        "btn_active": "#D0D0D0",
        "primary_bg": "#A5D6A7",
        "primary_fg": "#1B5E20",
        "primary_active": "#81C784",
        "emk_bg": "#FFCC80",
        "emk_fg": "#E65100",
        "emk_active": "#FFB74D",
        "status_bg": "#E8E8E8",
        "select_bg": "#BBDEFB",
        "heading_bg": "#E8E8E8",
        "heading_fg": "#1A1A1A",
        "tab_bg": "#E0E0E0",
        "tab_fg": "#1A1A1A",
        "tab_selected_bg": "#FFFFFF",
        "combo_bg": "#FFFFFF",
        "combo_fg": "#1A1A1A",
    },
    "dark": {
        "bg": "#2B2B2B",
        "fg": "#EDEDED",
        "muted": "#A0A0A0",
        "panel": "#333333",
        "toolbar": "#3A3A3A",
        "entry_bg": "#424242",
        "entry_fg": "#EDEDED",
        # кнопки шапки в тёмной теме: белый фон, чёрный текст
        "btn_bg": "#FFFFFF",
        "btn_fg": "#000000",
        "btn_active": "#F0F0F0",
        "primary_bg": "#FFFFFF",
        "primary_fg": "#000000",
        "primary_active": "#E8F5E9",
        "emk_bg": "#FFFFFF",
        "emk_fg": "#000000",
        "emk_active": "#FFE0B2",
        "status_bg": "#3A3A3A",
        "select_bg": "#1565C0",
        # шапка таблиц и вкладки-переключатели: чёрный текст
        "heading_bg": "#F0F0F0",
        "heading_fg": "#000000",
        "tab_bg": "#FFFFFF",
        "tab_fg": "#000000",
        "tab_selected_bg": "#FFFFFF",
        # отделение / месяц — белое поле, чёрный текст
        "combo_bg": "#FFFFFF",
        "combo_fg": "#000000",
    },
}


def _btn(parent, text, command, **pack):
    """Обычная tk.Button — всегда видна на macOS."""
    b = tk.Button(parent, text=text, command=command, padx=8, pady=4)
    b.pack(**pack)
    return b


class DesktopApp:
    def __init__(self, root):
        self.root = root
        self.app_version = read_local_version(APP_DIR)
        self.root.title(f"Сводная операций  v{self.app_version}")
        self.root.geometry("1200x780")
        self.root.minsize(900, 600)

        self.config = self.load_config()
        if self.config is None:
            root.destroy()
            return

        ensure_multi_dept_config(self.config)
        self.summary_key = "lor"
        self.summary_cfg = get_summary_cfg(self.config, summary_key="lor")
        self.df_emk = None
        self.emk_path = None
        self.store = OperationsStore()
        self.last_batch_span = (None, None)
        self.cat_table = None
        self.totals_df = None
        self.weeks = []
        self.last_emk_compare = None

        self.plan_mode = StringVar(value="template")
        self.hide_zeros = BooleanVar(value=False)
        self.filter_enabled = BooleanVar(value=False)
        default_summary = self.summary_cfg.get("default_path", "Операции сводная 2026.xlsx")
        self.summary_path = StringVar(value=str(APP_DIR / default_summary))
        self.dept_var = StringVar(value=self.config["departments"]["main"])
        self.preview_month = StringVar()
        self.status_var = StringVar(value="Готов к работе")
        self.start_date_var = StringVar(value="01.01.2026")
        self.end_date_var = StringVar(value="31.12.2026")
        self.year_var = StringVar(value=str(self.summary_cfg.get("year", 2026)))
        self._preview_clipboard = ""
        self.write_weeks_var = BooleanVar(value=True)
        self.write_form_var = BooleanVar(value=True)
        self.last_surg_dir = str(APP_DIR)
        self.last_emk_dir = str(APP_DIR)
        self._year_hint_key = None
        self.loaded_department = None
        self.summary_paths_by_dept = {}
        self.last_update_check = ""
        self.last_seen_version = ""
        self._prev_dept = None
        self.theme_var = StringVar(value="light")
        self._role_buttons: list = []
        self._theme_labels: list = []
        self._theme_frames: list = []
        self._theme_entries: list = []
        self._theme_checks: list = []
        self._theme_radios: list = []
        self._theme_texts: list = []

        self._apply_saved_settings()
        self._prev_dept = self.dept_var.get()
        self._sync_dept_context()

        self.kpi_ops_var = StringVar(value="—")
        self.kpi_patients_var = StringVar(value="—")
        self.kpi_plan_var = StringVar(value="—")
        self.kpi_emerg_var = StringVar(value="—")
        self.kpi_period_var = StringVar(value="—")
        self.kpi_files_var = StringVar(value="—")
        self.kpi_diff_var = StringVar(value="—")

        self._build_menu()
        self._build_layout()
        self._apply_theme()
        self._set_date_widgets_state("normal" if self.filter_enabled.get() else "disabled")
        self._bind_shortcuts()
        self._refresh_sources_list()
        self._load_log_into_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.log_message("Приложение готово. Нажмите «Опержурнал(ы)» для загрузки.")
        self.root.after(600, self._maybe_show_whats_new)
        self.root.after(900, self._maybe_startup_update_check)

    def load_config(self):
        try:
            with open(APP_DIR / "config.yaml", "r", encoding="utf-8") as f:
                cfg = yaml.safe_load(f)
            ensure_multi_dept_config(cfg)
            return cfg
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить config.yaml:\n{e}")
            logging.critical(f"Config load error: {e}")
            return None

    def _surgery_categories(self) -> list:
        return get_surgery_categories(self.config, summary_key=self.summary_key)

    def _sync_dept_context(self):
        """Активное отделение → summary_key и summary_cfg."""
        ensure_multi_dept_config(self.config)
        self.summary_key = dept_summary_key(self.config, self.dept_var.get())
        self.summary_cfg = get_summary_cfg(self.config, summary_key=self.summary_key)
        if not form_4001_enabled(self.summary_cfg):
            self.write_form_var.set(False)
        self._refresh_dept_hint()

    def _build_menu(self):
        bar = Menu(self.root)
        self.root.config(menu=bar)
        file_m = Menu(bar, tearoff=0)
        file_m.add_command(label="Опержурнал(ы)…", command=self.load_surg)
        file_m.add_command(label="Опержурналы из папки…", command=self.load_surg_folder)
        file_m.add_command(label="ЭМК…", command=self.load_emk)
        file_m.add_command(label="Сводная…", command=self.choose_summary)
        file_m.add_separator()
        file_m.add_command(label="Записать в Excel…", command=self.update_summary)
        file_m.add_command(label="Открыть Excel", command=self.open_summary_file)
        file_m.add_command(label="Восстановить из бэкапа…", command=self.restore_summary_backup)
        file_m.add_command(label="Экспорт простого отчёта…", command=self.export_simple)
        file_m.add_command(label="Создать сводную на год…", command=self.create_year_summary_dialog)
        file_m.add_command(label="Экспорт неклассифицированных…", command=self.export_unclassified)
        file_m.add_command(label="Экспорт проблемных кодов…", command=self.export_problem_codes)
        file_m.add_separator()
        file_m.add_command(label="Инвентаризация отделения…", command=self.inventory_department_dialog)
        file_m.add_command(label="Создать сводную для отделения…", command=self.create_dept_summary_dialog)
        file_m.add_command(
            label="План/экстр по ЭМК…", command=self.classify_kinds_from_emk_dialog
        )
        file_m.add_separator()
        file_m.add_command(label="Добавить операцию в отчёт…", command=self.add_category_dialog)
        file_m.add_command(label="Удалить операцию из отчёта…", command=self.delete_category_dialog)
        file_m.add_separator()
        file_m.add_command(label="Очистить", command=self.clear_store)
        file_m.add_command(label="Выход", command=self._on_close)
        bar.add_cascade(label="Файл", menu=file_m)
        help_m = Menu(bar, tearoff=0)
        help_m.add_command(label="Проверить обновления…", command=lambda: self.check_updates(force=True))
        help_m.add_command(label="Что нового…", command=lambda: self.show_whats_new(force=True))
        help_m.add_command(label="О программе", command=self.show_about)
        bar.add_cascade(label="Помощь", menu=help_m)

    def _bind_shortcuts(self):
        self.root.bind("<Command-o>", lambda e: self.load_surg())
        self.root.bind("<Control-o>", lambda e: self.load_surg())
        self.root.bind("<Command-s>", lambda e: self.update_summary())
        self.root.bind("<Control-s>", lambda e: self.update_summary())
        self.root.bind("<Command-c>", self._copy_focused_tree)
        self.root.bind("<Control-c>", self._copy_focused_tree)

    def _tool_btn(self, parent, text, command, role: str = "default", **pack):
        b = tk.Button(parent, text=text, command=command, padx=8, pady=4)
        self._role_buttons.append((b, role))
        b.pack(**pack)
        return b

    def _theme_label(self, parent, text, muted: bool = False, **kwargs):
        lbl = tk.Label(parent, text=text, **kwargs)
        self._theme_labels.append((lbl, muted))
        return lbl

    def _theme_toggle_label(self) -> str:
        return "Светлая" if self.theme_var.get() == "dark" else "Тёмная"

    def _toggle_theme(self):
        self.theme_var.set("dark" if self.theme_var.get() == "light" else "light")
        if getattr(self, "theme_btn", None) is not None:
            self.theme_btn.config(text=self._theme_toggle_label())
        self._apply_theme()
        self._persist_settings()

    def _current_theme(self) -> dict:
        key = self.theme_var.get() if self.theme_var.get() in UI_THEMES else "light"
        return UI_THEMES[key]

    def _apply_theme(self):
        t = self._current_theme()
        try:
            self.root.configure(bg=t["bg"])
        except tk.TclError:
            pass
        if getattr(self, "status_bar", None) is not None:
            self.status_bar.configure(bg=t["status_bg"], fg=t["fg"])

        for fr in self._theme_frames:
            try:
                if isinstance(fr, tk.LabelFrame):
                    fr.configure(bg=t["panel"], fg=t["fg"])
                else:
                    fr.configure(bg=t["bg"])
            except tk.TclError:
                pass

        # toolbar + два ряда — фон панели инструментов
        if self._theme_frames:
            for fr in self._theme_frames[:3]:
                try:
                    fr.configure(bg=t["toolbar"])
                except tk.TclError:
                    pass

        for lbl, muted in self._theme_labels:
            try:
                parent_bg = t["bg"]
                try:
                    pbg = lbl.master.cget("bg")
                    if pbg:
                        parent_bg = pbg
                except Exception:
                    pass
                lbl.configure(bg=parent_bg, fg=t["muted"] if muted else t["fg"])
            except tk.TclError:
                pass

        for b, role in self._role_buttons:
            try:
                if role == "primary":
                    b.configure(
                        bg=t["primary_bg"],
                        fg=t["primary_fg"],
                        activebackground=t["primary_active"],
                        activeforeground=t["primary_fg"],
                        highlightbackground=t["primary_bg"],
                    )
                elif role == "emk":
                    b.configure(
                        bg=t["emk_bg"],
                        fg=t["emk_fg"],
                        activebackground=t["emk_active"],
                        activeforeground=t["emk_fg"],
                        highlightbackground=t["emk_bg"],
                    )
                else:
                    b.configure(
                        bg=t["btn_bg"],
                        fg=t["btn_fg"],
                        activebackground=t["btn_active"],
                        activeforeground=t["btn_fg"],
                        highlightbackground=t["btn_bg"],
                    )
            except tk.TclError:
                pass

        # ttk notebook / combobox
        try:
            style = ttk.Style(self.root)
            try:
                style.theme_use("clam")
            except tk.TclError:
                pass
            style.configure(".", background=t["bg"], foreground=t["fg"], fieldbackground=t["entry_bg"])
            style.configure("TFrame", background=t["bg"])
            style.configure("TLabel", background=t["bg"], foreground=t["fg"])
            style.configure("TNotebook", background=t["bg"])
            style.configure(
                "TNotebook.Tab",
                background=t["tab_bg"],
                foreground=t["tab_fg"],
                padding=[8, 4],
            )
            style.map(
                "TNotebook.Tab",
                background=[("selected", t["tab_selected_bg"])],
                foreground=[("selected", t["tab_fg"])],
            )
            style.configure(
                "TCombobox",
                fieldbackground=t["combo_bg"],
                foreground=t["combo_fg"],
                background=t["combo_bg"],
                arrowcolor=t["combo_fg"],
            )
            style.map(
                "TCombobox",
                fieldbackground=[("readonly", t["combo_bg"]), ("!disabled", t["combo_bg"])],
                foreground=[("readonly", t["combo_fg"]), ("!disabled", t["combo_fg"])],
                selectbackground=[("readonly", t["combo_bg"])],
                selectforeground=[("readonly", t["combo_fg"])],
            )
            # выпадающий список combobox (macOS/clam)
            try:
                self.root.option_add("*TCombobox*Listbox.background", t["combo_bg"])
                self.root.option_add("*TCombobox*Listbox.foreground", t["combo_fg"])
                self.root.option_add("*TCombobox*Listbox.selectBackground", t["select_bg"])
                self.root.option_add("*TCombobox*Listbox.selectForeground", t["combo_fg"])
            except tk.TclError:
                pass
            # тело таблицы — по теме; шапка колонок — чёрный текст
            style.configure(
                "Treeview",
                background=t["panel"],
                foreground=t["fg"],
                fieldbackground=t["panel"],
            )
            style.configure(
                "Treeview.Heading",
                background=t["heading_bg"],
                foreground=t["heading_fg"],
            )
            style.map("Treeview", background=[("selected", t["select_bg"])])
            style.map("Treeview.Heading", foreground=[("active", t["heading_fg"])])
        except tk.TclError:
            pass

        # рекурсивно подкрасить основные контейнеры под notebook / settings
        self._paint_tk_tree(self.root, t, skip_buttons=True)
        self._apply_combo_caption_colors(t)

    def _apply_combo_caption_colors(self, t: dict) -> None:
        """Подписи «Отделение» / «Месяц» / «Год» — белые в тёмной теме (поля списков не трогаем)."""
        for attr in ("dept_caption", "month_caption", "year_caption"):
            lbl = getattr(self, attr, None)
            if lbl is None:
                continue
            try:
                bg = t["bg"]
                try:
                    bg = lbl.master.cget("bg") or bg
                except Exception:
                    pass
                lbl.configure(bg=bg, fg=t["fg"])
            except tk.TclError:
                pass

    def _paint_tk_tree(self, widget, t: dict, skip_buttons: bool = False):
        for child in widget.winfo_children():
            cls = child.winfo_class()
            try:
                if cls in ("Frame", "Labelframe", "Toplevel"):
                    if isinstance(child, tk.LabelFrame):
                        child.configure(bg=t["panel"], fg=t["fg"])
                    elif child not in self._theme_frames[:3]:
                        child.configure(bg=t["bg"])
                    self._paint_tk_tree(child, t, skip_buttons=skip_buttons)
                elif cls == "Label":
                    # уже в _theme_labels или нет — подогнать фон родителя
                    if not any(child is x for x, _ in self._theme_labels):
                        muted = str(child.cget("fg") or "") in ("#555", "#666", "#A0A0A0")
                        bg = t["bg"]
                        try:
                            bg = child.master.cget("bg") or bg
                        except Exception:
                            pass
                        child.configure(bg=bg, fg=t["muted"] if muted else t["fg"])
                elif cls == "Button" and not skip_buttons:
                    if not any(child is b for b, _ in self._role_buttons):
                        child.configure(
                            bg=t["btn_bg"],
                            fg=t["btn_fg"],
                            activebackground=t["btn_active"],
                            activeforeground=t["btn_fg"],
                        )
                elif cls == "Entry":
                    child.configure(
                        bg=t["entry_bg"],
                        fg=t["entry_fg"],
                        insertbackground=t["entry_fg"],
                    )
                elif cls in ("Checkbutton", "Radiobutton"):
                    bg = t["bg"]
                    try:
                        bg = child.master.cget("bg") or bg
                    except Exception:
                        pass
                    child.configure(
                        bg=bg,
                        fg=t["fg"],
                        activebackground=bg,
                        activeforeground=t["fg"],
                        selectcolor=t["panel"],
                    )
                elif cls == "Text":
                    child.configure(
                        bg=t["panel"],
                        fg=t["fg"],
                        insertbackground=t["fg"],
                    )
                elif cls in ("TFrame", "TNotebook", "Treeview"):
                    self._paint_tk_tree(child, t, skip_buttons=skip_buttons)
                else:
                    self._paint_tk_tree(child, t, skip_buttons=skip_buttons)
            except tk.TclError:
                try:
                    self._paint_tk_tree(child, t, skip_buttons=skip_buttons)
                except Exception:
                    pass

    def _build_layout(self):
        # ВАЖНО для macOS Tk: виджеты side=BOTTOM паковать ПЕРВЫМИ,
        # иначе expand-область «съедает» окно и оно выглядит пустым.
        self.status_bar = tk.Label(
            self.root, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor="w"
        )
        self.status_bar.pack(fill=tk.X, side=tk.BOTTOM)

        # --- шапка: два ряда кнопок ---
        toolbar = tk.Frame(self.root, bd=1, relief=tk.RAISED)
        toolbar.pack(fill=tk.X, padx=4, pady=4, side=tk.TOP)
        self._theme_frames.append(toolbar)

        row1 = tk.Frame(toolbar)
        row1.pack(fill=tk.X, padx=2, pady=(4, 2))
        self._theme_frames.append(row1)
        self._theme_label(row1, "Действия:", font=("Helvetica", 12, "bold")).pack(side=tk.LEFT, padx=6)
        # основные — зелёные
        self._tool_btn(row1, "Опержурнал(ы)", self.load_surg, role="primary", side=tk.LEFT, padx=3, pady=2)
        self._tool_btn(row1, "Из папки…", self.load_surg_folder, role="primary", side=tk.LEFT, padx=3, pady=2)
        self._tool_btn(row1, "Записать в Excel…", self.update_summary, role="primary", side=tk.LEFT, padx=3, pady=2)
        self._tool_btn(row1, "Открыть Excel", self.open_summary_file, role="primary", side=tk.LEFT, padx=3, pady=2)
        self._tool_btn(row1, "Обновить превью", self.run_analysis, side=tk.LEFT, padx=3, pady=2)
        self._tool_btn(row1, "Бэкап…", self.restore_summary_backup, side=tk.LEFT, padx=3, pady=2)
        self._tool_btn(row1, "Очистить", self.clear_store, side=tk.LEFT, padx=3, pady=2)

        row2 = tk.Frame(toolbar)
        row2.pack(fill=tk.X, padx=2, pady=(2, 4))
        self._theme_frames.append(row2)
        self._theme_label(row2, "Операции / ЭМК:", font=("Helvetica", 12, "bold")).pack(side=tk.LEFT, padx=6)
        self._tool_btn(row2, "Добавить операцию…", self.add_category_dialog, side=tk.LEFT, padx=3, pady=2)
        self._tool_btn(row2, "Удалить операцию…", self.delete_category_dialog, side=tk.LEFT, padx=3, pady=2)
        # ЭМК — светло-оранжевые
        self._tool_btn(row2, "ЭМК", self.load_emk, role="emk", side=tk.LEFT, padx=3, pady=2)
        self._tool_btn(row2, "Расхождения ЭМК", self.show_emk_diff, role="emk", side=tk.LEFT, padx=3, pady=2)
        self._theme_label(row2, "Тема:").pack(side=tk.LEFT, padx=(16, 4))
        self.theme_btn = self._tool_btn(
            row2, self._theme_toggle_label(), self._toggle_theme, side=tk.LEFT, padx=3, pady=2
        )

        # --- заголовок / отделение ---
        top = tk.Frame(self.root)
        top.pack(fill=tk.X, padx=8, pady=2)
        self._theme_frames.append(top)
        self._theme_label(top, "Сводная операционной деятельности", font=("Helvetica", 14, "bold")).pack(
            side=tk.LEFT
        )
        self._theme_label(
            top,
            f"  v{getattr(self, 'app_version', read_local_version(APP_DIR))}",
            font=("Helvetica", 12),
            muted=True,
        ).pack(side=tk.LEFT, padx=(4, 12))
        self.dept_caption = self._theme_label(top, "Отделение:")
        self.dept_caption.pack(side=tk.LEFT)
        dept_list = self.config["departments"]["list"]
        self.dept_combo = ttk.Combobox(top, textvariable=self.dept_var, values=dept_list, width=40, state="readonly")
        self.dept_combo.pack(side=tk.LEFT, padx=4)
        self.dept_combo.bind("<<ComboboxSelected>>", lambda e: self._on_department_changed())
        self.dept_hint_var = StringVar(value="")
        self.dept_hint_lbl = tk.Label(
            top, textvariable=self.dept_hint_var, wraplength=420, justify=tk.LEFT
        )
        self.dept_hint_lbl.pack(side=tk.LEFT, padx=(8, 0))
        self._theme_labels.append((self.dept_hint_lbl, True))
        self._refresh_dept_hint()

        # --- KPI ---
        kpi = tk.LabelFrame(self.root, text="Сводка", padx=6, pady=4)
        kpi.pack(fill=tk.X, padx=8, pady=4)
        self._theme_frames.append(kpi)
        self._kpi_title_labels = []
        self._kpi_value_labels = []
        for i, (title, var) in enumerate(
            (
                ("Операций", self.kpi_ops_var),
                ("Пациентов", self.kpi_patients_var),
                ("План %", self.kpi_plan_var),
                ("Экстр. %", self.kpi_emerg_var),
                ("Период", self.kpi_period_var),
                ("Файлы/ЭМК", self.kpi_files_var),
                ("Расхожд.", self.kpi_diff_var),
            )
        ):
            f = tk.Frame(kpi)
            f.grid(row=0, column=i, padx=8, sticky="w")
            self._theme_frames.append(f)
            lt = tk.Label(f, text=title)
            lt.pack(anchor="w")
            self._theme_labels.append((lt, True))
            lv = tk.Label(f, textvariable=var, font=("Helvetica", 14, "bold"))
            lv.pack(anchor="w")
            self._theme_labels.append((lv, False))

        # --- настройки ---
        opts = tk.LabelFrame(self.root, text="Настройки", padx=6, pady=4)
        opts.pack(fill=tk.X, padx=8, pady=2)

        r1 = tk.Frame(opts)
        r1.pack(fill=tk.X, pady=2)
        tk.Checkbutton(
            r1, text="Фильтр дат", variable=self.filter_enabled, command=self.toggle_date_widgets
        ).pack(side=tk.LEFT)
        tk.Label(r1, text="с").pack(side=tk.LEFT, padx=(8, 2))
        self.start_entry = tk.Entry(r1, textvariable=self.start_date_var, width=11, state="disabled")
        self.start_entry.pack(side=tk.LEFT)
        self.start_cal_btn = tk.Button(
            r1, text="…", width=2, command=lambda: self._pick_date(self.start_date_var), state="disabled"
        )
        self.start_cal_btn.pack(side=tk.LEFT, padx=(1, 0))
        tk.Label(r1, text="по").pack(side=tk.LEFT, padx=(6, 2))
        self.end_entry = tk.Entry(r1, textvariable=self.end_date_var, width=11, state="disabled")
        self.end_entry.pack(side=tk.LEFT)
        self.end_cal_btn = tk.Button(
            r1, text="…", width=2, command=lambda: self._pick_date(self.end_date_var), state="disabled"
        )
        self.end_cal_btn.pack(side=tk.LEFT, padx=(1, 0))
        tk.Checkbutton(
            r1, text="Скрыть нулевые", variable=self.hide_zeros, command=self.refresh_preview
        ).pack(side=tk.LEFT, padx=(12, 0))
        self.month_caption = tk.Label(r1, text="Месяц:")
        self.month_caption.pack(side=tk.LEFT, padx=(12, 2))
        self.month_combo = ttk.Combobox(r1, textvariable=self.preview_month, width=16, state="readonly")
        self.month_combo.pack(side=tk.LEFT)
        self.month_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_preview())
        self.year_caption = tk.Label(r1, text="Год:")
        self.year_caption.pack(side=tk.LEFT, padx=(12, 2))
        self.year_combo = ttk.Combobox(
            r1, textvariable=self.year_var, width=6, values=[str(y) for y in range(2024, 2036)]
        )
        self.year_combo.pack(side=tk.LEFT)
        self.year_combo.bind("<<ComboboxSelected>>", lambda e: self._on_year_changed())
        self.year_combo.bind("<Return>", lambda e: self._on_year_changed())

        r2 = tk.Frame(opts)
        r2.pack(fill=tk.X, pady=2)
        tk.Label(r2, text="План/экстренно:").pack(side=tk.LEFT)
        tk.Radiobutton(
            r2, text="По шаблону", variable=self.plan_mode, value="template", command=self._on_plan_mode_change
        ).pack(side=tk.LEFT)
        tk.Radiobutton(
            r2, text="Сверка с ЭМК", variable=self.plan_mode, value="emk", command=self.on_emk_mode
        ).pack(side=tk.LEFT, padx=8)

        r3 = tk.Frame(opts)
        r3.pack(fill=tk.X, pady=2)
        tk.Label(r3, text="Файл сводной:").pack(side=tk.LEFT)
        tk.Entry(r3, textvariable=self.summary_path).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)
        _btn(r3, "Обзор…", self.choose_summary, side=tk.LEFT)
        _btn(r3, "Сводная на год…", self.create_year_summary_dialog, side=tk.LEFT, padx=(4, 0))

        # --- тело: источники + вкладки ---
        body = tk.Frame(self.root)
        body.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)

        left = tk.LabelFrame(body, text="Источники", padx=4, pady=4)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 6))
        # Text с переносом — Listbox не умеет wrap длинных имён
        src_wrap = tk.Frame(left)
        src_wrap.pack(fill=tk.BOTH, expand=True)
        self.sources_text = tk.Text(
            src_wrap, width=36, height=18, wrap=tk.WORD, cursor="arrow", relief=tk.SUNKEN, bd=1
        )
        src_vs = ttk.Scrollbar(src_wrap, orient=tk.VERTICAL, command=self.sources_text.yview)
        self.sources_text.configure(yscrollcommand=src_vs.set)
        self.sources_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        src_vs.pack(side=tk.RIGHT, fill=tk.Y)
        self.sources_text.bind("<Double-Button-1>", self._on_source_dblclick)
        self.sources_text.configure(state=tk.DISABLED)
        self.emk_status = tk.Label(left, text="ЭМК: не загружен", anchor="w", wraplength=260, justify=tk.LEFT)
        self.emk_status.pack(fill=tk.X, pady=4)

        right = tk.Frame(body)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.notebook = ttk.Notebook(right)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        self.tab_preview_cat = tk.Frame(self.notebook)
        self.tab_preview_tot = tk.Frame(self.notebook)
        self.tab_preview_form = tk.Frame(self.notebook)
        self.tab_preview = self.tab_preview_cat
        self.tab_emk = tk.Frame(self.notebook)
        self.tab_uncl = tk.Frame(self.notebook)
        self.tab_dispute = tk.Frame(self.notebook)
        self.tab_log = tk.Frame(self.notebook)
        self.notebook.add(self.tab_preview_cat, text="Превью: категории")
        self.notebook.add(self.tab_preview_tot, text="Превью: итоги")
        self.notebook.add(self.tab_preview_form, text="Превью: форма 4001")
        self.notebook.add(self.tab_emk, text="Расхождения ЭМК")
        self.notebook.add(self.tab_uncl, text="Не классифицировано")
        self.notebook.add(self.tab_dispute, text="Спорные")
        self.notebook.add(self.tab_log, text="Журнал")

        pbtns = tk.Frame(self.tab_preview_cat)
        pbtns.pack(fill=tk.X, pady=2)
        _btn(pbtns, "Записать в Excel…", self.update_summary, side=tk.LEFT, padx=2)
        _btn(pbtns, "Экспорт простого отчёта", self.export_simple, side=tk.LEFT, padx=2)
        _btn(pbtns, "Копировать превью", self.copy_preview, side=tk.LEFT, padx=2)
        self.preview_info = StringVar(value="Превью: нет данных — загрузите опержурнал")
        tk.Label(self.tab_preview_cat, textvariable=self.preview_info, anchor="w").pack(fill=tk.X, padx=2)

        self.tree_preview_cat = self._make_tree(self.tab_preview_cat)
        self.tree_preview_tot = self._make_tree(self.tab_preview_tot)
        form_btns = tk.Frame(self.tab_preview_form)
        form_btns.pack(fill=tk.X, pady=2)
        _btn(form_btns, "Записать в Excel…", self.update_summary, side=tk.LEFT, padx=2)
        _btn(form_btns, "Копировать превью", self.copy_preview, side=tk.LEFT, padx=2)
        self.tree_form = self._make_tree(self.tab_preview_form)

        emk_top = tk.Frame(self.tab_emk)
        emk_top.pack(fill=tk.X, pady=2)
        _btn(emk_top, "Обновить сверку", self.show_emk_diff, side=tk.LEFT, padx=2)
        _btn(emk_top, "Экспорт таблицы…", self.export_emk_mismatches, side=tk.LEFT, padx=2)
        self.emk_info = StringVar(value="Загрузите ЭМК и опержурнал, затем обновите сверку")
        tk.Label(self.tab_emk, textvariable=self.emk_info, anchor="w").pack(fill=tk.X, padx=2)
        emk_cols = ("Дата", "КВС", "Категория", "Код", "Шаблон", "ЭМК", "Диагноз", "Услуга")
        emk_body = tk.Frame(self.tab_emk)
        emk_body.pack(fill=tk.BOTH, expand=True)
        self.tree_emk = ttk.Treeview(emk_body, columns=emk_cols, show="headings")
        for c, w in zip(emk_cols, (90, 90, 160, 100, 90, 90, 200, 220)):
            self.tree_emk.heading(c, text=c)
            self.tree_emk.column(c, width=w, anchor="w")
        emk_vs = ttk.Scrollbar(emk_body, orient=tk.VERTICAL, command=self.tree_emk.yview)
        self.tree_emk.configure(yscrollcommand=emk_vs.set)
        self.tree_emk.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        emk_vs.pack(side=tk.RIGHT, fill=tk.Y)

        uncl_top = tk.Frame(self.tab_uncl)
        uncl_top.pack(fill=tk.X, pady=2)
        _btn(uncl_top, "Экспорт списка…", self.export_unclassified, side=tk.LEFT, padx=2)
        _btn(uncl_top, "Экспорт проблемных кодов…", self.export_problem_codes, side=tk.LEFT, padx=2)
        _btn(uncl_top, "Править ключи…", self.edit_keywords_dialog, side=tk.LEFT, padx=2)
        uncl_cols = ("Дата", "КВС", "Код", "Название КСГ", "КСГ", "Услуга")
        uncl_body = tk.Frame(self.tab_uncl)
        uncl_body.pack(fill=tk.BOTH, expand=True)
        self.tree_uncl = ttk.Treeview(uncl_body, columns=uncl_cols, show="headings")
        for c, w in zip(uncl_cols, (90, 90, 110, 220, 120, 320)):
            self.tree_uncl.heading(c, text=c)
            self.tree_uncl.column(c, width=w, anchor="w")
        vs = ttk.Scrollbar(uncl_body, orient=tk.VERTICAL, command=self.tree_uncl.yview)
        self.tree_uncl.configure(yscrollcommand=vs.set)
        self.tree_uncl.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vs.pack(side=tk.RIGHT, fill=tk.Y)

        disp_top = tk.Frame(self.tab_dispute)
        disp_top.pack(fill=tk.X, pady=2)
        _btn(disp_top, "Назначить категорию…", self.assign_disputed_category_dialog, side=tk.LEFT, padx=2)
        _btn(disp_top, "Править ключи…", self.edit_keywords_dialog_from_dispute, side=tk.LEFT, padx=2)
        self.dispute_info = StringVar(
            value="Операции, где несколько категорий набрали одинаковый счёт по ключам"
        )
        tk.Label(self.tab_dispute, textvariable=self.dispute_info, anchor="w", fg="#555").pack(
            fill=tk.X, padx=2
        )
        disp_cols = ("Дата", "КВС", "Код", "Услуга", "Категория", "Кандидаты")
        disp_body = tk.Frame(self.tab_dispute)
        disp_body.pack(fill=tk.BOTH, expand=True)
        self.tree_dispute = ttk.Treeview(disp_body, columns=disp_cols, show="headings")
        for c, w in zip(disp_cols, (90, 90, 110, 280, 140, 220)):
            self.tree_dispute.heading(c, text=c)
            self.tree_dispute.column(c, width=w, anchor="w")
        dvs = ttk.Scrollbar(disp_body, orient=tk.VERTICAL, command=self.tree_dispute.yview)
        self.tree_dispute.configure(yscrollcommand=dvs.set)
        self.tree_dispute.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        dvs.pack(side=tk.RIGHT, fill=tk.Y)

        log_top = tk.Frame(self.tab_log)
        log_top.pack(fill=tk.X, pady=2)
        _btn(log_top, "Обновить", self._load_log_into_ui, side=tk.LEFT, padx=2)
        _btn(log_top, "Открыть файл", self._open_log_file, side=tk.LEFT, padx=2)
        _btn(log_top, "Очистить журнал", self._clear_log_file, side=tk.LEFT, padx=2)
        log_wrap = tk.Frame(self.tab_log)
        log_wrap.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(log_wrap, height=16, wrap=tk.WORD)
        log_vs = ttk.Scrollbar(log_wrap, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_vs.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_vs.pack(side=tk.RIGHT, fill=tk.Y)

    def _make_tree(self, parent):
        wrap = tk.Frame(parent)
        wrap.pack(fill=tk.BOTH, expand=True)
        # Колонки задаём безопасными id (без точек) — иначе ttk на macOS рисует пусто.
        tree = ttk.Treeview(wrap, columns=("c0",), show="headings")
        tree.heading("c0", text="")
        tree.column("c0", width=120)
        vs = ttk.Scrollbar(wrap, orient=tk.VERTICAL, command=tree.yview)
        hs = ttk.Scrollbar(wrap, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vs.grid(row=0, column=1, sticky="ns")
        hs.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)
        tree.tag_configure("zero", foreground="#888888")
        tree.tag_configure("total", font=("Helvetica", 11, "bold"))
        wrap._tree = tree  # noqa: SLF001 — ссылка для пересоздания при необходимости
        return tree

    def _configure_tree_columns(self, tree, col_ids, headings, widths=None):
        """col_ids без точек/спецсимволов; headings — подписи."""
        tree.configure(columns=col_ids, displaycolumns=col_ids)
        for i, cid in enumerate(col_ids):
            tree.heading(cid, text=headings[i])
            w = 220 if i == 0 else 90
            if widths and i < len(widths):
                w = widths[i]
            tree.column(cid, width=w, anchor="w" if i == 0 else "center", stretch=True)

    def _set_date_widgets_state(self, state: str):
        for w in (self.start_entry, self.end_entry, self.start_cal_btn, self.end_cal_btn):
            try:
                w.config(state=state)
            except tk.TclError:
                pass

    def toggle_date_widgets(self):
        self._set_date_widgets_state("normal" if self.filter_enabled.get() else "disabled")
        if not self.store.ops.empty:
            self.run_analysis()

    def _parse_date(self, text: str):
        text = (text or "").strip()
        for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Неверная дата: {text}")

    def _pick_date(self, variable: StringVar):
        """Календарь в отдельном окне — без DateEntry (он зависает на macOS Tk)."""
        if Calendar is None:
            messagebox.showinfo(
                "Календарь",
                "Пакет tkcalendar не установлен.\nВведите дату вручную: ДД.ММ.ГГГГ",
            )
            return
        try:
            current = self._parse_date(variable.get())
        except Exception:
            current = datetime.now().date()

        top = tk.Toplevel(self.root)
        top.title("Выбор даты")
        top.transient(self.root)
        top.resizable(False, False)
        # Тёмный фон: белые цифры, выходные жирным, выбранный день — голубым.
        cal = Calendar(
            top,
            selectmode="day",
            year=current.year,
            month=current.month,
            day=current.day,
            date_pattern="dd.mm.yyyy",
            background="#1e1e1e",
            foreground="white",
            bordercolor="#1e1e1e",
            headersbackground="#2a2a2a",
            headersforeground="white",
            normalbackground="#1e1e1e",
            normalforeground="white",
            weekendbackground="#1e1e1e",
            weekendforeground="white",
            othermonthbackground="#1e1e1e",
            othermonthforeground="#888888",
            othermonthwebackground="#1e1e1e",
            othermonthweforeground="#888888",
            selectbackground="#1a73e8",
            selectforeground="#7ec8ff",
            font=("Helvetica", 13),
        )
        try:
            p = cal._style_prefixe
            day_font = ("Helvetica", 13)
            we_font = ("Helvetica", 13, "bold")
            cal.style.configure(f"normal.{p}.TLabel", foreground="white", font=day_font)
            cal.style.configure(f"we.{p}.TLabel", foreground="white", font=we_font)
            cal.style.configure(f"normal_om.{p}.TLabel", foreground="#888888", font=day_font)
            cal.style.configure(f"we_om.{p}.TLabel", foreground="#888888", font=we_font)
            cal.style.configure(f"headers.{p}.TLabel", foreground="white")
            cal.style.configure(f"main.{p}.TLabel", foreground="white")
            # На aqua фон sel часто не виден — выделяем цветом текста.
            cal.style.configure(f"sel.{p}.TLabel", foreground="#7ec8ff", font=we_font)
        except Exception:
            pass
        cal.pack(padx=8, pady=8)
        # Повторно вешаем клик: на части macOS Tk событие <1> у ttk.Label теряется.
        for row in cal._calendar:
            for label in row:
                label.bind("<Button-1>", cal._on_click)

        applied = {"done": False}

        def apply_and_close(_event=None):
            if applied["done"]:
                return "break"
            applied["done"] = True
            try:
                d = cal.selection_get()
            except Exception:
                d = None
            if d is None:
                d = current
            variable.set(d.strftime("%d.%m.%Y"))
            try:
                top.destroy()
            except tk.TclError:
                pass
            if self.filter_enabled.get() and not self.store.ops.empty:
                self.run_analysis()
            return "break"

        # Клик по дню сразу выбирает дату (без обязательного OK).
        cal.bind("<<CalendarSelected>>", apply_and_close)

        bf = tk.Frame(top)
        bf.pack(pady=(0, 8))
        _btn(bf, "OK", apply_and_close, side=tk.LEFT, padx=4)
        _btn(bf, "Отмена", top.destroy, side=tk.LEFT, padx=4)
        top.protocol("WM_DELETE_WINDOW", top.destroy)
        # Без grab_set — на macOS он часто блокирует клики по дням.
        top.focus_force()
        top.lift()

    def _summary_dir(self) -> Path:
        """Папка текущей сводной (любая; не обязана быть рядом с программой)."""
        p = Path(self.summary_path.get().strip() or "")
        if p.is_file() or p.suffix.lower() in {".xlsx", ".xlsm"}:
            return p.parent if str(p.parent) not in ("", ".") else APP_DIR
        if p.is_dir():
            return p
        return APP_DIR

    def _on_year_changed(self):
        try:
            year = int(self.year_var.get())
        except ValueError:
            return
        self.summary_cfg["year"] = year
        suggested = suggest_summary_path(self._summary_dir(), year)
        if suggested.exists():
            self.summary_path.set(str(suggested))
            self.log_message(f"Год {year}: сводная {suggested}")
        else:
            self.log_message(
                f"Год {year}: файла {suggested.name} нет в {suggested.parent}.",
                level="WARNING",
            )
            if messagebox.askyesno(
                "Нет файла сводной",
                f"Файла за {year} нет:\n{suggested}\n\n"
                "Создать сводную на этот год из текущего шаблона?",
            ):
                self.create_year_summary_dialog()
                return
        if not self.store.ops.empty:
            self.run_analysis()

    def _summary_file_year(self) -> int | None:
        path = self.summary_path.get().strip()
        if not path or not os.path.exists(path):
            return None
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
            try:
                names = self.summary_cfg.get("sheet_names") or {}
                for m in range(1, 13):
                    sheet = names.get(m) or names.get(str(m))
                    if not sheet or sheet not in wb.sheetnames:
                        continue
                    d = _as_date(wb[sheet]["C2"].value)
                    if d and d.year >= 2000:
                        return int(d.year)
            finally:
                wb.close()
        except Exception:
            return None
        return None

    def _maybe_hint_year_mismatch(self, ops: pd.DataFrame | None = None):
        """Подсказка: данные/файл сводной и год в настройках не совпадают."""
        if ops is None:
            ops = self.store.ops
        if ops is None or getattr(ops, "empty", True):
            return
        try:
            cfg_year = int(self.year_var.get())
        except ValueError:
            return
        op_years = sorted(
            {int(y) for y in pd.to_datetime(ops["Дата"], errors="coerce").dt.year.dropna().unique()}
        )
        if not op_years:
            return
        data_year = int(max(op_years))
        file_year = self._summary_file_year()
        key = (tuple(op_years), cfg_year, file_year)
        if key == self._year_hint_key:
            return
        problems = []
        if data_year != cfg_year:
            problems.append(f"в журнале год {data_year}, в настройках — {cfg_year}")
        if file_year is not None and file_year != data_year:
            problems.append(f"файл сводной на {file_year}, данные — {data_year}")
        if not problems:
            return
        self._year_hint_key = key
        msg = "Несовпадение года:\n• " + "\n• ".join(problems)
        msg += f"\n\nСоздать/переключить сводную на {data_year}?"
        if messagebox.askyesno("Подсказка: год", msg):
            self.year_var.set(str(data_year))
            self.summary_cfg["year"] = data_year
            suggested = suggest_summary_path(self._summary_dir(), data_year)
            if suggested.exists():
                self.summary_path.set(str(suggested))
                self._persist_settings()
                self.log_message(f"Переключено на сводную {suggested}")
                self.run_analysis()
            else:
                self.create_year_summary_dialog()

    def _tree_to_tsv(self, tree) -> str:
        cols = list(tree["columns"] or ())
        if not cols:
            return ""
        headers = [tree.heading(c)["text"] for c in cols]
        lines = ["\t".join(headers)]
        for item in tree.get_children():
            vals = tree.item(item).get("values") or []
            lines.append("\t".join("" if v is None else str(v) for v in vals))
        return "\n".join(lines)

    def copy_preview(self):
        parts = []
        t_cat = self._tree_to_tsv(getattr(self, "tree_preview_cat", None))
        t_tot = self._tree_to_tsv(getattr(self, "tree_preview_tot", None))
        t_form = self._tree_to_tsv(getattr(self, "tree_form", None))
        if t_cat:
            parts.append("Категории по неделям\n" + t_cat)
        if t_tot:
            parts.append("Итоги по неделям\n" + t_tot)
        if t_form:
            parts.append("Форма 4001\n" + t_form)
        text = "\n\n".join(parts)
        if not text.strip():
            messagebox.showinfo("Копирование", "Нечего копировать — сначала обновите превью")
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.root.update_idletasks()
        self._preview_clipboard = text
        self.status_var.set("Превью скопировано в буфер")
        self.log_message("Превью скопировано в буфер обмена")

    def _copy_focused_tree(self, event=None):
        # Если активна вкладка превью — копируем всё превью одной пачкой.
        try:
            tab_text = self.notebook.tab(self.notebook.select(), "text")  # type: ignore[arg-type]
        except Exception:
            tab_text = ""
        if str(tab_text).startswith("Превью"):
            self.copy_preview()
            return "break"

        w = self.root.focus_get()
        for t in (
            getattr(self, "tree_preview_cat", None),
            getattr(self, "tree_preview_tot", None),
            getattr(self, "tree_form", None),
            self.tree_uncl,
        ):
            if t is None:
                continue
            if str(w).startswith(str(t)):
                text = self._tree_to_tsv(t)
                if text:
                    self.root.clipboard_clear()
                    self.root.clipboard_append(text)
                    self.status_var.set("Таблица скопирована")
                return "break"
        return None

    def _load_log_into_ui(self):
        """При открытии — показать хвост analysis.log и прокрутить к последнему событию."""
        self.log_text.delete("1.0", tk.END)
        lines = APP_LOG.read_lines()
        if lines:
            self.log_text.insert(tk.END, "\n".join(lines) + "\n")
            self.log_text.insert(tk.END, "—" * 40 + "\n")
        self.log_text.see(tk.END)
        self.root.after(100, lambda: self.log_text.see(tk.END))

    def log_message(self, msg: str, tag: str = None, level: str = "INFO"):
        line = APP_LOG.append(msg, level=level)
        if hasattr(self, "log_text"):
            self.log_text.insert(tk.END, line + "\n")
            self.log_text.see(tk.END)
        logging.log(getattr(logging, level, logging.INFO), msg)

    def _on_tab_changed(self, _event=None):
        try:
            if self.notebook.index(self.notebook.select()) == self.notebook.index(self.tab_log):
                self.log_text.see(tk.END)
        except Exception:
            pass

    def _busy(self, on: bool):
        self.root.config(cursor="watch" if on else "")
        self.root.update_idletasks()

    def _set_kpis_empty(self):
        for v in (
            self.kpi_ops_var,
            self.kpi_patients_var,
            self.kpi_plan_var,
            self.kpi_emerg_var,
            self.kpi_period_var,
            self.kpi_files_var,
            self.kpi_diff_var,
        ):
            v.set("—")

    def _refresh_sources_list(self):
        self.sources_text.configure(state=tk.NORMAL)
        self.sources_text.delete("1.0", tk.END)
        if self.store.ops.empty:
            self.sources_text.insert(tk.END, "Нет загруженных журналов")
        else:
            self.store.refresh_source_meta()
            blocks = []
            for src in self.store.sources:
                meta = self.store.source_meta.get(src) or {}
                d0 = meta.get("date_from")
                d1 = meta.get("date_to")
                n = meta.get("count", 0)
                if d0 is not None and d1 is not None:
                    period = f"{d0.strftime('%d.%m.%Y')} – {d1.strftime('%d.%m.%Y')}"
                    blocks.append(f"• {src}\n  {period}\n  операций: {n}")
                else:
                    blocks.append(f"• {src}\n  операций: {n}")
            self.sources_text.insert(tk.END, "\n\n".join(blocks))
        self.sources_text.configure(state=tk.DISABLED)
        if self.df_emk is not None:
            name = os.path.basename(self.emk_path) if self.emk_path else "загружен"
            self.emk_status.config(text=f"ЭМК: {name} ({len(self.df_emk)} стр.)")
        else:
            self.emk_status.config(text="ЭМК: не загружен")
        self.kpi_files_var.set(f"{len(self.store.sources)} / {'да' if self.df_emk is not None else 'нет'}")

    def _on_source_dblclick(self, _e=None):
        # строка под курсором
        try:
            idx = self.sources_text.index(f"@{_e.x},{_e.y}") if _e else self.sources_text.index(tk.INSERT)
            line = self.sources_text.get(f"{idx} linestart", f"{idx} lineend").strip()
        except Exception:
            line = ""
        if line:
            self.log_message(f"Источник: {line}")
            self.notebook.select(self.tab_log)

    def _on_plan_mode_change(self):
        if self.plan_mode.get() == "emk":
            self.on_emk_mode()
        elif not self.store.ops.empty:
            self.run_analysis()

    def _weeks_for_month(self, year: int, month: int):
        path = self.summary_path.get().strip()
        sheet_names = self.summary_cfg.get("sheet_names", {})
        sheet = sheet_names.get(month) or sheet_names.get(str(month))
        if path and os.path.exists(path) and sheet:
            try:
                wb = openpyxl.load_workbook(path, data_only=False)
                if sheet in wb.sheetnames:
                    weeks = read_sheet_weeks(wb[sheet])
                    wb.close()
                    if weeks:
                        return weeks
                wb.close()
            except Exception:
                pass
        return compute_month_weeks(year, month)

    def choose_summary(self):
        initial = self._summary_dir()
        path = filedialog.askopenfilename(
            initialdir=str(initial),
            filetypes=[("Excel", "*.xlsx")],
            title="Выберите файл сводной",
        )
        if path:
            self.summary_path.set(path)
            self._persist_settings()
            self.log_message(f"Сводная: {path}")

    def clear_store(self):
        if messagebox.askyesno("Очистка", "Удалить накопленные операции?"):
            self.store.clear()
            self.cat_table = self.totals_df = None
            self.weeks = []
            self.loaded_department = None
            self.last_emk_compare = None
            self._refresh_sources_list()
            self._set_kpis_empty()
            self._clear_trees()
            if hasattr(self, "tab_emk"):
                self.notebook.tab(self.tab_emk, text="Расхождения ЭМК")
                self.emk_info.set("Загрузите ЭМК и опержурнал, затем обновите сверку")
            if hasattr(self, "tab_uncl"):
                self.notebook.tab(self.tab_uncl, text="Не классифицировано")
            if hasattr(self, "tab_dispute"):
                self.notebook.tab(self.tab_dispute, text="Спорные")
            self.log_message("Накопитель очищен")
            self.status_var.set("Очищено")

    def _clear_trees(self):
        for tree in (
            getattr(self, "tree_preview_cat", None),
            getattr(self, "tree_preview_tot", None),
            getattr(self, "tree_form", None),
            getattr(self, "tree_emk", None),
            getattr(self, "tree_uncl", None),
            getattr(self, "tree_dispute", None),
        ):
            if tree is None:
                continue
            for item in tree.get_children():
                tree.delete(item)

    def load_emk(self):
        path = filedialog.askopenfilename(
            initialdir=self.last_emk_dir,
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("Excel", "*.xlsx *.xls"), ("CSV", "*.csv")],
            title="ЭМК — отчёт по заполнению в стационаре",
        )
        if not path:
            return
        self.last_emk_dir = str(Path(path).parent)
        self._persist_settings()
        try:
            self._busy(True)
            if str(path).lower().endswith(".csv"):
                self.df_emk = read_table(path)
            else:
                self.df_emk = read_emk_stationary_report(path)
            self.emk_path = path
            self._refresh_sources_list()
            stats = emk_department_stats(self.df_emk)
            dept = self.dept_var.get()
            dept_rows = stats.get(dept)
            if dept_rows is None:
                # частичное совпадение по подстроке
                dept_rows = sum(v for k, v in stats.items() if dept and dept in str(k))
            extra = f", отделение «{dept}»: {dept_rows} стр." if dept_rows else ""
            self.log_message(f"ЭМК: {path} ({len(self.df_emk)} строк{extra})")
            if not self.store.ops.empty:
                n = self._rebind_emk_to_store()
                self.log_message(f"Привязка ЭМК: {n} операций с типом госпитализации")
                self.run_analysis()
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
            logging.error(traceback.format_exc())
        finally:
            self._busy(False)

    def _rebind_emk_to_store(self) -> int:
        if self.df_emk is None or self.store.ops.empty:
            return 0
        from analyzers.surgery import SurgeryAnalyzer as SA

        analyzer = SA.__new__(SA)
        analyzer.emk_df = self.df_emk
        analyzer.categories = self._surgery_categories()
        analyzer._code_index = analyzer._build_code_index()
        analyzer._emk_hosp_index = analyzer._build_emk_hosp_index()
        updated = 0
        ops = self.store.ops
        for i, row in ops.iterrows():
            hosp = analyzer.hosp_type_for(row["КВС"], row["Дата"])
            diag = analyzer.diagnosis_for(row["КВС"], row["Дата"])
            changed = False
            if (hosp or "") != str(row.get("Тип_ЭМК") or ""):
                ops.at[i, "Тип_ЭМК"] = hosp or ""
                changed = True
            if (diag or "") != str(row.get("Диагноз") or ""):
                ops.at[i, "Диагноз"] = diag or ""
                changed = True
            cat = str(row.get("Категория") or "")
            if cat in ("Миринготомия план", "Миринготомия экстр"):
                day = pd.Timestamp(row["Дата"]).normalize()
                has_adeno = (
                    (ops["КВС"].astype(str) == str(row["КВС"]))
                    & (pd.to_datetime(ops["Дата"]).dt.normalize() == day)
                    & (ops["Категория"] == "Аденотомия")
                ).any()
                new_cat = (
                    "Миринготомия план"
                    if has_adeno
                    else ("Миринготомия экстр" if hosp == "экстренная" else "Миринготомия план")
                )
                if new_cat != cat:
                    ops.at[i, "Категория"] = new_cat
                    changed = True
            if changed:
                updated += 1
        return updated

    def load_surg(self):
        paths = filedialog.askopenfilenames(
            initialdir=self.last_surg_dir,
            filetypes=[("Excel/CSV", "*.xlsx *.csv")],
            title="Опержурнал(ы)",
        )
        if not paths:
            return
        self.last_surg_dir = str(Path(paths[0]).parent)
        self._persist_settings()
        self._ingest_surg_paths(list(paths))

    def load_surg_folder(self):
        folder = filedialog.askdirectory(
            initialdir=self.last_surg_dir,
            title="Папка с опержурналами",
        )
        if not folder:
            return
        self.last_surg_dir = folder
        self._persist_settings()
        base = Path(folder)
        paths = sorted(
            [*(base.glob("*.xlsx")), *(base.glob("*.xls")), *(base.glob("*.csv"))],
            key=lambda p: p.name.lower(),
        )
        paths = [p for p in paths if not p.name.startswith("~$") and ".bak." not in p.name.lower()]
        if not paths:
            messagebox.showinfo("Папка", "В папке нет Excel/CSV журналов")
            return
        if not messagebox.askyesno(
            "Загрузка из папки",
            f"Найдено файлов: {len(paths)}\n\nЗагрузить все?",
        ):
            return
        self._ingest_surg_paths([str(p) for p in paths])

    def _ingest_surg_paths(self, paths: list):
        try:
            self._busy(True)
            dept = self.dept_var.get()
            any_added = False
            for path in paths:
                try:
                    df = read_table(path)
                    n_all = len(df)
                    analyzer = SurgeryAnalyzer(
                        df, dept, self._surgery_categories(), emk_df=self.df_emk
                    )
                    n_dept = len(analyzer.df)
                    self.log_message(
                        f"{os.path.basename(path)}: фильтр «{dept}» — {n_dept} из {n_all} строк журнала"
                    )
                    if n_dept == 0:
                        self.log_message(
                            f"  нет строк отделения — проверьте название в журнале (колонки "
                            f"«Отделение госпитализации» / «Оперблок»)"
                        )
                    ops = analyzer.extract_operations()
                    if ops.empty:
                        self.log_message(f"Нет операций: {os.path.basename(path)}")
                        continue
                    info = self.store.add(ops, path)
                    any_added = True
                    self.last_batch_span = (info.get("date_from"), info.get("date_to"))
                    msg = f"{os.path.basename(path)}: +{info['added']}, вытеснено {info['removed']}, всего {info['total']}"
                    if info.get("date_from") is not None:
                        msg += f" | {info['date_from'].strftime('%d.%m.%Y')}–{info['date_to'].strftime('%d.%m.%Y')}"
                    self.log_message(msg)
                    uncl = ops[ops["Категория"] == "Не классифицировано"]
                    if len(uncl):
                        codes = sorted({c for c in uncl["Код"].dropna().unique() if c})
                        self.log_message(f"  не классифицировано: {len(uncl)} опер., коды: {codes}")
                        for code in codes:
                            hint = ""
                            if "КСГ_подсказка" in uncl.columns:
                                rows = uncl[uncl["Код"] == code]
                                if len(rows):
                                    hint = str(rows.iloc[0].get("КСГ_подсказка", "") or "")
                            if hint:
                                self.log_message(f"    {code}: {hint}")
                            else:
                                self.log_message(f"    {code}: нет в KSGoperacii.csv — добавьте в config.yaml")
                    if "Спор_ключей" in ops.columns:
                        n_disp = int(ops["Спор_ключей"].fillna(False).astype(bool).sum())
                        if n_disp:
                            self.log_message(f"  спорных по ключам: {n_disp} (вкладка «Спорные»)")
                except Exception as e:
                    messagebox.showerror("Ошибка", f"{os.path.basename(path)}:\n{e}")
                    logging.error(traceback.format_exc())
            if any_added:
                self.loaded_department = dept
            self._refresh_sources_list()
            self.run_analysis()
            self.notebook.select(self.tab_preview)
            self._maybe_hint_year_mismatch()
        finally:
            self._busy(False)

    def get_view_ops(self) -> pd.DataFrame:
        ops = self.store.ops.copy()
        if ops.empty:
            return ops
        if self.filter_enabled.get():
            try:
                start = pd.Timestamp(self._parse_date(self.start_date_var.get()))
                end = pd.Timestamp(self._parse_date(self.end_date_var.get())) + timedelta(days=1) - timedelta(seconds=1)
            except Exception:
                messagebox.showerror("Дата", "Формат даты: ДД.ММ.ГГГГ")
                return ops
            dates = pd.to_datetime(ops["Дата"], errors="coerce")
            ops = ops.loc[(dates >= start) & (dates <= end)].copy()
        return ops

    def run_analysis(self):
        ops = self.get_view_ops()
        if ops.empty:
            messagebox.showwarning("Нет данных", "Добавьте опержурнал(ы)")
            return
        try:
            self._busy(True)
            cat_table, totals_df, weeks = build_summary_tables(
                ops, self.summary_cfg, self._surgery_categories()
            )
            if self.plan_mode.get() == "emk" and self.df_emk is not None:
                totals_df = self._totals_from_emk(ops, totals_df, weeks)
            self.cat_table, self.totals_df, self.weeks = cat_table, totals_df, weeks
            self._update_month_choices(ops)
            self.refresh_preview()
            self._update_unclassified(ops)
            self._update_disputed(ops)
            self._update_kpis(ops, totals_df)
            self._maybe_update_emk_kpi(ops)
            self.status_var.set(f"Готово: {len(ops)} операций, {len(weeks)} нед.")
            d0 = pd.to_datetime(ops["Дата"]).min()
            d1 = pd.to_datetime(ops["Дата"]).max()
            prev_cat_rows = len(self.tree_preview_cat.get_children()) if hasattr(self, "tree_preview_cat") else 0
            prev_tot_rows = len(self.tree_preview_tot.get_children()) if hasattr(self, "tree_preview_tot") else 0
            self.log_message(
                f"Отчёт: {len(ops)} опер. | {d0.strftime('%d.%m.%Y')}–{d1.strftime('%d.%m.%Y')} | "
                f"месяц превью={self.preview_month.get()!r} | категории превью={prev_cat_rows} | итоги превью={prev_tot_rows}"
            )
        except Exception as e:
            self.log_message(traceback.format_exc(), level="ERROR")
            messagebox.showerror("Ошибка", str(e))
        finally:
            self._busy(False)

    def _update_month_choices(self, ops):
        months = sorted({int(m) for m in pd.to_datetime(ops["Дата"]).dt.month.dropna().unique()})
        try:
            year = int(self.year_var.get())
        except ValueError:
            year = int(self.summary_cfg.get("year", 2026))
        self.summary_cfg["year"] = year
        labels = []
        self._month_label_to_num = {}
        for m in months:
            if m not in MONTH_RU:
                continue
            lab = f"{MONTH_RU[m].capitalize()} {year}"
            labels.append(lab)
            self._month_label_to_num[lab] = m
        self.month_combo["values"] = labels
        if labels:
            prefer = None
            if self.last_batch_span[0] is not None:
                prefer = f"{MONTH_RU[int(self.last_batch_span[0].month)].capitalize()} {year}"
            self.preview_month.set(prefer if prefer in labels else labels[0])
        else:
            self.preview_month.set("")
            self.log_message("Превью: не удалось определить месяц по датам операций", level="WARNING")

    def _totals_from_emk(self, ops, totals_df, weeks):
        df = totals_df.copy()
        emerg = pd.Series(0, index=weeks, dtype=int)
        plan = pd.Series(0, index=weeks, dtype=int)
        ops = ops.copy()
        ops["Неделя_начало"] = pd.to_datetime(ops["Дата"]).apply(lambda d: d - pd.Timedelta(days=d.weekday()))
        for _, row in ops.iterrows():
            w = row["Неделя_начало"]
            t = str(row.get("Тип_ЭМК", "") or "").lower()
            if t.startswith("экстр"):
                emerg[w] = emerg.get(w, 0) + 1
            elif t.startswith("план"):
                plan[w] = plan.get(w, 0) + 1
        df.loc["Экстренно операций"] = emerg.reindex(weeks, fill_value=0).astype(int).values
        df.loc["План операций"] = plan.reindex(weeks, fill_value=0).astype(int).values
        return df.astype(int)

    def _update_kpis(self, ops, totals_df):
        self.kpi_ops_var.set(str(len(ops)))
        self.kpi_patients_var.set(str(int(ops["КВС"].nunique())) if len(ops) else "0")
        if totals_df is not None and not totals_df.empty and "Всего операций" in totals_df.index:
            total = int(totals_df.loc["Всего операций"].sum())
            plan = int(totals_df.loc["План операций"].sum()) if "План операций" in totals_df.index else 0
            emerg = int(totals_df.loc["Экстренно операций"].sum()) if "Экстренно операций" in totals_df.index else 0
            self.kpi_plan_var.set(f"{plan * 100 / total:.0f}%" if total else "—")
            self.kpi_emerg_var.set(f"{emerg * 100 / total:.0f}%" if total else "—")
        if len(ops):
            d0 = pd.to_datetime(ops["Дата"]).min().strftime("%d.%m.%Y")
            d1 = pd.to_datetime(ops["Дата"]).max().strftime("%d.%m.%Y")
            self.kpi_period_var.set(f"{d0} – {d1}")

    def _maybe_update_emk_kpi(self, ops):
        if self.df_emk is None or ops.empty:
            self.kpi_diff_var.set("—")
            return
        if ops["Тип_ЭМК"].astype(str).str.strip().eq("").all():
            self.kpi_diff_var.set("нет связи")
            return
        result = compare_plan_emergency(ops, self.summary_cfg, department=self.dept_var.get())
        self.kpi_diff_var.set(str(len(result["mismatches"])))
        self._fill_emk_tree(result, select_tab=False)

    def on_emk_mode(self):
        if self.df_emk is None:
            messagebox.showinfo("ЭМК", "Сначала загрузите ЭМК")
            self.plan_mode.set("template")
            return
        self.run_analysis()
        self.show_emk_diff()

    def _fill_emk_tree(self, result: dict, *, select_tab: bool = False):
        self.last_emk_compare = result
        mismatches = result.get("mismatches") or []
        n = len(mismatches)
        compared = int(result.get("compared") or 0)
        if hasattr(self, "tree_emk"):
            for item in self.tree_emk.get_children():
                self.tree_emk.delete(item)
            for m in mismatches:
                dt = m.get("Дата")
                dt_s = dt.strftime("%d.%m.%Y") if hasattr(dt, "strftime") else str(dt or "")
                self.tree_emk.insert(
                    "",
                    tk.END,
                    values=(
                        dt_s,
                        m.get("КВС"),
                        m.get("Категория"),
                        m.get("Код"),
                        m.get("Шаблон"),
                        m.get("ЭМК"),
                        str(m.get("Диагноз") or "")[:100],
                        str(m.get("Услуга") or "")[:100],
                    ),
                )
            linked = int(result.get("emk_linked") or 0)
            total = int(result.get("total_ops") or 0)
            self.notebook.tab(self.tab_emk, text=f"Расхождения ЭМК ({n})")
            if n:
                self.emk_info.set(
                    f"Расхождений: {n} из {compared} сравнений "
                    f"(связано с ЭМК: {linked} из {total} операций)"
                )
            else:
                self.emk_info.set(
                    f"Расхождений нет (сравнено {compared}, связано {linked}/{total})"
                )
        self.kpi_diff_var.set(str(n))
        if select_tab and hasattr(self, "tab_emk"):
            self.notebook.select(self.tab_emk)

    def show_emk_diff(self):
        ops = self.get_view_ops()
        if ops.empty:
            messagebox.showwarning("Нет данных", "Нет операций")
            return
        if self.df_emk is None:
            messagebox.showwarning("Нет ЭМК", "Загрузите ЭМК")
            return
        if ops["Тип_ЭМК"].astype(str).str.strip().eq("").all():
            messagebox.showwarning("Нет связи", "КВС журнала и ЭМК не пересекаются")
            return
        result = compare_plan_emergency(ops, self.summary_cfg, department=self.dept_var.get())
        self._fill_emk_tree(result, select_tab=True)
        self.log_message(format_mismatch_report(result, limit=30))
        if result["mismatches"]:
            messagebox.showwarning(
                "Расхождения",
                f"Найдено: {len(result['mismatches'])}\nСм. вкладку «Расхождения ЭМК»",
            )
        else:
            messagebox.showinfo("Сверка", "Расхождений нет")

    def display_tables(self):
        """Раньше заполнял дублирующие вкладки Категории/Итоги — оставлены только превью."""
        return

    def _fill_tree(self, tree, first_col, week_headers, df, week_keys, totals=False):
        for item in tree.get_children():
            tree.delete(item)
        if df is None or df.empty:
            return
        col_ids = ["c0"] + [f"w{i}" for i in range(len(week_headers))]
        headings = [first_col] + list(week_headers)
        self._configure_tree_columns(tree, col_ids, headings)
        for idx, row in df.iterrows():
            values = [idx] + [int(row[week]) for week in week_keys]
            tags = []
            if totals or idx in (
                "Всего операций", "Экстренно операций", "План операций", "Дети всего", "Взрослые", "Человек"
            ):
                tags.append("total")
            elif sum(values[1:]) == 0:
                tags.append("zero")
            tree.insert("", tk.END, values=values, tags=tuple(tags))

    def refresh_preview(self):
        # очищаем 3 таблицы превью
        for tree in (self.tree_preview_cat, self.tree_preview_tot, self.tree_form):
            for item in tree.get_children():
                tree.delete(item)
        ops = self.get_view_ops()
        if ops.empty:
            self.preview_info.set("Превью: нет операций")
            return
        label = self.preview_month.get()
        month_map = getattr(self, "_month_label_to_num", {})
        if not label or label not in month_map:
            self.preview_info.set(f"Превью: месяц не выбран (label={label!r})")
            self.log_message(f"Превью пусто: label={label!r}, map={list(month_map.keys())}", level="WARNING")
            return
        month = int(month_map[label])
        try:
            year = int(self.year_var.get())
        except ValueError:
            year = int(self.summary_cfg.get("year", 2026))
        weeks = self._weeks_for_month(year, month)
        if not weeks:
            weeks = compute_month_weeks(year, month)
        week_headers = [f"{s.strftime('%d.%m')}-{e.strftime('%d.%m')}" for s, e in weeks]
        # --- превью категории ---
        col_ids_cat = ["c0"] + [f"w{i}" for i in range(len(weeks))] + ["tot"]
        headings_cat = ["Категория"] + week_headers + ["ИТОГ"]
        self._configure_tree_columns(
            self.tree_preview_cat, col_ids_cat, headings_cat, widths=[240] + [90] * (len(weeks) + 1)
        )

        month_ops = ops[pd.to_datetime(ops["Дата"]).dt.month == month].copy()
        cat_order = list(self.summary_cfg.get("category_rows", {}).keys())
        counts_map = {cat: [0] * len(weeks) for cat in cat_order}
        unmapped = 0
        for _, r in month_ops.iterrows():
            cat = r["Категория"]
            if cat not in counts_map:
                continue
            d = pd.Timestamp(r["Дата"]).date()
            hit = False
            for wi, (s, e) in enumerate(weeks):
                if s <= d <= e:
                    counts_map[cat][wi] += 1
                    hit = True
                    break
            if not hit:
                unmapped += 1

        rows_cat = 0
        for cat in cat_order:
            counts = counts_map[cat]
            total = sum(counts)
            if self.hide_zeros.get() and total == 0:
                continue
            self.tree_preview_cat.insert(
                "", tk.END, values=[cat] + counts + [total], tags=("zero",) if total == 0 else ()
            )
            rows_cat += 1

        emerg_set = set(self.summary_cfg.get("emergency_categories", []))
        plan_set = set(self.summary_cfg.get("plan_categories", []))

        def week_slice(wi):
            s, e = weeks[wi]
            return month_ops[month_ops["Дата"].map(lambda x: s <= pd.Timestamp(x).date() <= e)]

        arrays = {k: [] for k in ("ops", "emerg", "plan", "kids", "people")}
        for i in range(len(weeks)):
            wops = week_slice(i)
            arrays["ops"].append(len(wops))
            arrays["emerg"].append(int(wops["Категория"].isin(emerg_set).sum()))
            arrays["plan"].append(int(wops["Категория"].isin(plan_set).sum()))
            arrays["kids"].append(int(wops.loc[wops["Возраст"].fillna(99) < 18, "КВС"].nunique()))
            arrays["people"].append(int(wops["КВС"].nunique()))
        adults = [max(0, arrays["people"][i] - arrays["kids"][i]) for i in range(len(weeks))]

        # --- превью итоги ---
        col_ids_tot = ["c0"] + [f"w{i}" for i in range(len(weeks))] + ["tot"]
        headings_tot = ["Показатель"] + week_headers + ["ИТОГ"]
        self._configure_tree_columns(
            self.tree_preview_tot, col_ids_tot, headings_tot, widths=[260] + [90] * (len(weeks) + 1)
        )
        rows_tot = 0
        for name, arr in (
            ("Всего операций", arrays["ops"]),
            ("Экстренно операций", arrays["emerg"]),
            ("План операций", arrays["plan"]),
            ("Дети всего", arrays["kids"]),
            ("Взрослые", adults),
            ("Человек", arrays["people"]),
        ):
            self.tree_preview_tot.insert("", tk.END, values=[name] + arr + [sum(arr)], tags=("total",))
            rows_tot += 1

        # --- форма 4001 ---
        form_cfg = self.summary_cfg.get("form_4001") or {}
        stats = compute_form_4001(
            month_ops,
            self._surgery_categories(),
            pension_age=int(self.config.get("thresholds", {}).get("pension_age", 60)),
            form_cfg=form_cfg,
        )
        # Колонки как в шаблоне: 1,2,3,4,5,6,28 + S (т.4000)
        form_cols = ["c0", "line", "n", "o", "p", "q", "r", "s"]
        form_heads = [
            "Наименование операции",
            "№ строки",
            "всего",
            "0–14 лет",
            "до 1 года",
            "15–17 лет",
            "морфол. (28)",
            "Всего (т.4000)",
        ]
        self._configure_tree_columns(
            self.tree_form, form_cols, form_heads, widths=[280, 70, 70, 70, 70, 70, 90, 90]
        )
        for row in form_4001_preview_rows(stats):
            name = row["name"]
            bold = name in (
                "операции на органах уха, горла, носа",
                "операции на органах дыхания",
                "операции на коже и подкожной клетчатке",
                "Всего операций",
            )
            self.tree_form.insert(
                "",
                tk.END,
                values=(
                    name,
                    row.get("line", ""),
                    row.get("total", ""),
                    row.get("age_0_14", ""),
                    row.get("age_under_1", ""),
                    row.get("age_15_17", ""),
                    row.get("histology", ""),
                    row.get("senior", ""),
                ),
                tags=("total",) if bold else (),
            )

        self.preview_info.set(
            f"Превью: {label} | операций месяца {len(month_ops)} | недель {len(weeks)} | категории {rows_cat} | итоги {rows_tot}"
            + (f" | вне недель: {unmapped}" if unmapped else "")
        )
        self.log_message(
            f"Превью: {label}, недель={[(str(s), str(e)) for s, e in weeks]}, "
            f"ops_month={len(month_ops)}, categories={rows_cat}, totals={rows_tot}, unmapped={unmapped}, "
            f"form4001={stats}"
        )

    def _update_unclassified(self, ops):
        for item in self.tree_uncl.get_children():
            self.tree_uncl.delete(item)
        uncl = ops[ops["Категория"] == "Не классифицировано"]
        self.notebook.tab(self.tab_uncl, text=f"Не классифицировано ({len(uncl)})")
        for _, r in uncl.iterrows():
            dt = r["Дата"]
            dt_s = dt.strftime("%d.%m.%Y") if hasattr(dt, "strftime") else str(dt)
            svc = str(r.get("Услуга", "") or "")[:80]
            self.tree_uncl.insert(
                "",
                tk.END,
                values=(
                    dt_s,
                    r.get("КВС"),
                    r.get("Код"),
                    r.get("КСГ_название", ""),
                    r.get("КСГ", ""),
                    svc,
                ),
            )

    def _update_disputed(self, ops):
        for item in self.tree_dispute.get_children():
            self.tree_dispute.delete(item)
        if ops is None or ops.empty or "Спор_ключей" not in ops.columns:
            self.notebook.tab(self.tab_dispute, text="Спорные")
            return
        disp = ops[ops["Спор_ключей"].fillna(False).astype(bool)]
        self.notebook.tab(self.tab_dispute, text=f"Спорные ({len(disp)})")
        for idx, r in disp.iterrows():
            dt = r["Дата"]
            dt_s = dt.strftime("%d.%m.%Y") if hasattr(dt, "strftime") else str(dt)
            svc = str(r.get("Услуга", "") or "")[:100]
            self.tree_dispute.insert(
                "",
                tk.END,
                iid=str(idx),
                values=(
                    dt_s,
                    r.get("КВС"),
                    r.get("Код"),
                    svc,
                    r.get("Категория"),
                    r.get("Спорные_категории", ""),
                ),
            )

    def update_summary(self):
        """Один диалог записи: галочки «Недели» и «Форма 4001»."""
        if self.store.ops.empty:
            messagebox.showwarning("Нет данных", "Сначала добавьте опержурнал(ы)")
            return
        path = self.summary_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Нет файла", f"Сводная не найдена:\n{path}")
            return

        selected = self.dept_var.get()
        if self.loaded_department and self.loaded_department != selected:
            if not messagebox.askyesno(
                "Отделение",
                f"В накопителе данные для «{self.loaded_department}»,\n"
                f"а выбрано «{selected}».\n\n"
                "Записать всё равно?",
            ):
                return

        d_min, d_max = self.last_batch_span
        if d_min is None or d_max is None:
            d_min, d_max = self.store.date_span(self.store.ops)
        try:
            cfg_year = int(self.year_var.get())
        except ValueError:
            cfg_year = int(self.summary_cfg.get("year", 2026))

        if d_min is not None:
            op_years = sorted({int(y) for y in pd.to_datetime(self.store.ops["Дата"]).dt.year.dropna().unique()})
            foreign = [y for y in op_years if y != cfg_year]
            if foreign:
                if not messagebox.askyesno(
                    "Год данных",
                    f"В накопителе есть операции за {op_years}, а сводная/год в настройках — {cfg_year}.\n\n"
                    f"Записать всё равно в:\n{path}\n\n"
                    "Для другого года: смените «Год» или «Создать сводную на год…».",
                ):
                    return

        period = ""
        if d_min is not None:
            period = f"{d_min.strftime('%d.%m.%Y')} – {d_max.strftime('%d.%m.%Y')}"

        top = tk.Toplevel(self.root)
        top.title("Запись в Excel")
        top.transient(self.root)
        top.resizable(False, False)
        tk.Label(top, text=f"Файл:\n{path}", justify=tk.LEFT, wraplength=480).pack(padx=12, pady=(12, 4), anchor="w")
        if period:
            tk.Label(top, text=f"Период перезаписи: {period}", justify=tk.LEFT).pack(padx=12, pady=2, anchor="w")
        tk.Label(top, text=f"Год настроек: {cfg_year}", justify=tk.LEFT).pack(padx=12, pady=2, anchor="w")

        weeks_var = BooleanVar(value=self.write_weeks_var.get())
        form_enabled = form_4001_enabled(self.summary_cfg)
        form_var = BooleanVar(value=self.write_form_var.get() if form_enabled else False)
        tk.Checkbutton(top, text="Недели / категории (столбцы C–G)", variable=weeks_var).pack(
            padx=12, pady=4, anchor="w"
        )
        if form_enabled:
            tk.Checkbutton(top, text="Форма 4001 (формулы N/R не затираются)", variable=form_var).pack(
                padx=12, pady=4, anchor="w"
            )
        else:
            tk.Label(top, text="Форма 4001 для этого отделения не используется.", fg="#555").pack(
                padx=12, pady=4, anchor="w"
            )
        tk.Label(
            top,
            text="Закройте файл в Excel перед записью. Бэкап — в папку backups/ (до 20 шт.).",
            fg="#555",
            wraplength=480,
            justify=tk.LEFT,
        ).pack(padx=12, pady=6, anchor="w")

        def do_write():
            write_weeks = bool(weeks_var.get())
            write_form = bool(form_var.get()) if form_enabled else False
            if not write_weeks and not write_form:
                messagebox.showwarning("Запись", "Отметьте хотя бы один пункт", parent=top)
                return
            self.write_weeks_var.set(write_weeks)
            self.write_form_var.set(write_form)
            top.destroy()
            self._do_write_excel(path, d_min, d_max, write_weeks=write_weeks, write_form=write_form)

        bf = tk.Frame(top)
        bf.pack(pady=10)
        _btn(bf, "Записать", do_write, side=tk.LEFT, padx=4)
        _btn(bf, "Отмена", top.destroy, side=tk.LEFT, padx=4)
        top.grab_set()
        top.focus_force()

    def _do_write_excel(self, path, d_min, d_max, *, write_weeks: bool, write_form: bool):
        if excel_file_locked(path):
            messagebox.showerror(
                "Файл занят",
                f"Сводная открыта в Excel или заблокирована:\n{path}\n\n"
                "Закройте файл и повторите запись.",
            )
            return
        try:
            self._busy(True)
            parts = []
            if write_weeks:
                parts.append("недели")
            if write_form:
                parts.append("форма 4001")
            self.log_message(
                f"Запись в сводную ({', '.join(parts)}): {path} | ops={len(self.store.ops)} | "
                f"overwrite={d_min}..{d_max}"
            )
            writer = SummaryWriter(
                path,
                self.summary_cfg,
                department=self.dept_var.get(),
                categories=self._surgery_categories(),
                pension_age=int(self.config.get("thresholds", {}).get("pension_age", 60)),
            )
            report = writer.write(
                self.store.ops,
                output_path=path,
                overwrite_from=d_min,
                overwrite_to=d_max,
                backup=True,
                write_weeks=write_weeks,
                write_form=write_form,
            )
            blank_delta = int(report.get("blank_delta") or 0)
            if blank_delta:
                shift_totals_rows_by_delta(self.config, blank_delta, summary_key=self.summary_key)
                save_config(self.config, APP_DIR / "config.yaml")
                self._sync_dept_context()
                self.log_message(
                    f"Разделитель перед итогами: сдвиг totals_rows на {blank_delta:+d}"
                )
            months = ", ".join(report.get("months", {}).keys()) or "—"
            self.log_message(
                f"Сводная обновлена: {months}, ячеек {report.get('cells_written', 0)}, "
                f"вне недель {report.get('unmapped_dates', 0)}"
            )
            for sheet, info in report.get("months", {}).items():
                self.log_message(
                    f"  [{sheet}] cols={info.get('cols')} weeks={info.get('weeks')} "
                    f"ops={info.get('ops')} sample={info.get('sample')}"
                )
                form = info.get("form_4001")
                if form:
                    self.log_message(f"  [{sheet}] форма 4001: {form}")
            self.status_var.set(f"Запись: {report.get('cells_written', 0)} ячеек")
            self._persist_settings()
            verify_msg = ""
            if write_weeks:
                try:
                    vres = verify_write_report(path, report)
                    verify_msg = format_verify_message(vres)
                    self.log_message(verify_msg, level="INFO" if vres.get("ok") else "WARNING")
                except Exception as ve:
                    verify_msg = f"Проверка записи не выполнена: {ve}"
                    self.log_message(verify_msg, level="WARNING")
            bak = report.get("backup")
            done = (
                f"Обновлено: {months}\nЯчеек: {report.get('cells_written', 0)}\n"
            )
            if bak:
                done += f"Бэкап: {Path(bak).name}\n"
            if verify_msg:
                done += f"\n{verify_msg}\n"
            done += "\nОткрыть файл? (если Excel был открыт — закройте и откройте заново)"
            if messagebox.askyesno("Готово", done):
                self.open_summary_file()
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
            self.log_message(traceback.format_exc(), level="ERROR")
        finally:
            self._busy(False)

    def restore_summary_backup(self):
        path = self.summary_path.get().strip()
        if not path:
            messagebox.showerror("Сводная", "Сначала укажите файл сводной")
            return
        baks = list_backups(path)
        if not baks:
            from analyzers.backup_utils import backups_dir

            messagebox.showinfo(
                "Бэкапы",
                f"Нет бэкапов (.bak.xlsx).\nПапка: {backups_dir(path)}",
            )
            return
        top = tk.Toplevel(self.root)
        top.title("Восстановить из бэкапа")
        top.transient(self.root)
        tk.Label(
            top,
            text="Выберите бэкап (папка backups/, хранятся последние 20). "
            "Текущий файл сохранится туда же перед восстановлением.",
            wraplength=520,
            justify=tk.LEFT,
        ).pack(padx=12, pady=8, anchor="w")
        lb = tk.Listbox(top, width=72, height=min(12, max(4, len(baks))))
        lb.pack(fill=tk.BOTH, expand=True, padx=12, pady=4)
        for b in baks:
            mtime = datetime.fromtimestamp(b.stat().st_mtime).strftime("%d.%m.%Y %H:%M")
            lb.insert(tk.END, f"{mtime}  —  {b.name}")
        lb.selection_set(0)

        def do_restore():
            sel = lb.curselection()
            if not sel:
                return
            bak = baks[int(sel[0])]
            if excel_file_locked(path):
                messagebox.showerror("Файл занят", "Закройте сводную в Excel", parent=top)
                return
            if not messagebox.askyesno(
                "Восстановление",
                f"Заменить текущую сводную содержимым:\n{bak.name}?",
                parent=top,
            ):
                return
            try:
                restore_backup(bak, path)
                self.log_message(f"Восстановлено из бэкапа: {bak}")
                messagebox.showinfo("Готово", f"Восстановлено:\n{path}", parent=top)
                top.destroy()
            except Exception as e:
                messagebox.showerror("Ошибка", str(e), parent=top)

        bf = tk.Frame(top)
        bf.pack(pady=10)
        _btn(bf, "Восстановить", do_restore, side=tk.LEFT, padx=4)
        _btn(bf, "Отмена", top.destroy, side=tk.LEFT, padx=4)

    def open_summary_file(self):
        path = self.summary_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Нет файла", path)
            return
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", path])
            elif sys.platform.startswith("win"):
                os.startfile(path)  # type: ignore
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def export_simple(self):
        if self.store.ops.empty:
            messagebox.showwarning("Нет данных", "Сначала анализ / загрузка журнала")
            return
        path_tpl = self.summary_path.get().strip()
        if not path_tpl or not os.path.exists(path_tpl):
            messagebox.showerror("Нет шаблона", f"Нужен файл сводной как образец стиля:\n{path_tpl}")
            return
        label = self.preview_month.get()
        month_map = getattr(self, "_month_label_to_num", {})
        month = int(month_map[label]) if label in month_map else None
        try:
            year = int(self.year_var.get())
        except ValueError:
            year = int(self.summary_cfg.get("year", 2026))
        default_name = f"Отчёт_{MONTH_RU.get(month or 0, 'период')}_{year}.xlsx"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not file_path:
            return
        try:
            self._busy(True)
            report = export_month_like_summary(
                path_tpl,
                file_path,
                self.get_view_ops(),
                self.summary_cfg,
                department=self.dept_var.get(),
                categories=self._surgery_categories(),
                pension_age=int(self.config.get("thresholds", {}).get("pension_age", 60)),
                month=month,
                year=year,
            )
            self.log_message(
                f"Экспорт как сводная: {file_path} | ячеек {report.get('cells_written', 0)} | "
                f"листы {list(report.get('months', {}))}"
            )
            messagebox.showinfo(
                "Готово",
                f"Сохранён отчёт в стиле сводной:\n{file_path}\n\n"
                "Листы/формулы как в шаблоне; заполнен выбранный месяц.",
            )
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
            self.log_message(traceback.format_exc(), level="ERROR")
        finally:
            self._busy(False)

    def create_year_summary_dialog(self):
        """Создать файл «Операции сводная YYYY.xlsx» на новый год из текущего шаблона."""
        tpl = self.summary_path.get().strip()
        if not tpl or not os.path.exists(tpl):
            messagebox.showerror(
                "Нет шаблона",
                "Укажите существующую сводную через «Обзор…» — файл может лежать в любой папке.",
            )
            return
        try:
            cur = int(self.year_var.get())
        except ValueError:
            cur = 2026
        new_year = cur + 1
        base_dir = self._summary_dir()

        top = tk.Toplevel(self.root)
        top.title("Сводная на год")
        top.transient(self.root)
        tk.Label(
            top,
            text="Новый файл будет создан в той же папке, что и текущая сводная.",
            wraplength=480,
            justify=tk.LEFT,
        ).pack(padx=12, pady=(12, 4), anchor="w")
        tk.Label(top, text="Год:").pack(padx=12, anchor="w")
        yvar = StringVar(value=str(new_year))
        ttk.Combobox(top, textvariable=yvar, values=[str(y) for y in range(cur, cur + 6)], width=8).pack(
            padx=12, anchor="w"
        )
        out_hint = StringVar(value=str(suggest_summary_path(base_dir, new_year)))

        def refresh_hint(*_a):
            try:
                out_hint.set(str(suggest_summary_path(base_dir, int(yvar.get()))))
            except ValueError:
                pass

        yvar.trace_add("write", refresh_hint)
        tk.Label(top, text="Путь:", anchor="w").pack(padx=12, pady=(8, 0), anchor="w")
        tk.Label(top, textvariable=out_hint, wraplength=480, justify=tk.LEFT, fg="#333").pack(
            padx=12, pady=4, anchor="w"
        )

        def pick_other_folder():
            folder = filedialog.askdirectory(initialdir=str(base_dir), parent=top, title="Папка для новой сводной")
            if not folder:
                return
            try:
                y = int(yvar.get())
            except ValueError:
                y = new_year
            out_hint.set(str(suggest_summary_path(Path(folder), y)))

        def do_create():
            try:
                y = int(yvar.get())
            except ValueError:
                messagebox.showerror("Год", "Укажите год числом", parent=top)
                return
            dest = Path(out_hint.get())
            if dest.exists() and not messagebox.askyesno(
                "Файл есть", f"{dest.name} уже существует. Перезаписать?", parent=top
            ):
                return
            try:
                rows_map = self.summary_cfg.get("category_rows") or {}
                cat_max = max(int(v) for v in rows_map.values()) if rows_map else 37
                path = create_year_summary(
                    tpl,
                    y,
                    output_path=str(dest),
                    sheet_names=self.summary_cfg.get("sheet_names"),
                    clear_values=True,
                    category_row_max=cat_max,
                    totals_rows=self.summary_cfg.get("totals_rows"),
                )
                self.year_var.set(str(y))
                self.summary_cfg["year"] = y
                self.summary_path.set(str(path))
                self._persist_settings()
                self.log_message(f"Создана сводная на {y}: {path}")
                messagebox.showinfo("Готово", f"Создан файл:\n{path}", parent=top)
                top.destroy()
            except Exception as e:
                messagebox.showerror("Ошибка", str(e), parent=top)

        bf = tk.Frame(top)
        bf.pack(pady=10)
        _btn(bf, "Другая папка…", pick_other_folder, side=tk.LEFT, padx=4)
        _btn(bf, "Создать", do_create, side=tk.LEFT, padx=4)
        _btn(bf, "Отмена", top.destroy, side=tk.LEFT, padx=4)

    def show_about(self):
        ver = read_local_version(APP_DIR)
        upd = self.config.get("updates") or {}
        repo = f"{upd.get('github_owner', '')}/{upd.get('github_repo', '')}".strip("/")
        repo_line = f"\nGitHub: {repo}" if repo != "/" and repo else ""
        messagebox.showinfo(
            "О программе",
            f"Сводная операций\nВерсия: {ver}{repo_line}\n\n"
            "Запись: один диалог с галочками «Недели» и «Форма 4001».\n"
            "Обновления: Помощь → Проверить обновления…",
        )

    def check_updates(self, silent: bool = False, force: bool = False):
        """Проверить GitHub и предложить установить обновление."""
        cfg = self.config.get("updates") or {}
        if not cfg.get("enabled", True):
            if not silent:
                messagebox.showinfo("Обновления", "Проверка обновлений отключена в config.yaml")
            return
        try:
            self._busy(True)
            self.status_var.set("Проверка обновлений…")
            info = check_for_update(APP_DIR, cfg)
        except Exception as e:
            self.log_message(f"Обновление: ошибка проверки — {e}", level="WARNING")
            if not silent:
                messagebox.showerror("Обновления", str(e))
            return
        finally:
            self._busy(False)
            self.status_var.set("Готов к работе")

        self._mark_update_checked()

        if info is None:
            ver = read_local_version(APP_DIR)
            msg = f"У вас актуальная версия: {ver}"
            self.log_message(f"Обновления: {msg}")
            if not silent:
                messagebox.showinfo("Обновления", msg)
            return

        self.log_message(
            f"Обновления: доступна {info.remote_version} (сейчас {info.local_version})"
        )
        notes = format_update_notes(info)
        top = tk.Toplevel(self.root)
        top.title("Доступно обновление")
        top.transient(self.root)
        top.resizable(True, True)
        tk.Label(top, text="Найдена новая версия", font=("Helvetica", 13, "bold")).pack(
            padx=12, pady=(12, 4), anchor="w"
        )
        txt = tk.Text(top, width=64, height=14, wrap=tk.WORD)
        txt.pack(fill=tk.BOTH, expand=True, padx=12, pady=4)
        txt.insert("1.0", notes)
        txt.configure(state=tk.DISABLED)
        include_cfg = BooleanVar(value=False)
        tk.Checkbutton(
            top,
            text="Также заменить config.yaml (рубрикатор) — обычно не нужно",
            variable=include_cfg,
        ).pack(padx=12, pady=4, anchor="w")

        def do_install():
            if not messagebox.askyesno(
                "Обновление",
                "Установить обновление?\n\n"
                "Будут обновлены код и VERSION.\n"
                "Excel, журналы, .venv и ui_settings.json не трогаются.\n"
                "Создаётся папка .update_backup_…",
                parent=top,
            ):
                return
            try:
                self._busy(True)
                self.status_var.set("Загрузка обновления…")
                report = apply_update_from_zip(
                    APP_DIR,
                    info.zip_url,
                    token=resolve_token(cfg),
                    include_config=bool(include_cfg.get()),
                    backup=True,
                    sha256_url=info.sha256_url,
                    require_sha256=bool(info.source == "release-asset"),
                    mode="release-asset" if info.source == "release-asset" else "auto",
                )
                self.log_message(
                    f"Обновление установлено: v{report.get('new_version')} "
                    f"({report.get('count')} файлов), backup={report.get('backup')}"
                )
                top.destroy()
                new_ver = str(report.get("new_version") or read_local_version(APP_DIR))
                old_ver = info.local_version
                self.app_version = new_ver
                self.root.title(f"Сводная операций  v{new_ver}")
                self.show_whats_new(
                    version=new_ver,
                    previous_version=old_ver,
                    force=True,
                    title=f"Установлена версия {new_ver}",
                )
                if messagebox.askyesno(
                    "Перезапуск",
                    f"Версия {new_ver} установлена.\n\nПерезапустить приложение сейчас?",
                ):
                    self._restart_app()
            except Exception as e:
                self.log_message(traceback.format_exc(), level="ERROR")
                messagebox.showerror("Обновление", str(e), parent=top)
            finally:
                self._busy(False)
                self.status_var.set("Готов к работе")

        bf = tk.Frame(top)
        bf.pack(pady=10)
        _btn(bf, "Установить", do_install, side=tk.LEFT, padx=4)
        _btn(bf, "Позже", top.destroy, side=tk.LEFT, padx=4)
        if info.html_url:
            def open_page():
                try:
                    if sys.platform == "darwin":
                        subprocess.Popen(["open", info.html_url])
                    elif sys.platform.startswith("win"):
                        os.startfile(info.html_url)  # type: ignore
                    else:
                        subprocess.Popen(["xdg-open", info.html_url])
                except Exception as e:
                    messagebox.showerror("Ошибка", str(e), parent=top)

            _btn(bf, "Открыть на GitHub", open_page, side=tk.LEFT, padx=4)

    def show_whats_new(
        self,
        version: str | None = None,
        previous_version: str | None = None,
        *,
        force: bool = False,
        title: str | None = None,
    ):
        """Окно с описанием новых функций из RELEASE_NOTES.md."""
        ver = (version or self.app_version or read_local_version(APP_DIR)).strip()
        text = format_whats_new(
            ver,
            path=APP_DIR / "RELEASE_NOTES.md",
            previous_version=previous_version,
        )
        top = tk.Toplevel(self.root)
        top.title(title or f"Что нового — v{ver}")
        top.transient(self.root)
        top.resizable(True, True)
        tk.Label(
            top,
            text=f"Новые возможности версии {ver}",
            font=("Helvetica", 13, "bold"),
        ).pack(padx=12, pady=(12, 4), anchor="w")
        frame = tk.Frame(top)
        frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=4)
        txt = tk.Text(frame, width=72, height=16, wrap=tk.WORD)
        vs = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=txt.yview)
        txt.configure(yscrollcommand=vs.set)
        txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vs.pack(side=tk.RIGHT, fill=tk.Y)
        txt.insert("1.0", text)
        txt.configure(state=tk.DISABLED)

        def close():
            self.last_seen_version = ver
            try:
                self._persist_settings()
            except Exception:
                pass
            top.destroy()

        bf = tk.Frame(top)
        bf.pack(pady=10)
        _btn(bf, "Понятно", close, side=tk.LEFT, padx=4)
        top.protocol("WM_DELETE_WINDOW", close)
        if force:
            top.lift()
            top.focus_force()

    def _maybe_show_whats_new(self):
        """После обновления — показать «Что нового» один раз при старте."""
        current = (self.app_version or read_local_version(APP_DIR)).strip()
        last = (self.last_seen_version or "").strip()
        if not last:
            # первый запуск / нет метки — запомнить, не беспокоить
            self.last_seen_version = current
            try:
                self._persist_settings()
            except Exception:
                pass
            return
        if last == current:
            return
        self.show_whats_new(version=current, previous_version=last, force=True)

    def _mark_update_checked(self):
        self.last_update_check = datetime.now().strftime("%Y-%m-%d")
        try:
            self._persist_settings()
        except Exception:
            pass

    def _maybe_startup_update_check(self):
        cfg = self.config.get("updates") or {}
        if not cfg.get("enabled", True) or not cfg.get("check_on_startup", True):
            return
        try:
            days = int(cfg.get("check_interval_days", 7) or 7)
        except (TypeError, ValueError):
            days = 7
        days = max(1, days)
        last = (self.last_update_check or "").strip()
        if last:
            try:
                last_d = datetime.strptime(last[:10], "%Y-%m-%d").date()
                if (datetime.now().date() - last_d).days < days:
                    self.log_message(
                        f"Обновления: автопроверка пропущена (последняя {last_d.isoformat()}, интервал {days} дн.)"
                    )
                    return
            except ValueError:
                pass
        self.check_updates(silent=True)

    def _dept_profile(self, name: str | None = None) -> dict:
        profiles = self.config.get("department_profiles") or {}
        return dict(profiles.get(name or self.dept_var.get()) or {})

    def _refresh_dept_hint(self):
        if not hasattr(self, "dept_hint_var"):
            return
        key = dept_summary_key(self.config, self.dept_var.get())
        n_cats = len(get_surgery_categories(self.config, summary_key=key))
        if key == "lor":
            self.dept_hint_var.set(f"Рубрикатор: ЛОР ({n_cats} операций)")
        else:
            label = (DEPT_REPORT_SOURCES.get(key) or {}).get("label") or key
            form = "форма 4001" if form_4001_enabled(get_summary_cfg(self.config, summary_key=key)) else "без 4001"
            self.dept_hint_var.set(f"Рубрикатор: {label} ({n_cats} операций, {form})")

    def _on_department_changed(self):
        old = self._prev_dept
        new = self.dept_var.get()
        if old and old != new:
            self.summary_paths_by_dept[old] = self.summary_path.get()
            saved = self.summary_paths_by_dept.get(new)
            if saved:
                self.summary_path.set(saved)
            else:
                try:
                    y = int(self.year_var.get())
                except (TypeError, ValueError):
                    y = 2026
                sk = dept_summary_key(self.config, new)
                self.summary_path.set(
                    str(APP_DIR / default_summary_filename(self.config, sk, y))
                )
            if not self.store.ops.empty and self.loaded_department and self.loaded_department != new:
                messagebox.showwarning(
                    "Отделение",
                    f"В накопителе данные для «{self.loaded_department}».\n"
                    f"Выбрано «{new}».\n\n"
                    "Очистите накопитель и загрузите журналы заново.",
                )
        self._prev_dept = new
        self._sync_dept_context()
        if not self.store.ops.empty:
            self.run_analysis()
        self._persist_settings()

    def _restart_app(self):
        try:
            self._persist_settings()
        except Exception:
            pass
        if getattr(sys, "frozen", False):
            cmd = [sys.executable]
            cwd = str(APP_DIR)
        else:
            cmd = [sys.executable, str(APP_DIR / "app_desktop.py")]
            cwd = str(APP_DIR)
        try:
            subprocess.Popen(cmd, cwd=cwd)
        except Exception as e:
            messagebox.showerror(
                "Перезапуск",
                f"Не удалось перезапустить автоматически:\n{e}\n\n"
                "Закройте программу и запустите снова.",
            )
            return
        self.root.destroy()

    def _apply_saved_settings(self):
        s = load_settings(APP_DIR)
        if not s:
            return
        if s.get("summary_path"):
            self.summary_path.set(str(s["summary_path"]))
        if s.get("department"):
            self.dept_var.set(str(s["department"]))
        if s.get("year"):
            self.year_var.set(str(s["year"]))
            try:
                self.summary_cfg["year"] = int(s["year"])
            except (TypeError, ValueError):
                pass
        if "hide_zeros" in s:
            self.hide_zeros.set(bool(s["hide_zeros"]))
        if "filter_enabled" in s:
            self.filter_enabled.set(bool(s["filter_enabled"]))
        if s.get("start_date"):
            self.start_date_var.set(str(s["start_date"]))
        if s.get("end_date"):
            self.end_date_var.set(str(s["end_date"]))
        if "write_weeks" in s:
            self.write_weeks_var.set(bool(s["write_weeks"]))
        if "write_form" in s:
            self.write_form_var.set(bool(s["write_form"]))
        if s.get("plan_mode") in ("template", "emk"):
            self.plan_mode.set(s["plan_mode"])
        if s.get("last_surg_dir"):
            self.last_surg_dir = str(s["last_surg_dir"])
        if s.get("last_emk_dir"):
            self.last_emk_dir = str(s["last_emk_dir"])
        if s.get("last_update_check"):
            self.last_update_check = str(s["last_update_check"])
        if s.get("last_seen_version"):
            self.last_seen_version = str(s["last_seen_version"])
        if isinstance(s.get("summary_paths_by_dept"), dict):
            self.summary_paths_by_dept = {
                str(k): str(v) for k, v in s["summary_paths_by_dept"].items() if v
            }
        if s.get("theme") in UI_THEMES:
            self.theme_var.set(str(s["theme"]))

    def _persist_settings(self):
        # актуальный путь для текущего отделения
        self.summary_paths_by_dept[self.dept_var.get()] = self.summary_path.get()
        save_settings(
            APP_DIR,
            {
                "summary_path": self.summary_path.get(),
                "department": self.dept_var.get(),
                "year": self.year_var.get(),
                "hide_zeros": bool(self.hide_zeros.get()),
                "filter_enabled": bool(self.filter_enabled.get()),
                "start_date": self.start_date_var.get(),
                "end_date": self.end_date_var.get(),
                "write_weeks": bool(self.write_weeks_var.get()),
                "write_form": bool(self.write_form_var.get()),
                "plan_mode": self.plan_mode.get(),
                "last_surg_dir": self.last_surg_dir,
                "last_emk_dir": self.last_emk_dir,
                "last_update_check": self.last_update_check,
                "last_seen_version": self.last_seen_version,
                "summary_paths_by_dept": self.summary_paths_by_dept,
                "theme": self.theme_var.get() if self.theme_var.get() in UI_THEMES else "light",
            },
        )

    def _on_close(self):
        try:
            self._persist_settings()
        except Exception:
            pass
        self.root.destroy()

    def _open_log_file(self):
        path = APP_DIR / "analysis.log"
        if not path.exists():
            messagebox.showinfo("Журнал", "Файл analysis.log ещё не создан")
            return
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", str(path)])
            elif sys.platform.startswith("win"):
                os.startfile(str(path))  # type: ignore
            else:
                subprocess.Popen(["xdg-open", str(path)])
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def _clear_log_file(self):
        if not messagebox.askyesno("Журнал", "Очистить analysis.log?"):
            return
        APP_LOG.clear()
        self.log_text.delete("1.0", tk.END)
        self.log_message("Журнал очищен")

    def export_unclassified(self):
        ops = self.get_view_ops()
        if ops.empty:
            messagebox.showwarning("Нет данных", "Нет операций")
            return
        uncl = ops[ops["Категория"] == "Не классифицировано"]
        if uncl.empty:
            messagebox.showinfo("Экспорт", "Неклассифицированных операций нет")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile="неклассифицировано.csv",
            filetypes=[("CSV", "*.csv"), ("Excel", "*.xlsx")],
        )
        if not path:
            return
        cols = [c for c in ("Дата", "КВС", "Код", "КСГ_название", "КСГ", "Услуга", "КСГ_подсказка") if c in uncl.columns]
        out = uncl[cols].copy()
        try:
            if path.lower().endswith(".xlsx"):
                out.to_excel(path, index=False)
            else:
                out.to_csv(path, index=False, encoding="utf-8-sig")
            self.log_message(f"Экспорт неклассифицированных: {path} ({len(out)} строк)")
            messagebox.showinfo("Готово", f"Сохранено {len(out)} строк:\n{path}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def export_problem_codes(self):
        ops = self.get_view_ops()
        if ops.empty:
            messagebox.showwarning("Нет данных", "Нет операций")
            return
        table = build_problem_codes_table(ops)
        if table.empty:
            messagebox.showinfo("Экспорт", "Неклассифицированных кодов нет")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="проблемные_коды.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
        )
        if not path:
            return
        try:
            if path.lower().endswith(".xlsx"):
                table.to_excel(path, index=False)
            else:
                table.to_csv(path, index=False, encoding="utf-8-sig")
            draft_path = str(Path(path).with_suffix(".yaml"))
            Path(draft_path).write_text(format_config_draft(table), encoding="utf-8")
            self.log_message(
                f"Экспорт проблемных кодов: {path} ({len(table)} кодов), черновик: {draft_path}"
            )
            messagebox.showinfo(
                "Готово",
                f"Таблица: {path}\n"
                f"Кодов: {len(table)}\n\n"
                f"Черновик для config.yaml:\n{draft_path}",
            )
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def _selected_unclassified_defaults(self) -> dict:
        """Подсказки из выделенной строки вкладки «Не классифицировано»."""
        out = {"code": "", "service": "", "name": ""}
        if not hasattr(self, "tree_uncl"):
            return out
        sel = self.tree_uncl.selection()
        if not sel:
            return out
        vals = self.tree_uncl.item(sel[0], "values") or ()
        # Дата, КВС, Код, Название КСГ, КСГ, Услуга
        if len(vals) >= 3:
            out["code"] = str(vals[2] or "").strip()
        if len(vals) >= 4 and vals[3]:
            out["name"] = str(vals[3]).strip()
        if len(vals) >= 6:
            out["service"] = str(vals[5] or "").strip()
        if not out["name"] and out["service"]:
            out["name"] = out["service"][:60]
        return out

    def _selected_dispute_defaults(self) -> dict:
        out = {"category": "", "candidates": [], "store_index": None}
        if not hasattr(self, "tree_dispute"):
            return out
        sel = self.tree_dispute.selection()
        if not sel:
            return out
        iid = sel[0]
        try:
            out["store_index"] = int(iid)
        except (TypeError, ValueError):
            out["store_index"] = iid
        vals = self.tree_dispute.item(sel[0], "values") or ()
        # Дата, КВС, Код, Услуга, Категория, Кандидаты
        if len(vals) >= 5:
            out["category"] = str(vals[4] or "").strip()
        if len(vals) >= 6 and vals[5]:
            out["candidates"] = [c.strip() for c in str(vals[5]).split("|") if c.strip()]
        return out

    def assign_disputed_category_dialog(self):
        """Назначить категорию выбранной спорной операции вручную."""
        if self.store.ops.empty:
            messagebox.showinfo("Спорные", "Нет данных — загрузите опержурнал")
            return
        defaults = self._selected_dispute_defaults()
        if defaults["store_index"] is None:
            messagebox.showinfo("Спорные", "Выберите строку в таблице")
            return
        idx = defaults["store_index"]
        if idx not in self.store.ops.index:
            messagebox.showwarning("Спорные", "Строка не найдена в накопителе — обновите отчёт")
            return

        rows = list((self.summary_cfg.get("category_rows") or {}).keys())
        # кандидаты сверху
        ordered = []
        for c in defaults["candidates"]:
            if c and c not in ordered:
                ordered.append(c)
        for c in rows:
            if c not in ordered:
                ordered.append(c)
        if defaults["category"] and defaults["category"] not in ordered:
            ordered.insert(0, defaults["category"])

        top = tk.Toplevel(self.root)
        top.title("Назначить категорию")
        top.transient(self.root)
        cat_var = StringVar(value=defaults["candidates"][0] if defaults["candidates"] else (defaults["category"] or (ordered[0] if ordered else "")))
        tk.Label(top, text="Категория для выбранной операции:").pack(anchor="w", padx=10, pady=(10, 2))
        ttk.Combobox(top, textvariable=cat_var, values=ordered, width=48, state="readonly").pack(
            fill=tk.X, padx=10, pady=4
        )
        tk.Label(
            top,
            text="Назначение сохранится как ручное и не сбросится при правке ключей.",
            fg="#555",
            wraplength=420,
            justify=tk.LEFT,
        ).pack(anchor="w", padx=10, pady=4)

        def apply():
            name = cat_var.get().strip()
            if not name:
                messagebox.showwarning("Категория", "Выберите категорию", parent=top)
                return
            meta = lookup_category_meta(self._surgery_categories(), name)
            ops = self.store.ops
            if "Ручная_категория" not in ops.columns:
                ops["Ручная_категория"] = False
            if "Спор_ключей" not in ops.columns:
                ops["Спор_ключей"] = False
            if "Спорные_категории" not in ops.columns:
                ops["Спорные_категории"] = ""
            ops.at[idx, "Категория"] = name
            ops.at[idx, "Спор_ключей"] = False
            ops.at[idx, "Спорные_категории"] = ""
            ops.at[idx, "Ручная_категория"] = True
            if meta:
                ops.at[idx, "Группа"] = meta.get("group", ops.at[idx, "Группа"] if "Группа" in ops.columns else "")
                ops.at[idx, "Строка_4001"] = meta.get("line", "")
                ops.at[idx, "Гистология"] = bool(meta.get("histology", False))
            self.log_message(f"Спорные: строка {idx} → «{name}» (вручную)")
            top.destroy()
            self.run_analysis()

        bf = tk.Frame(top)
        bf.pack(pady=10)
        _btn(bf, "Назначить", apply, side=tk.LEFT, padx=4)
        _btn(bf, "Отмена", top.destroy, side=tk.LEFT, padx=4)

    def edit_keywords_dialog_from_dispute(self):
        defaults = self._selected_dispute_defaults()
        prefer = ""
        if defaults["candidates"]:
            prefer = defaults["candidates"][0]
        elif defaults["category"]:
            prefer = defaults["category"]
        self.edit_keywords_dialog(preselect=prefer)

    def edit_keywords_dialog(self, preselect: str = ""):
        """Правка name_keywords у существующей категории + переклассификация."""
        cats = self._surgery_categories()
        names = [str(c.get("category") or "").strip() for c in cats if c.get("category")]
        if not names:
            messagebox.showinfo("Ключи", "В config нет категорий")
            return

        top = tk.Toplevel(self.root)
        top.title("Ключевые слова категории")
        top.transient(self.root)
        top.resizable(True, True)
        pad = {"padx": 10, "pady": 3}

        name_var = StringVar(value=preselect if preselect in names else names[0])
        kw_var = StringVar()

        def load_kw(*_a):
            meta = lookup_category_meta(cats, name_var.get())
            kws = (meta or {}).get("name_keywords") or []
            kw_var.set(", ".join(str(x) for x in kws))

        tk.Label(top, text="Категория:").pack(anchor="w", **pad)
        cb = ttk.Combobox(top, textvariable=name_var, values=names, width=48, state="readonly")
        cb.pack(fill=tk.X, **pad)
        cb.bind("<<ComboboxSelected>>", load_kw)
        tk.Label(top, text="Ключевые слова (через запятую):").pack(anchor="w", **pad)
        tk.Entry(top, textvariable=kw_var, width=56).pack(fill=tk.X, **pad)
        tk.Label(
            top,
            text=(
                "Каждый ключ после запятой — отдельно («или»). "
                "Словосочетание целиком — одним ключом без запятой. "
                "После сохранения спорные/неклассифицированные строки без ручной метки пересчитаются."
            ),
            fg="#555",
            wraplength=520,
            justify=tk.LEFT,
        ).pack(anchor="w", padx=10, pady=(0, 4))
        load_kw()

        def save():
            try:
                kws = [c.strip() for c in kw_var.get().replace(";", ",").split(",") if c.strip()]
                cfg, normed = update_category_keywords_file(
                    APP_DIR / "config.yaml",
                    name_var.get(),
                    kws,
                    config=self.config,
                    summary_key=self.summary_key,
                )
                self.config = cfg
                self._sync_dept_context()
                if not self.store.ops.empty:
                    self.store.ops = reclassify_ops_by_keywords(
                        self.store.ops, self._surgery_categories()
                    )
                self.log_message(
                    f"Ключи «{name_var.get()}»: {normed or '—'}; переклассификация накопителя"
                )
                top.destroy()
                if not self.store.ops.empty:
                    self.run_analysis()
                else:
                    messagebox.showinfo("Ключи", f"Сохранено: {', '.join(normed) or 'пусто'}")
            except CategoryRegistryError as e:
                messagebox.showwarning("Проверка", str(e), parent=top)
            except Exception as e:
                self.log_message(traceback.format_exc(), level="ERROR")
                messagebox.showerror("Ошибка", str(e), parent=top)

        bf = tk.Frame(top)
        bf.pack(pady=10)
        _btn(bf, "Сохранить", save, side=tk.LEFT, padx=4)
        _btn(bf, "Отмена", top.destroy, side=tk.LEFT, padx=4)

    def inventory_department_dialog(self):
        """Инвентаризация кодов операций из отчёта отделения."""
        keys = [k for k in DEPT_REPORT_SOURCES if k in (self.config.get("summaries") or {})]
        if not keys:
            keys = list(DEPT_REPORT_SOURCES.keys())
        labels = [
            f"{DEPT_REPORT_SOURCES[k]['label']} ({DEPT_REPORT_SOURCES[k]['department']})" for k in keys
        ]
        top = tk.Toplevel(self.root)
        top.title("Инвентаризация отделения")
        top.transient(self.root)
        tk.Label(
            top,
            text="Сформировать таблицу уникальных кодов услуг с сопоставлением КСГ\n"
            "из папки «Отчеты других отделений».",
            justify=tk.LEFT,
            wraplength=480,
        ).pack(padx=12, pady=8, anchor="w")
        dept_var = StringVar(value=labels[0] if labels else "")
        ttk.Combobox(top, textvariable=dept_var, values=labels, width=52, state="readonly").pack(
            padx=12, pady=4, anchor="w"
        )

        def run():
            if not labels:
                return
            idx = labels.index(dept_var.get()) if dept_var.get() in labels else 0
            sk = keys[idx]
            reports = APP_DIR / "Отчеты других отделений"
            try:
                table, cats, _meta = inventory_from_source(sk, reports)
                default_out = APP_DIR / f"инвентарь_{sk}.xlsx"
                out = filedialog.asksaveasfilename(
                    parent=top,
                    title="Сохранить инвентаризацию",
                    initialdir=str(APP_DIR),
                    initialfile=default_out.name,
                    defaultextension=".xlsx",
                    filetypes=[("Excel", "*.xlsx")],
                )
                if not out:
                    return
                export_inventory_excel(table, out, summary_key=sk)
                self.log_message(
                    f"Инвентаризация {sk}: {len(table)} кодов, {len(cats)} категорий → {out}"
                )
                messagebox.showinfo(
                    "Готово",
                    f"Кодов: {len(table)}\nКатегорий (1 код = 1 строка): {len(cats)}\n\n{out}",
                    parent=top,
                )
                top.destroy()
            except Exception as e:
                self.log_message(traceback.format_exc(), level="ERROR")
                messagebox.showerror("Ошибка", str(e), parent=top)

        bf = tk.Frame(top)
        bf.pack(pady=10)
        _btn(bf, "Сформировать…", run, side=tk.LEFT, padx=4)
        _btn(bf, "Отмена", top.destroy, side=tk.LEFT, padx=4)

    def create_dept_summary_dialog(self):
        """Создать Excel-сводную для текущего или выбранного отделения."""
        sk = self.summary_key
        summary_cfg = get_summary_cfg(self.config, summary_key=sk)
        cats = summary_cfg.get("category_rows") or {}
        if not cats:
            messagebox.showwarning(
                "Сводная",
                f"Для отделения «{self.dept_var.get()}» нет category_rows в config.\n"
                "Сначала выполните scripts/generate_dept_config.py или инвентаризацию.",
            )
            return
        try:
            year = int(self.year_var.get())
        except (TypeError, ValueError):
            year = int(summary_cfg.get("year") or 2026)
        default_name = default_summary_filename(self.config, sk, year)
        out = filedialog.asksaveasfilename(
            title="Создать сводную для отделения",
            initialdir=str(APP_DIR),
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not out:
            return
        try:
            create_from_summary_cfg(out, summary_cfg, self.dept_var.get())
            self.summary_path.set(out)
            self.summary_paths_by_dept[self.dept_var.get()] = out
            self._persist_settings()
            self.log_message(f"Создана сводная: {out} ({len(cats)} операций)")
            messagebox.showinfo(
                "Готово",
                f"Создан файл:\n{out}\n\nСтрок операций: {len(cats)}\nФорма 4001 не включена.",
            )
        except Exception as e:
            self.log_message(traceback.format_exc(), level="ERROR")
            messagebox.showerror("Ошибка", str(e))

    def classify_kinds_from_emk_dialog(self):
        """По загруженной ЭМК: plan/emergency для категорий; спорные — на ручную сверку."""
        if self.df_emk is None:
            messagebox.showwarning(
                "ЭМК",
                "Сначала загрузите отчёт «Заполнение ЭМК в стационаре».",
            )
            return
        ops = self.get_view_ops() if hasattr(self, "get_view_ops") else self.store.ops
        if ops is None or ops.empty:
            messagebox.showwarning("Нет данных", "Загрузите опержурнал текущего отделения")
            return
        if ops["Тип_ЭМК"].astype(str).str.strip().eq("").all():
            messagebox.showwarning(
                "Нет связи",
                "КВС журнала и ЭМК не пересекаются — сверка план/экстр невозможна.",
            )
            return

        names = list((self.summary_cfg.get("category_rows") or {}).keys())
        kind = classify_categories_by_emk(ops, category_names=names)
        report = format_kind_report(kind)
        self.log_message(report)

        n_disp = len(kind.get("disputed") or [])
        msg = (
            f"Плановые: {len(kind['plan'])}\n"
            f"Экстренные: {len(kind['emergency'])}\n"
            f"Без ЭМК (оставлены план): {len(kind['no_emk'])}\n"
            f"Спорные (план и экстр): {n_disp}\n\n"
            "Записать plan/emergency в config.yaml?"
        )
        if n_disp:
            msg += "\n\nСпорные будут сохранены в Excel для ручной сверки."
        if not messagebox.askyesno("План/экстр по ЭМК", msg):
            return
        try:
            updated = apply_kind_to_summary_cfg(dict(self.summary_cfg), kind)
            set_summary_cfg(self.config, self.summary_key, updated)
            save_config(self.config, APP_DIR / "config.yaml")
            self._sync_dept_context()

            if n_disp:
                out = filedialog.asksaveasfilename(
                    title="Спорные план/экстр",
                    initialdir=str(APP_DIR),
                    initialfile=f"спорные_план_экстр_{self.summary_key}.xlsx",
                    defaultextension=".xlsx",
                    filetypes=[("Excel", "*.xlsx")],
                )
                if out:
                    disputed_to_dataframe(kind).to_excel(out, index=False)
                    self.log_message(f"Спорные план/экстр: {out}")

            # обновить цвета/формулы: предложить пересоздать сводную
            recreate = messagebox.askyesno(
                "Сводная",
                "Config обновлён.\n\nПересоздать Excel-сводную с новыми plan/экстр "
                "(цвета и формулы итогов)?\n"
                "Текущие числа недель при этом сбросятся — потом снова «Записать в Excel…».",
            )
            if recreate:
                path = self.summary_path.get().strip()
                if not path:
                    path = str(
                        APP_DIR
                        / default_summary_filename(
                            self.config, self.summary_key, int(self.year_var.get() or 2026)
                        )
                    )
                create_from_summary_cfg(path, self.summary_cfg, self.dept_var.get())
                self.summary_path.set(path)
                self.summary_paths_by_dept[self.dept_var.get()] = path
                self._persist_settings()
                self.log_message(f"Сводная пересоздана: {path}")

            if not self.store.ops.empty:
                self.run_analysis()
            messagebox.showinfo("Готово", report)
        except Exception as e:
            self.log_message(traceback.format_exc(), level="ERROR")
            messagebox.showerror("Ошибка", str(e))

    def add_category_dialog(self):
        """Мастер: добавить неизвестную операцию в config (± Excel)."""
        defaults = self._selected_unclassified_defaults()
        cat = get_catalog()
        hint = cat.hint_for(defaults["code"]) if defaults["code"] else ""
        if defaults["code"] and not defaults["name"]:
            info = cat.lookup(defaults["code"])
            if info and info.get("name"):
                defaults["name"] = str(info["name"])[:60]

        top = tk.Toplevel(self.root)
        top.title("Добавить операцию в отчёт")
        top.transient(self.root)
        top.resizable(True, True)

        name_var = StringVar(value=defaults.get("name") or "")
        codes_var = StringVar(value=defaults.get("code") or "")
        kw_var = StringVar(value=", ".join(suggest_keywords_from_name(name_var.get())))
        kind_var = StringVar(value="plan")
        line_var = StringVar(value="6")
        hist_var = BooleanVar(value=False)
        endo_var = BooleanVar(value=False)
        anchor_var = StringVar(value=default_anchor_category(self.config, self.summary_key))
        ksg_query = StringVar(value=defaults.get("code") or defaults.get("name") or "")
        ksg_hint = StringVar(value=hint or "Поиск по KSGoperacii.csv или заполните поля вручную")

        pad = {"padx": 10, "pady": 3}

        tk.Label(top, text="Название в отчёте (как в сводной):").pack(anchor="w", **pad)
        tk.Entry(top, textvariable=name_var, width=56).pack(fill=tk.X, **pad)

        tk.Label(top, text="Код(ы) через запятую:").pack(anchor="w", **pad)
        tk.Entry(top, textvariable=codes_var, width=56).pack(fill=tk.X, **pad)

        tk.Label(top, text="Ключевые слова (через запятую):").pack(anchor="w", **pad)
        tk.Entry(top, textvariable=kw_var, width=56).pack(fill=tk.X, **pad)
        tk.Label(
            top,
            text=(
                "Как ищутся: каждый ключ после запятой проверяется отдельно как подстрока в названии услуги "
                "(без учёта регистра). Достаточно совпадения хотя бы одного ключа («или», не «и»). "
                "Несколько совпадений усиливают выбор этой категории. "
                "Словосочетание целиком — пишите одним ключом без запятой "
                "(например: «резекция гортани»)."
            ),
            fg="#555",
            wraplength=520,
            justify=tk.LEFT,
        ).pack(anchor="w", padx=10, pady=(0, 4))

        row_kind = tk.Frame(top)
        row_kind.pack(fill=tk.X, **pad)
        tk.Label(row_kind, text="Тип:").pack(side=tk.LEFT)
        tk.Radiobutton(row_kind, text="Плановая", variable=kind_var, value="plan").pack(side=tk.LEFT, padx=6)
        tk.Radiobutton(row_kind, text="Экстренная", variable=kind_var, value="emergency").pack(side=tk.LEFT)

        row_line = tk.Frame(top)
        row_line.pack(fill=tk.X, **pad)
        tk.Label(row_line, text="Строка формы 4001:").pack(side=tk.LEFT)
        ttk.Combobox(
            row_line,
            textvariable=line_var,
            values=["5.1", "5.2", "6", "6.1", "17"],
            width=8,
            state="readonly",
        ).pack(side=tk.LEFT, padx=6)
        tk.Checkbutton(row_line, text="Гистология", variable=hist_var).pack(side=tk.LEFT, padx=8)
        tk.Checkbutton(row_line, text="Эндоскопия", variable=endo_var).pack(side=tk.LEFT)

        row_anchor = tk.Frame(top)
        row_anchor.pack(fill=tk.X, **pad)
        tk.Label(row_anchor, text="Вставить после:").pack(side=tk.LEFT)
        anchors = list((self.summary_cfg.get("category_rows") or {}).keys())
        ttk.Combobox(row_anchor, textvariable=anchor_var, values=anchors, width=40, state="readonly").pack(
            side=tk.LEFT, padx=6
        )

        ksg_fr = tk.LabelFrame(top, text="Рубрикатор КСГ", padx=6, pady=4)
        ksg_fr.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)
        qrow = tk.Frame(ksg_fr)
        qrow.pack(fill=tk.X)
        ksg_entry = tk.Entry(qrow, textvariable=ksg_query, width=40)
        ksg_entry.pack(side=tk.LEFT, padx=(0, 4))
        ksg_list = tk.Listbox(ksg_fr, height=6)

        def do_search(_event=None):
            ksg_list.delete(0, tk.END)
            hits = cat.search(ksg_query.get(), limit=40)
            if not hits:
                ksg_hint.set("Ничего не найдено — заполните поля вручную")
                return
            ksg_hint.set(f"Найдено: {len(hits)}. Выберите строку (Enter / двойной клик) — подставить")
            for h in hits:
                ksg_list.insert(tk.END, f"{h['code']} | {h['name']}")

        def apply_hit(_event=None):
            sel = ksg_list.curselection()
            if not sel:
                return
            text = ksg_list.get(sel[0])
            code = text.split("|", 1)[0].strip()
            info = cat.lookup(code)
            if not info:
                return
            codes_var.set(code)
            # всегда обновляем название при выборе из рубрикатора
            name_var.set(str(info.get("name") or "")[:60])
            kw_var.set(", ".join(suggest_keywords_from_name(info.get("name") or name_var.get())))
            ksg_hint.set(cat.hint_for(code))

        _btn(qrow, "Поиск", do_search, side=tk.LEFT, padx=2)
        ksg_entry.bind("<Return>", do_search)
        ksg_list.pack(fill=tk.BOTH, expand=True, pady=4)
        ksg_list.bind("<Double-Button-1>", apply_hit)
        ksg_list.bind("<Return>", apply_hit)
        tk.Label(ksg_fr, textvariable=ksg_hint, fg="#555", wraplength=520, justify=tk.LEFT).pack(anchor="w")
        # автопоиск при открытии не запускаем — только по кнопке / Enter

        def build_spec() -> CategorySpec:
            codes = [c.strip() for c in codes_var.get().replace(";", ",").split(",") if c.strip()]
            kws = [c.strip() for c in kw_var.get().replace(";", ",").split(",") if c.strip()]
            return CategorySpec(
                name=name_var.get().strip(),
                codes=codes,
                name_keywords=kws,
                kind=kind_var.get(),
                form_line=line_var.get(),
                histology=bool(hist_var.get()),
                endoscopic=bool(endo_var.get()),
                anchor_category=anchor_var.get(),
            )

        def reload_after():
            cfg = self.load_config()
            if cfg is None:
                return
            self.config = cfg
            self._sync_dept_context()
            if not self.store.ops.empty:
                self.run_analysis()

        def do_config_only():
            try:
                _cfg, result = register_category(
                    APP_DIR / "config.yaml", build_spec(), config=self.config, summary_key=self.summary_key
                )
                self.config = _cfg
                self._sync_dept_context()
                msg = f"В программе: «{result.name}» → строка Excel {result.excel_row}"
                if result.warnings:
                    msg += "\n" + "\n".join(result.warnings)
                self.log_message(msg)
                reload_after()
                messagebox.showinfo(
                    "Готово",
                    msg + "\n\nДобавьте строку в Excel или нажмите «В программе + Excel».",
                    parent=top,
                )
                top.destroy()
            except CategoryRegistryError as e:
                messagebox.showwarning("Проверка", str(e), parent=top)
            except Exception as e:
                self.log_message(traceback.format_exc(), level="ERROR")
                messagebox.showerror("Ошибка", str(e), parent=top)

        def do_config_and_excel():
            path = self.summary_path.get().strip()
            if not path or not os.path.exists(path):
                messagebox.showerror("Сводная", f"Файл сводной не найден:\n{path}", parent=top)
                return
            if excel_file_locked(path):
                messagebox.showerror("Файл занят", "Закройте сводную в Excel и повторите.", parent=top)
                return
            try:
                spec = build_spec()
                _cfg, result = register_category(
                    APP_DIR / "config.yaml", spec, config=self.config, summary_key=self.summary_key
                )
                self.config = _cfg
                self._sync_dept_context()
                sheets = {
                    int(k): v for k, v in (self.summary_cfg.get("sheet_names") or {}).items()
                }
                anchor_row = find_anchor_row(self.summary_cfg.get("category_rows") or {}, spec.anchor_category)
                xrep = add_category_row_to_summary(
                    path,
                    category_name=result.name,
                    excel_row=result.excel_row,
                    form_line=spec.form_line,
                    sheet_names=sheets,
                    form_cfg=self.summary_cfg.get("form_4001") or {},
                    kind=spec.kind,
                    histology=bool(spec.histology),
                    endoscopic=bool(spec.endoscopic),
                    anchor_row=anchor_row,
                    backup=True,
                    backup_keep=int(self.summary_cfg.get("backup_keep", 20)),
                )
                blank_delta = int(xrep.get("blank_delta") or 0)
                if blank_delta:
                    shift_totals_rows_by_delta(self.config, blank_delta, summary_key=self.summary_key)
                    save_config(self.config, APP_DIR / "config.yaml")
                    self._sync_dept_context()
                msg = (
                    f"«{result.name}» → вставлена строка {result.excel_row} "
                    f"({ 'экстр.' if spec.kind == 'emergency' else 'план' })\n"
                    f"Месяцы: {len(xrep.get('sheets') or [])}, "
                    f"сводные: {', '.join(xrep.get('overview') or []) or '—'}\n"
                    f"Формулы 4001: N+{xrep['formulas']['n']}, R+{xrep['formulas']['r']}; "
                    f"итоги+{xrep['formulas']['total']}, план/экстр+{xrep['formulas']['kind']}; "
                    f"графики+{xrep.get('charts', 0)}"
                )
                if xrep.get("backup"):
                    msg += f"\nБэкап: {xrep['backup']}"
                self.log_message(msg)
                reload_after()
                messagebox.showinfo(
                    "Готово",
                    msg + "\n\nДальше: «Обновить превью» и при необходимости «Записать в Excel…».",
                    parent=top,
                )
                top.destroy()
            except CategoryRegistryError as e:
                messagebox.showwarning("Проверка", str(e), parent=top)
            except Exception as e:
                self.log_message(traceback.format_exc(), level="ERROR")
                messagebox.showerror("Ошибка", str(e), parent=top)

        bf = tk.Frame(top)
        bf.pack(pady=10)
        _btn(bf, "Только в программе", do_config_only, side=tk.LEFT, padx=4)
        _btn(bf, "В программе + Excel", do_config_and_excel, side=tk.LEFT, padx=4)
        _btn(bf, "Отмена", top.destroy, side=tk.LEFT, padx=4)

    def delete_category_dialog(self):
        """Удалить категорию операции из программы и (опционально) из Excel."""
        rows = dict(self.summary_cfg.get("category_rows") or {})
        if not rows:
            messagebox.showinfo("Удаление", "В config нет категорий")
            return
        # сортировка по номеру строки
        names = [k for k, _ in sorted(rows.items(), key=lambda kv: int(kv[1]))]

        top = tk.Toplevel(self.root)
        top.title("Удалить операцию из отчёта")
        top.transient(self.root)
        top.resizable(True, True)
        tk.Label(
            top,
            text="Выберите операцию. Удалится из программы; при выборе «В программе + Excel» — "
            "и строка из сводной на всех листах.",
            wraplength=520,
            justify=tk.LEFT,
        ).pack(padx=12, pady=8, anchor="w")

        name_var = StringVar(value=names[-1] if names else "")
        ttk.Combobox(top, textvariable=name_var, values=names, width=48, state="readonly").pack(
            padx=12, pady=4, anchor="w"
        )
        row_info = StringVar(value="")

        def refresh_info(*_a):
            n = name_var.get()
            r = rows.get(n)
            row_info.set(f"Строка Excel: {r}" if r is not None else "")

        name_var.trace_add("write", refresh_info)
        refresh_info()
        tk.Label(top, textvariable=row_info, fg="#555").pack(padx=12, anchor="w")

        def reload_after():
            cfg = self.load_config()
            if cfg is None:
                return
            self.config = cfg
            self._sync_dept_context()
            if not self.store.ops.empty:
                self.run_analysis()

        def do_program_only():
            name = name_var.get()
            if not name:
                return
            if not messagebox.askyesno(
                "Удаление",
                f"Удалить «{name}» из программы (config)?\nСтрока в Excel не трогается.",
                parent=top,
            ):
                return
            try:
                _cfg, result = unregister_category(
                    APP_DIR / "config.yaml", name, config=self.config, summary_key=self.summary_key
                )
                self.config = _cfg
                self._sync_dept_context()
                self.log_message(
                    f"Удалено из программы: «{result.name}» (была строка {result.excel_row})"
                )
                reload_after()
                messagebox.showinfo("Готово", f"Удалено: «{result.name}»", parent=top)
                top.destroy()
            except CategoryRegistryError as e:
                messagebox.showwarning("Проверка", str(e), parent=top)
            except Exception as e:
                self.log_message(traceback.format_exc(), level="ERROR")
                messagebox.showerror("Ошибка", str(e), parent=top)

        def do_program_and_excel():
            name = name_var.get()
            if not name:
                return
            path = self.summary_path.get().strip()
            if not path or not os.path.exists(path):
                messagebox.showerror("Сводная", f"Файл сводной не найден:\n{path}", parent=top)
                return
            if excel_file_locked(path):
                messagebox.showerror("Файл занят", "Закройте сводную в Excel и повторите.", parent=top)
                return
            excel_row = int(rows.get(name) or 0)
            if not excel_row:
                messagebox.showwarning("Удаление", "Неизвестна строка Excel", parent=top)
                return
            if not messagebox.askyesno(
                "Удаление",
                f"Удалить «{name}» из программы и строку {excel_row} из Excel "
                f"(все месяцы, ОБЩАЯ, графики)?",
                parent=top,
            ):
                return
            try:
                _cfg, result = unregister_category(
                    APP_DIR / "config.yaml", name, config=self.config, summary_key=self.summary_key
                )
                self.config = _cfg
                self._sync_dept_context()
                sheets = {
                    int(k): v for k, v in (self.summary_cfg.get("sheet_names") or {}).items()
                }
                xrep = delete_category_row_from_summary(
                    path,
                    excel_row=result.excel_row,
                    sheet_names=sheets,
                    backup=True,
                    backup_keep=int(self.summary_cfg.get("backup_keep", 20)),
                )
                blank_delta = int(xrep.get("blank_delta") or 0)
                if blank_delta:
                    shift_totals_rows_by_delta(self.config, blank_delta, summary_key=self.summary_key)
                    save_config(self.config, APP_DIR / "config.yaml")
                    self._sync_dept_context()
                msg = (
                    f"Удалено «{result.name}» (строка {result.excel_row})\n"
                    f"Листов: {len(xrep.get('sheets') or [])}, "
                    f"сводные: {', '.join(xrep.get('overview') or []) or '—'}"
                )
                if xrep.get("backup"):
                    msg += f"\nБэкап: {xrep['backup']}"
                self.log_message(msg)
                reload_after()
                messagebox.showinfo("Готово", msg, parent=top)
                top.destroy()
            except CategoryRegistryError as e:
                messagebox.showwarning("Проверка", str(e), parent=top)
            except Exception as e:
                self.log_message(traceback.format_exc(), level="ERROR")
                messagebox.showerror("Ошибка", str(e), parent=top)

        bf = tk.Frame(top)
        bf.pack(pady=10)
        _btn(bf, "Только в программе", do_program_only, side=tk.LEFT, padx=4)
        _btn(bf, "В программе + Excel", do_program_and_excel, side=tk.LEFT, padx=4)
        _btn(bf, "Отмена", top.destroy, side=tk.LEFT, padx=4)

    def export_emk_mismatches(self):
        result = self.last_emk_compare
        if not result or not result.get("mismatches"):
            ops = self.get_view_ops()
            if ops.empty or self.df_emk is None:
                messagebox.showinfo("Экспорт", "Сначала выполните сверку ЭМК")
                return
            result = compare_plan_emergency(ops, self.summary_cfg, department=self.dept_var.get())
            self._fill_emk_tree(result, select_tab=False)
        mismatches = result.get("mismatches") or []
        if not mismatches:
            messagebox.showinfo("Экспорт", "Расхождений нет")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="расхождения_эмк.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
        )
        if not path:
            return
        rows = []
        for m in mismatches:
            dt = m.get("Дата")
            dt_s = dt.strftime("%d.%m.%Y") if hasattr(dt, "strftime") else str(dt or "")
            rows.append(
                {
                    "Дата": dt_s,
                    "КВС": m.get("КВС"),
                    "Категория": m.get("Категория"),
                    "Код": m.get("Код"),
                    "Шаблон": m.get("Шаблон"),
                    "ЭМК": m.get("ЭМК"),
                    "Диагноз": m.get("Диагноз"),
                    "Услуга": m.get("Услуга"),
                }
            )
        out = pd.DataFrame(rows)
        try:
            if path.lower().endswith(".xlsx"):
                out.to_excel(path, index=False)
            else:
                out.to_csv(path, index=False, encoding="utf-8-sig")
            self.log_message(f"Экспорт расхождений ЭМК: {path} ({len(out)} строк)")
            messagebox.showinfo("Готово", f"Сохранено {len(out)} строк:\n{path}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    try:
        app = DesktopApp(root)
    except Exception:
        traceback.print_exc()
        messagebox.showerror("Старт", traceback.format_exc())
        root.destroy()
        raise
    root.mainloop()
